#!/usr/bin/env python3
"""
Telegram Alert Backtester - Proper Daily Simulation

Simulates day-by-day trading with:
- Capital tracking (5% position size from available capital)
- Daily SL checks against KC Lower/Middle
- Proper entry/exit with transaction history
- Margin support (up to 100% = 2x leverage)

Data Source: Daily/data/audit_vsr.db (telegram_alerts table)
"""

import os
import sys
import sqlite3
import logging
import configparser
from datetime import datetime, timedelta
from pathlib import Path
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from collections import defaultdict
import time

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pytz

sys.path.insert(0, str(Path(__file__).parent.parent))
sys.path.insert(0, str(Path(__file__).parent))

from kiteconnect import KiteConnect
from core.td_indicators import TDIndicatorCalculator, TDState

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def safe_float(val, default=0.0):
    """Convert value to float safely, handling bytes"""
    if val is None:
        return default
    if isinstance(val, bytes):
        try:
            return float(val.decode('utf-8'))
        except:
            return default
    try:
        return float(val)
    except:
        return default


@dataclass
class OpenPosition:
    """Represents an open position"""
    ticker: str
    entry_date: str
    entry_price: float
    quantity: int  # Current remaining quantity
    position_value: float  # Current remaining position value
    stop_loss: float  # Current stop loss level
    kc_lower: float
    kc_middle: float
    overnight_charges: float = 0.0  # Accumulated overnight charges
    tdst_support: float = 0.0  # TDST Support level (dynamic, updated as new setups form)
    # Tranche exit tracking for TD strategy (3-tier: 30%, 45%, 25%)
    original_quantity: int = 0  # Original full quantity at entry
    original_value: float = 0.0  # Original full position value
    # Tranche state: 1=Full(100%), 2=De-risked(70%), 3=Runner(25%), 4=Flat(0%)
    position_state: int = 1
    setup_lowest_low: float = 0.0  # Lowest low of bars 1-9 for Setup validity
    # Tranche 1 (30%) tracking
    tranche1_exited: bool = False
    tranche1_exit_date: str = ""
    tranche1_exit_price: float = 0.0
    tranche1_pnl: float = 0.0
    tranche1_reason: str = ""
    # Tranche 2 (45%) tracking
    tranche2_exited: bool = False
    tranche2_exit_date: str = ""
    tranche2_exit_price: float = 0.0
    tranche2_pnl: float = 0.0
    tranche2_reason: str = ""
    # TDST violation flag (prevents re-entry after TDST breach)
    tdst_violated: bool = False


@dataclass
class ClosedTrade:
    """Represents a closed trade"""
    ticker: str
    entry_date: str
    entry_price: float
    exit_date: str
    exit_price: float
    quantity: int
    position_value: float
    pnl: float
    pnl_pct: float
    exit_reason: str
    days_held: int
    kc_lower_at_entry: float
    kc_middle_at_entry: float
    overnight_charges: float = 0.0  # Total overnight charges incurred
    tranche: str = ""  # "TRANCHE_1" or "TRANCHE_2" for TD strategy partial exits


class TelegramAlertBacktester:
    """
    Day-by-day backtester for Telegram alert strategies
    """

    def __init__(self, initial_capital: float = 10000000, position_size_pct: float = 5.0,
                 charges_per_leg_pct: float = 0.1, overnight_charge_pct: float = 0.2,
                 lookback_days: int = 60, margin_pct: float = 100.0):
        """
        Initialize backtester

        Args:
            initial_capital: Starting capital (default 1 Crore)
            position_size_pct: Position size as % of portfolio (default 5%)
            charges_per_leg_pct: Trading charges per leg (default 0.1%, round trip 0.2%)
            overnight_charge_pct: Overnight holding charge % (default 0.2%)
            lookback_days: Days to look back for alerts
            margin_pct: Margin allowed as % of capital (default 100% = 2x leverage)
        """
        self.initial_capital = initial_capital
        self.position_size_pct = position_size_pct
        self.charges_per_leg_pct = charges_per_leg_pct
        self.overnight_charge_pct = overnight_charge_pct
        self.lookback_days = lookback_days
        self.margin_pct = margin_pct
        self.max_capital = initial_capital * (1 + margin_pct / 100)

        self.base_dir = Path("/Users/maverick/PycharmProjects/India-TS/Daily")
        self.db_path = self.base_dir / "data" / "audit_vsr.db"
        self.output_dir = self.base_dir / "analysis" / "Efficiency"

        self.ist = pytz.timezone('Asia/Kolkata')
        self.kite = None
        self.instrument_map = {}

        # Historical data cache
        self.price_cache = {}
        self.td_cache = {}  # Cache for TD indicator calculations

        # Rate limiting
        self.last_api_call = 0
        self.rate_limit = 3

        # TD Indicator Calculator
        self.td_calculator = TDIndicatorCalculator()

        self._init_kite()

    def _init_kite(self):
        """Initialize Zerodha connection"""
        try:
            config = configparser.ConfigParser()
            config.read(self.base_dir / 'config.ini')

            if 'API_CREDENTIALS_Sai' in config:
                api_key = config['API_CREDENTIALS_Sai']['api_key']
                access_token = config['API_CREDENTIALS_Sai']['access_token']
            else:
                api_key = config['DEFAULT']['api_key']
                access_token = config['DEFAULT']['access_token']

            self.kite = KiteConnect(api_key=api_key)
            self.kite.set_access_token(access_token)

            instruments = self.kite.instruments("NSE")
            self.instrument_map = {i['tradingsymbol']: i['instrument_token'] for i in instruments}

            logger.info(f"Kite connected, loaded {len(self.instrument_map)} instruments")
        except Exception as e:
            logger.error(f"Failed to init Kite: {e}")

    def _throttle(self):
        """Rate limit API calls"""
        elapsed = time.time() - self.last_api_call
        if elapsed < 1.0 / self.rate_limit:
            time.sleep(1.0 / self.rate_limit - elapsed)
        self.last_api_call = time.time()

    def get_telegram_alerts_by_date(self) -> Dict[str, List[Dict]]:
        """Get Telegram alerts grouped by date"""
        if not self.db_path.exists():
            logger.error(f"Database not found: {self.db_path}")
            return {}

        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row

        start_date = (datetime.now() - timedelta(days=self.lookback_days)).strftime('%Y-%m-%d')

        query = """
            SELECT
                ticker,
                timestamp as alert_time,
                current_price as entry_price,
                score,
                momentum,
                liquidity_grade
            FROM telegram_alerts
            WHERE DATE(timestamp) >= ?
            ORDER BY timestamp
        """

        cursor = conn.execute(query, (start_date,))
        rows = cursor.fetchall()

        # Group by date
        alerts_by_date = defaultdict(list)
        seen_tickers_by_date = defaultdict(set)

        for row in rows:
            entry_price = safe_float(row['entry_price'], 0.0)
            if entry_price <= 0:
                continue

            try:
                alert_dt = datetime.fromisoformat(row['alert_time'].replace('Z', '+00:00'))
                date_str = alert_dt.strftime('%Y-%m-%d')
            except:
                date_str = row['alert_time'][:10]

            ticker = row['ticker']

            # Only take first alert per ticker per day
            if ticker not in seen_tickers_by_date[date_str]:
                seen_tickers_by_date[date_str].add(ticker)
                alerts_by_date[date_str].append({
                    'ticker': ticker,
                    'alert_time': row['alert_time'],
                    'entry_price': float(entry_price),
                    'score': row['score'],
                    'momentum': row['momentum']
                })

        conn.close()
        total_alerts = sum(len(v) for v in alerts_by_date.values())
        logger.info(f"Found {total_alerts} unique alerts across {len(alerts_by_date)} days")
        return dict(alerts_by_date)

    def get_historical_data(self, ticker: str, from_date: str, to_date: str) -> Optional[pd.DataFrame]:
        """Get historical daily OHLC data with caching"""
        cache_key = f"{ticker}_{from_date}_{to_date}"
        if cache_key in self.price_cache:
            return self.price_cache[cache_key]

        if not self.kite or ticker not in self.instrument_map:
            return None

        try:
            self._throttle()
            token = self.instrument_map[ticker]

            data = self.kite.historical_data(
                instrument_token=token,
                from_date=from_date,
                to_date=to_date,
                interval='day'
            )

            if not data:
                return None

            df = pd.DataFrame(data)
            df['date'] = pd.to_datetime(df['date']).dt.strftime('%Y-%m-%d')
            df.set_index('date', inplace=True)

            self.price_cache[cache_key] = df
            return df

        except Exception as e:
            logger.warning(f"Error getting historical data for {ticker}: {e}")
            return None

    def calculate_kc_for_date(self, ticker: str, target_date: str) -> Optional[Dict]:
        """Calculate KC bands for a specific date"""
        # Need 30 days before target date for KC calculation
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
        from_date = (target_dt - timedelta(days=50)).strftime('%Y-%m-%d')
        to_date = (target_dt + timedelta(days=1)).strftime('%Y-%m-%d')

        df = self.get_historical_data(ticker, from_date, to_date)
        if df is None or len(df) < 25:
            return None

        # Calculate KC
        df['kc_middle'] = df['close'].ewm(span=20, adjust=False).mean()
        df['tr'] = pd.concat([
            df['high'] - df['low'],
            abs(df['high'] - df['close'].shift(1)),
            abs(df['low'] - df['close'].shift(1))
        ], axis=1).max(axis=1)
        df['atr'] = df['tr'].rolling(window=10).mean()
        df['kc_lower'] = df['kc_middle'] - (2.0 * df['atr'])
        df['kc_upper'] = df['kc_middle'] + (2.0 * df['atr'])

        if target_date not in df.index:
            return None

        row = df.loc[target_date]
        return {
            'close': row['close'],
            'low': row['low'],
            'high': row['high'],
            'kc_lower': row['kc_lower'],
            'kc_middle': row['kc_middle'],
            'kc_upper': row['kc_upper']
        }

    def calculate_td_for_date(self, ticker: str, target_date: str) -> Optional[Dict]:
        """Calculate TD indicators for a specific date"""
        cache_key = f"td_{ticker}_{target_date}"
        if cache_key in self.td_cache:
            return self.td_cache[cache_key]

        # Need 60 days before target date for TD calculation
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
        from_date = (target_dt - timedelta(days=90)).strftime('%Y-%m-%d')
        to_date = (target_dt + timedelta(days=1)).strftime('%Y-%m-%d')

        df = self.get_historical_data(ticker, from_date, to_date)
        if df is None or len(df) < 20:
            return None

        # Calculate TD indicators
        df = self.td_calculator.calculate_all(df)

        if target_date not in df.index:
            return None

        row = df.loc[target_date]
        result = {
            'close': row['close'],
            'low': row['low'],
            'high': row['high'],
            # TD MA I
            'td_ma1_active': bool(row.get('td_ma1_active', False)),
            'td_ma1_value': float(row.get('td_ma1_value', 0.0)),
            # TD MA II
            'td_ma2_active': bool(row.get('td_ma2_active', False)),
            'td_ma2_value': float(row.get('td_ma2_value', 0.0)),
            # TD Setup
            'td_setup_count': int(row.get('td_setup_count', 0)),
            'td_setup_complete': bool(row.get('td_setup_complete', False)),
            'td_setup_bar9_close': float(row.get('td_setup_bar9_close', 0.0)),
            'td_setup_bar9_range_pct': float(row.get('td_setup_bar9_range_pct', 0.0)),
            'td_setup_lowest_low': float(row.get('td_setup_lowest_low', 0.0)),
            'bars_since_setup9': int(row.get('bars_since_setup9', 0)),
            'highest_close_since_setup9': float(row.get('highest_close_since_setup9', 0.0)),
            # TDST Support
            'tdst_support': float(row.get('tdst_support', 0.0)),
            'tdst_active': bool(row.get('tdst_active', False)),
            # TDST Resistance
            'tdst_resistance': float(row.get('tdst_resistance', 0.0)),
            'tdst_res_active': bool(row.get('tdst_res_active', False)),
            'tdst_res_broken': bool(row.get('tdst_res_broken', False)),
            # TD Countdown
            'td_countdown': int(row.get('td_countdown', 0)),
            'td_countdown_complete': bool(row.get('td_countdown_complete', False)),
            # Higher Lows
            'recent_higher_low': float(row.get('recent_higher_low', 0.0)),
            # Entry condition
            'td_entry_valid': bool(row.get('td_entry_valid', False))
        }

        self.td_cache[cache_key] = result
        return result

    def calculate_delta_cvd_for_date(self, ticker: str, target_date: str) -> Optional[Dict]:
        """
        Calculate Delta CVD (Cumulative Volume Delta) for a specific date

        For daily data, we approximate CVD using:
        - Buy volume: volume when close > open (bullish day)
        - Sell volume: volume when close < open (bearish day)
        - Delta = Buy volume - Sell volume
        - CVD = Cumulative sum of Delta
        - Delta CVD = Change in CVD
        - EMA50 of Delta CVD

        Returns exit signal when EMA50(Delta CVD) < 0
        """
        cache_key = f"cvd_{ticker}_{target_date}"
        if cache_key in self.price_cache:
            return self.price_cache[cache_key]

        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
        from_date = (target_dt - timedelta(days=90)).strftime('%Y-%m-%d')
        to_date = (target_dt + timedelta(days=1)).strftime('%Y-%m-%d')

        df = self.get_historical_data(ticker, from_date, to_date)
        if df is None or len(df) < 50:
            return None

        # Calculate Delta (volume direction based on candle color)
        df['delta'] = np.where(
            df['close'] > df['open'],
            df['volume'],  # Bullish: all volume is "buy"
            np.where(
                df['close'] < df['open'],
                -df['volume'],  # Bearish: all volume is "sell"
                0  # Doji: neutral
            )
        )

        # CVD = Cumulative Volume Delta
        df['cvd'] = df['delta'].cumsum()

        # Delta CVD = Daily change in CVD (essentially same as delta)
        df['delta_cvd'] = df['cvd'].diff()

        # EMA50 of Delta CVD
        df['ema50_delta_cvd'] = df['delta_cvd'].ewm(span=50, adjust=False).mean()

        if target_date not in df.index:
            return None

        row = df.loc[target_date]
        result = {
            'close': row['close'],
            'low': row['low'],
            'high': row['high'],
            'delta': float(row.get('delta', 0.0)),
            'cvd': float(row.get('cvd', 0.0)),
            'delta_cvd': float(row.get('delta_cvd', 0.0)),
            'ema50_delta_cvd': float(row.get('ema50_delta_cvd', 0.0)),
            'exit_signal': float(row.get('ema50_delta_cvd', 0.0)) < 0
        }

        self.price_cache[cache_key] = result
        return result

    def run_simulation(self, exit_type: str) -> Tuple[List[OpenPosition], List[ClosedTrade], Dict]:
        """
        Run day-by-day simulation

        Args:
            exit_type: 'td_strategy' (3-tier tranche), 'delta_cvd', 'kc_lower' or 'kc_middle'

        Returns:
            Tuple of (open_positions, closed_trades, portfolio_summary)
        """
        if exit_type == 'td_strategy':
            sim_name = "Sim 1: TD 3-Tier Tranche Exit"
        elif exit_type == 'delta_cvd':
            sim_name = "Sim 2: EMA50(Delta CVD) Exit"
        elif exit_type == 'kc_lower':
            sim_name = "Sim 1: KC Lower Exit"
        else:
            sim_name = "Sim 2: KC Middle + 2% Fixed SL"
        logger.info(f"Running simulation: {sim_name}")

        # Get alerts grouped by date
        alerts_by_date = self.get_telegram_alerts_by_date()
        if not alerts_by_date:
            logger.error("No alerts found")
            return [], [], {}

        # Sort dates
        all_dates = sorted(alerts_by_date.keys())
        start_date = datetime.strptime(all_dates[0], '%Y-%m-%d')
        end_date = datetime.now()

        # Initialize portfolio state
        cash = self.initial_capital
        invested = 0.0
        open_positions: Dict[str, OpenPosition] = {}
        closed_trades: List[ClosedTrade] = []
        total_charges = 0.0
        total_overnight_charges = 0.0

        # Generate all trading days
        current_date = start_date
        trading_days = []
        while current_date <= end_date:
            if current_date.weekday() < 5:  # Mon-Fri
                trading_days.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)

        processed = 0
        for date_str in trading_days:
            # 1. Check exits for existing positions
            positions_to_close = []
            for ticker, pos in list(open_positions.items()):
                kc_data = self.calculate_kc_for_date(ticker, date_str)
                if kc_data is None:
                    continue

                current_low = kc_data['low']
                current_close = kc_data['close']

                if exit_type == 'td_strategy':
                    # TD Strategy: 3-Tier Tranche Exit System
                    td_data = self.calculate_td_for_date(ticker, date_str)
                    if td_data is None:
                        continue

                    # Build TDState for exit checks
                    td_state = TDState(
                        td_ma1_active=td_data.get('td_ma1_active', False),
                        td_ma1_value=td_data.get('td_ma1_value', 0.0),
                        td_ma2_active=td_data.get('td_ma2_active', False),
                        td_ma2_value=td_data.get('td_ma2_value', 0.0),
                        td_setup_count=td_data.get('td_setup_count', 0),
                        td_setup_complete=td_data.get('td_setup_complete', False),
                        td_setup_bar9_close=td_data.get('td_setup_bar9_close', 0.0),
                        td_setup_bar9_range_pct=td_data.get('td_setup_bar9_range_pct', 0.0),
                        td_setup_lowest_low=td_data.get('td_setup_lowest_low', 0.0),
                        bars_since_setup9=td_data.get('bars_since_setup9', 0),
                        highest_close_since_setup9=td_data.get('highest_close_since_setup9', 0.0),
                        tdst_support=td_data.get('tdst_support', 0.0) if td_data.get('tdst_support', 0.0) > 0 else pos.tdst_support,
                        tdst_active=td_data.get('tdst_active', False),
                        td_countdown=td_data.get('td_countdown', 0),
                        td_countdown_complete=td_data.get('td_countdown_complete', False),
                        recent_higher_low=td_data.get('recent_higher_low', 0.0)
                    )

                    entry_dt = datetime.strptime(pos.entry_date, '%Y-%m-%d')
                    exit_dt = datetime.strptime(date_str, '%Y-%m-%d')
                    days_held = (exit_dt - entry_dt).days

                    # Check Tranche 1 exit (30%) - if not already exited
                    if not pos.tranche1_exited and pos.position_state == 1:
                        should_exit, reason = self.td_calculator.check_tranche1_exit(current_close, td_state)
                        if should_exit:
                            # Exit 30% of position
                            tranche1_qty = int(pos.original_quantity * 0.30)
                            if tranche1_qty > 0:
                                exit_price = current_close
                                gross_pnl = (exit_price - pos.entry_price) * tranche1_qty
                                exit_charges = (exit_price * tranche1_qty) * (self.charges_per_leg_pct / 100)
                                # Allocate 30% of overnight charges to this tranche
                                tranche_overnight = pos.overnight_charges * 0.30
                                net_pnl = gross_pnl - exit_charges - tranche_overnight

                                closed_trades.append(ClosedTrade(
                                    ticker=ticker,
                                    entry_date=pos.entry_date,
                                    entry_price=pos.entry_price,
                                    exit_date=date_str,
                                    exit_price=round(exit_price, 2),
                                    quantity=tranche1_qty,
                                    position_value=round(pos.original_value * 0.30, 2),
                                    pnl=round(net_pnl, 2),
                                    pnl_pct=round((net_pnl / (pos.original_value * 0.30)) * 100, 2),
                                    exit_reason=f"TRANCHE1_{reason}",
                                    days_held=days_held,
                                    kc_lower_at_entry=pos.kc_lower,
                                    kc_middle_at_entry=pos.kc_middle,
                                    overnight_charges=round(tranche_overnight, 2),
                                    tranche="TRANCHE_1"
                                ))

                                # Update position
                                cash += (exit_price * tranche1_qty) - exit_charges
                                invested -= pos.original_value * 0.30
                                total_charges += exit_charges
                                pos.tranche1_exited = True
                                pos.tranche1_exit_date = date_str
                                pos.tranche1_exit_price = exit_price
                                pos.tranche1_pnl = net_pnl
                                pos.tranche1_reason = reason
                                pos.quantity = pos.original_quantity - tranche1_qty
                                pos.position_value = pos.entry_price * pos.quantity
                                pos.position_state = 2  # De-risked (70%)
                                pos.overnight_charges = pos.overnight_charges * 0.70

                    # Check Tranche 2 exit (45%) - if tranche 1 done but not tranche 2
                    if pos.tranche1_exited and not pos.tranche2_exited and pos.position_state == 2:
                        setup_lowest = pos.setup_lowest_low if pos.setup_lowest_low > 0 else td_state.td_setup_lowest_low
                        should_exit, reason = self.td_calculator.check_tranche2_exit(current_close, td_state, setup_lowest)
                        if should_exit:
                            # Exit 45% of original position (which is ~64% of remaining)
                            tranche2_qty = int(pos.original_quantity * 0.45)
                            if tranche2_qty > 0 and tranche2_qty <= pos.quantity:
                                exit_price = current_close
                                gross_pnl = (exit_price - pos.entry_price) * tranche2_qty
                                exit_charges = (exit_price * tranche2_qty) * (self.charges_per_leg_pct / 100)
                                # Allocate proportional overnight charges
                                tranche_overnight = pos.overnight_charges * (tranche2_qty / pos.quantity)
                                net_pnl = gross_pnl - exit_charges - tranche_overnight

                                closed_trades.append(ClosedTrade(
                                    ticker=ticker,
                                    entry_date=pos.entry_date,
                                    entry_price=pos.entry_price,
                                    exit_date=date_str,
                                    exit_price=round(exit_price, 2),
                                    quantity=tranche2_qty,
                                    position_value=round(pos.original_value * 0.45, 2),
                                    pnl=round(net_pnl, 2),
                                    pnl_pct=round((net_pnl / (pos.original_value * 0.45)) * 100, 2),
                                    exit_reason=f"TRANCHE2_{reason}",
                                    days_held=days_held,
                                    kc_lower_at_entry=pos.kc_lower,
                                    kc_middle_at_entry=pos.kc_middle,
                                    overnight_charges=round(tranche_overnight, 2),
                                    tranche="TRANCHE_2"
                                ))

                                cash += (exit_price * tranche2_qty) - exit_charges
                                invested -= pos.original_value * 0.45
                                total_charges += exit_charges
                                pos.tranche2_exited = True
                                pos.tranche2_exit_date = date_str
                                pos.tranche2_exit_price = exit_price
                                pos.tranche2_pnl = net_pnl
                                pos.tranche2_reason = reason
                                pos.quantity = pos.original_quantity - int(pos.original_quantity * 0.30) - tranche2_qty
                                pos.position_value = pos.entry_price * pos.quantity
                                pos.position_state = 3  # Runner (25%)
                                pos.overnight_charges = pos.overnight_charges - tranche_overnight

                                # Mark TDST violated if that was the reason
                                if "TDST" in reason:
                                    pos.tdst_violated = True

                    # Check Tranche 3 exit (25%) - runner position
                    if pos.tranche2_exited and pos.position_state == 3 and pos.quantity > 0:
                        should_exit, reason = self.td_calculator.check_tranche3_exit(
                            current_close, td_state, pos.entry_price, days_held
                        )
                        if should_exit:
                            # Exit remaining 25%
                            tranche3_qty = pos.quantity
                            exit_price = current_close
                            gross_pnl = (exit_price - pos.entry_price) * tranche3_qty
                            exit_charges = (exit_price * tranche3_qty) * (self.charges_per_leg_pct / 100)
                            net_pnl = gross_pnl - exit_charges - pos.overnight_charges

                            closed_trades.append(ClosedTrade(
                                ticker=ticker,
                                entry_date=pos.entry_date,
                                entry_price=pos.entry_price,
                                exit_date=date_str,
                                exit_price=round(exit_price, 2),
                                quantity=tranche3_qty,
                                position_value=round(pos.original_value * 0.25, 2),
                                pnl=round(net_pnl, 2),
                                pnl_pct=round((net_pnl / (pos.original_value * 0.25)) * 100, 2),
                                exit_reason=f"TRANCHE3_{reason}",
                                days_held=days_held,
                                kc_lower_at_entry=pos.kc_lower,
                                kc_middle_at_entry=pos.kc_middle,
                                overnight_charges=round(pos.overnight_charges, 2),
                                tranche="TRANCHE_3"
                            ))

                            cash += (exit_price * tranche3_qty) - exit_charges
                            invested -= pos.position_value
                            total_charges += exit_charges
                            pos.position_state = 4  # Flat
                            positions_to_close.append(ticker)

                elif exit_type == 'delta_cvd':
                    # Sim 2: Exit when EMA50(Delta CVD) < 0
                    cvd_data = self.calculate_delta_cvd_for_date(ticker, date_str)
                    if cvd_data is None:
                        continue

                    if cvd_data.get('exit_signal', False):
                        exit_price = current_close
                        gross_pnl = (exit_price - pos.entry_price) * pos.quantity
                        exit_charges = (exit_price * pos.quantity) * (self.charges_per_leg_pct / 100)
                        net_pnl = gross_pnl - exit_charges - pos.overnight_charges

                        entry_dt = datetime.strptime(pos.entry_date, '%Y-%m-%d')
                        exit_dt = datetime.strptime(date_str, '%Y-%m-%d')
                        days_held = (exit_dt - entry_dt).days

                        closed_trades.append(ClosedTrade(
                            ticker=ticker,
                            entry_date=pos.entry_date,
                            entry_price=pos.entry_price,
                            exit_date=date_str,
                            exit_price=round(exit_price, 2),
                            quantity=pos.quantity,
                            position_value=pos.position_value,
                            pnl=round(net_pnl, 2),
                            pnl_pct=round((net_pnl / pos.position_value) * 100, 2),
                            exit_reason='EMA50_DELTA_CVD_NEGATIVE',
                            days_held=days_held,
                            kc_lower_at_entry=pos.kc_lower,
                            kc_middle_at_entry=pos.kc_middle,
                            overnight_charges=round(pos.overnight_charges, 2)
                        ))

                        cash += (exit_price * pos.quantity) - exit_charges
                        invested -= pos.position_value
                        total_charges += exit_charges
                        positions_to_close.append(ticker)

                elif exit_type == 'kc_lower':
                    # Sim 1 (KC): Only KC Lower stop
                    stop_level = kc_data['kc_lower']
                    if current_low <= stop_level:
                        exit_price = stop_level
                        gross_pnl = (exit_price - pos.entry_price) * pos.quantity
                        exit_charges = (exit_price * pos.quantity) * (self.charges_per_leg_pct / 100)
                        net_pnl = gross_pnl - exit_charges - pos.overnight_charges

                        entry_dt = datetime.strptime(pos.entry_date, '%Y-%m-%d')
                        exit_dt = datetime.strptime(date_str, '%Y-%m-%d')
                        days_held = (exit_dt - entry_dt).days

                        closed_trades.append(ClosedTrade(
                            ticker=ticker,
                            entry_date=pos.entry_date,
                            entry_price=pos.entry_price,
                            exit_date=date_str,
                            exit_price=round(exit_price, 2),
                            quantity=pos.quantity,
                            position_value=pos.position_value,
                            pnl=round(net_pnl, 2),
                            pnl_pct=round((net_pnl / pos.position_value) * 100, 2),
                            exit_reason='KC_LOWER_BREACH',
                            days_held=days_held,
                            kc_lower_at_entry=pos.kc_lower,
                            kc_middle_at_entry=pos.kc_middle,
                            overnight_charges=round(pos.overnight_charges, 2)
                        ))

                        cash += (exit_price * pos.quantity) - exit_charges
                        invested -= pos.position_value
                        total_charges += exit_charges
                        positions_to_close.append(ticker)

                else:
                    # Sim 2: KC Middle OR 2% below entry (whichever comes first)
                    stop_level = kc_data['kc_middle']
                    fixed_stop = pos.entry_price * 0.98  # 2% below entry

                    stop_hit = False
                    exit_reason = ''
                    if current_low <= fixed_stop:
                        stop_hit = True
                        exit_price = fixed_stop
                        exit_reason = 'FIXED_2PCT_SL'
                    elif current_low <= stop_level:
                        stop_hit = True
                        exit_price = stop_level
                        exit_reason = 'KC_MIDDLE_BREACH'

                    if stop_hit:
                        gross_pnl = (exit_price - pos.entry_price) * pos.quantity
                        exit_charges = (exit_price * pos.quantity) * (self.charges_per_leg_pct / 100)
                        net_pnl = gross_pnl - exit_charges - pos.overnight_charges

                        entry_dt = datetime.strptime(pos.entry_date, '%Y-%m-%d')
                        exit_dt = datetime.strptime(date_str, '%Y-%m-%d')
                        days_held = (exit_dt - entry_dt).days

                        closed_trades.append(ClosedTrade(
                            ticker=ticker,
                            entry_date=pos.entry_date,
                            entry_price=pos.entry_price,
                            exit_date=date_str,
                            exit_price=round(exit_price, 2),
                            quantity=pos.quantity,
                            position_value=pos.position_value,
                            pnl=round(net_pnl, 2),
                            pnl_pct=round((net_pnl / pos.position_value) * 100, 2),
                            exit_reason=exit_reason,
                            days_held=days_held,
                            kc_lower_at_entry=pos.kc_lower,
                            kc_middle_at_entry=pos.kc_middle,
                            overnight_charges=round(pos.overnight_charges, 2)
                        ))

                        cash += (exit_price * pos.quantity) - exit_charges
                        invested -= pos.position_value
                        total_charges += exit_charges
                        positions_to_close.append(ticker)

            # Remove fully closed positions
            for ticker in positions_to_close:
                del open_positions[ticker]

            # 2. Process new alerts for this date
            if date_str in alerts_by_date:
                for alert in alerts_by_date[date_str]:
                    ticker = alert['ticker']

                    # Skip if already have position
                    if ticker in open_positions:
                        continue

                    # Calculate available capital with margin
                    total_available = cash + (self.initial_capital * self.margin_pct / 100) - invested

                    # Calculate position size (5% of initial capital)
                    position_value = self.initial_capital * (self.position_size_pct / 100)

                    if position_value > total_available:
                        continue  # Not enough capital

                    # Get KC data for entry
                    kc_data = self.calculate_kc_for_date(ticker, date_str)
                    if kc_data is None:
                        continue

                    # Entry conditions based on strategy
                    tdst_support = 0.0
                    setup_lowest_low = 0.0

                    if exit_type == 'td_strategy':
                        # Sim 1: TD MA I AND TD MA II both active
                        td_data = self.calculate_td_for_date(ticker, date_str)
                        if td_data is None:
                            continue
                        if not td_data.get('td_entry_valid', False):
                            continue  # Skip if TD entry conditions not met
                        # Get TDST Support for exit
                        tdst_support = td_data.get('tdst_support', 0.0)
                        setup_lowest_low = td_data.get('td_setup_lowest_low', 0.0)
                        if tdst_support <= 0:
                            tdst_support = alert['entry_price'] * 0.98

                    elif exit_type == 'delta_cvd':
                        # Sim 2: EMA50(Delta CVD) > 0 AND TDST Resistance broken
                        cvd_data = self.calculate_delta_cvd_for_date(ticker, date_str)
                        if cvd_data is None:
                            continue
                        # Check EMA50(Delta CVD) > 0
                        if cvd_data.get('ema50_delta_cvd', 0) <= 0:
                            continue  # Skip if CVD not bullish

                        # Check TDST Resistance broken
                        td_data = self.calculate_td_for_date(ticker, date_str)
                        if td_data is None:
                            continue
                        if not td_data.get('tdst_res_broken', False):
                            continue  # Skip if TDST Resistance not broken

                    entry_price = alert['entry_price']
                    quantity = int(position_value / entry_price)
                    if quantity == 0:
                        continue

                    actual_value = quantity * entry_price
                    entry_charges = actual_value * (self.charges_per_leg_pct / 100)

                    # Determine stop loss based on strategy
                    if exit_type == 'td_strategy':
                        stop_loss = tdst_support
                    elif exit_type == 'delta_cvd':
                        stop_loss = entry_price * 0.95  # 5% default stop for CVD strategy
                    elif exit_type == 'kc_lower':
                        stop_loss = kc_data['kc_lower']
                    else:
                        stop_loss = kc_data['kc_middle']

                    # Open position
                    open_positions[ticker] = OpenPosition(
                        ticker=ticker,
                        entry_date=date_str,
                        entry_price=entry_price,
                        quantity=quantity,
                        position_value=actual_value,
                        stop_loss=stop_loss,
                        kc_lower=kc_data['kc_lower'],
                        kc_middle=kc_data['kc_middle'],
                        tdst_support=tdst_support,
                        original_quantity=quantity,  # For TD tranche tracking
                        original_value=actual_value,  # For TD tranche tracking
                        setup_lowest_low=setup_lowest_low  # For TD tranche 2 exit
                    )

                    cash -= (actual_value + entry_charges)
                    invested += actual_value
                    total_charges += entry_charges

            # 3. Apply overnight charges to all positions still open at end of day
            for ticker, pos in open_positions.items():
                overnight_charge = pos.position_value * (self.overnight_charge_pct / 100)
                pos.overnight_charges += overnight_charge
                cash -= overnight_charge
                total_overnight_charges += overnight_charge

            processed += 1
            if processed % 20 == 0:
                logger.info(f"Processed {processed}/{len(trading_days)} days, {len(open_positions)} open, {len(closed_trades)} closed")

        # Mark remaining positions with current prices
        final_open = []
        for ticker, pos in open_positions.items():
            # Get latest price
            kc_data = self.calculate_kc_for_date(ticker, trading_days[-1])
            if kc_data:
                current_price = kc_data['close']
                unrealized_pnl = (current_price - pos.entry_price) * pos.quantity
            else:
                current_price = pos.entry_price
                unrealized_pnl = 0

            # Include overnight charges in unrealized P&L
            net_unrealized_pnl = unrealized_pnl - pos.overnight_charges

            final_open.append({
                'ticker': ticker,
                'entry_date': pos.entry_date,
                'entry_price': pos.entry_price,
                'current_price': round(current_price, 2),
                'quantity': pos.quantity,
                'position_value': pos.position_value,
                'unrealized_pnl': round(net_unrealized_pnl, 2),
                'unrealized_pnl_pct': round((net_unrealized_pnl / pos.position_value) * 100, 2),
                'stop_loss': round(pos.stop_loss, 2),
                'kc_lower': round(pos.kc_lower, 2),
                'kc_middle': round(pos.kc_middle, 2),
                'overnight_charges': round(pos.overnight_charges, 2)
            })

        # Calculate summary
        realized_pnl = sum(t.pnl for t in closed_trades)
        unrealized_pnl = sum(p['unrealized_pnl'] for p in final_open)
        total_pnl = realized_pnl + unrealized_pnl

        winning_trades = [t for t in closed_trades if t.pnl > 0]
        losing_trades = [t for t in closed_trades if t.pnl < 0]

        summary = {
            'sim_name': sim_name,
            'exit_type': exit_type,
            'initial_capital': self.initial_capital,
            'final_value': self.initial_capital + total_pnl,
            'cash': cash,
            'invested': invested,
            'realized_pnl': realized_pnl,
            'unrealized_pnl': unrealized_pnl,
            'total_pnl': total_pnl,
            'total_pnl_pct': (total_pnl / self.initial_capital) * 100,
            'total_trades': len(closed_trades) + len(final_open),
            'closed_trades': len(closed_trades),
            'open_positions': len(final_open),
            'winning_trades': len(winning_trades),
            'losing_trades': len(losing_trades),
            'win_rate': (len(winning_trades) / len(closed_trades) * 100) if closed_trades else 0,
            'avg_win': np.mean([t.pnl for t in winning_trades]) if winning_trades else 0,
            'avg_loss': np.mean([t.pnl for t in losing_trades]) if losing_trades else 0,
            'max_win': max([t.pnl for t in winning_trades]) if winning_trades else 0,
            'max_loss': min([t.pnl for t in losing_trades]) if losing_trades else 0,
            'avg_days_held': np.mean([t.days_held for t in closed_trades]) if closed_trades else 0,
            'total_charges': total_charges,
            'total_overnight_charges': total_overnight_charges
        }

        return final_open, closed_trades, summary

    def generate_report(self, open_positions: List, closed_trades: List[ClosedTrade],
                       summary: Dict, filename: str):
        """Generate Excel report"""
        wb = Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        summary_data = [
            ["SIMULATION BACKTEST RESULTS", ""],
            ["", ""],
            ["Simulation", summary['sim_name']],
            ["Exit Strategy", summary['exit_type'].upper()],
            ["Lookback Period", f"{self.lookback_days} days"],
            ["Margin", f"{self.margin_pct}%"],
            ["", ""],
            ["CAPITAL", ""],
            ["Initial Capital", f"₹{summary['initial_capital']:,.0f}"],
            ["Final Value", f"₹{summary['final_value']:,.0f}"],
            ["Realized P&L", f"₹{summary['realized_pnl']:,.0f}"],
            ["Unrealized P&L", f"₹{summary['unrealized_pnl']:,.0f}"],
            ["Total P&L", f"₹{summary['total_pnl']:,.0f}"],
            ["Return %", f"{summary['total_pnl_pct']:.2f}%"],
            ["", ""],
            ["TRADE STATISTICS", ""],
            ["Total Trades", summary['total_trades']],
            ["Closed Trades", summary['closed_trades']],
            ["Open Positions", summary['open_positions']],
            ["Winning Trades", summary['winning_trades']],
            ["Losing Trades", summary['losing_trades']],
            ["Win Rate", f"{summary['win_rate']:.1f}%"],
            ["Avg Win", f"₹{summary['avg_win']:,.0f}"],
            ["Avg Loss", f"₹{summary['avg_loss']:,.0f}"],
            ["Max Win", f"₹{summary['max_win']:,.0f}"],
            ["Max Loss", f"₹{summary['max_loss']:,.0f}"],
            ["Avg Days Held", f"{summary['avg_days_held']:.1f}"],
            ["Total Charges", f"₹{summary['total_charges']:,.0f}"],
            ["Overnight Charges", f"₹{summary['total_overnight_charges']:,.0f}"],
        ]

        for row in summary_data:
            ws.append(row)

        # Open Positions sheet
        ws2 = wb.create_sheet("Open Positions")
        headers = ["Ticker", "Entry Date", "Entry Price", "Current Price", "Quantity",
                  "Value", "Unrealized P&L", "P&L %", "Overnight Charges", "Stop Loss", "KC Lower", "KC Middle"]
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, pos in enumerate(open_positions, 2):
            ws2.cell(row=i, column=1, value=pos['ticker'])
            ws2.cell(row=i, column=2, value=pos['entry_date'])
            ws2.cell(row=i, column=3, value=pos['entry_price'])
            ws2.cell(row=i, column=4, value=pos['current_price'])
            ws2.cell(row=i, column=5, value=pos['quantity'])
            ws2.cell(row=i, column=6, value=pos['position_value'])
            ws2.cell(row=i, column=7, value=pos['unrealized_pnl'])
            ws2.cell(row=i, column=8, value=pos['unrealized_pnl_pct'])
            ws2.cell(row=i, column=9, value=pos.get('overnight_charges', 0))
            ws2.cell(row=i, column=10, value=pos['stop_loss'])
            ws2.cell(row=i, column=11, value=pos['kc_lower'])
            ws2.cell(row=i, column=12, value=pos['kc_middle'])

        # Closed Trades sheet
        ws3 = wb.create_sheet("Trade History")
        headers = ["Ticker", "Entry Date", "Entry Price", "Exit Date", "Exit Price",
                  "Quantity", "P&L", "P&L %", "Exit Reason", "Days Held", "Overnight Charges"]
        for col, h in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, trade in enumerate(closed_trades, 2):
            ws3.cell(row=i, column=1, value=trade.ticker)
            ws3.cell(row=i, column=2, value=trade.entry_date)
            ws3.cell(row=i, column=3, value=trade.entry_price)
            ws3.cell(row=i, column=4, value=trade.exit_date)
            ws3.cell(row=i, column=5, value=trade.exit_price)
            ws3.cell(row=i, column=6, value=trade.quantity)
            ws3.cell(row=i, column=7, value=trade.pnl)
            ws3.cell(row=i, column=8, value=trade.pnl_pct)
            ws3.cell(row=i, column=9, value=trade.exit_reason)
            ws3.cell(row=i, column=10, value=trade.days_held)
            ws3.cell(row=i, column=11, value=trade.overnight_charges)

        filepath = self.output_dir / filename
        wb.save(filepath)
        logger.info(f"Report saved: {filepath}")
        return filepath


def main():
    """Run backtests for both strategies"""
    import argparse

    parser = argparse.ArgumentParser(description='Telegram Alert Backtester')
    parser.add_argument('--days', type=int, default=60, help='Lookback days (default: 60)')
    parser.add_argument('--capital', type=float, default=10000000, help='Initial capital (default: 1 Cr)')
    parser.add_argument('--position-size', type=float, default=5.0, help='Position size % (default: 5)')
    parser.add_argument('--margin', type=float, default=100.0, help='Margin allowed % (default: 100)')
    args = parser.parse_args()

    backtester = TelegramAlertBacktester(
        initial_capital=args.capital,
        position_size_pct=args.position_size,
        lookback_days=args.days,
        margin_pct=args.margin
    )

    print("\n" + "=" * 70)
    print("TELEGRAM ALERT BACKTESTER - DAY-BY-DAY SIMULATION")
    print(f"Capital: ₹{args.capital:,.0f} | Position Size: {args.position_size}%")
    print(f"Lookback: {args.days} days | Margin: {args.margin}%")
    print("=" * 70)

    today = datetime.now().strftime('%Y%m%d')

    # Run Sim 1: TD Strategy (3-Tier Tranche Exit)
    print("\n[1/2] Running Sim 1: TD 3-Tier Tranche Exit...")
    print("       Entry: TD MA I + TD MA II active")
    print("       Exit: 30% @ TD MA I breach, 45% @ TDST breach, 25% @ Countdown/Time")
    open1, closed1, summary1 = backtester.run_simulation('td_strategy')
    backtester.generate_report(open1, closed1, summary1, f"Backtest_Sim1_TD_Tranche_{today}.xlsx")

    # Run Sim 2: Delta CVD Strategy
    print("\n[2/2] Running Sim 2: EMA50(Delta CVD) Exit...")
    print("       Entry: EMA50(Delta CVD) > 0 AND TDST Resistance broken")
    print("       Exit: EMA50(Delta CVD) < 0")
    open2, closed2, summary2 = backtester.run_simulation('delta_cvd')
    backtester.generate_report(open2, closed2, summary2, f"Backtest_Sim2_DeltaCVD_{today}.xlsx")

    # Print comparison
    print("\n" + "=" * 70)
    print("BACKTEST RESULTS COMPARISON")
    print("=" * 70)

    print(f"\n{'Metric':<25} {'Sim 1 (TD Tranche)':<20} {'Sim 2 (Delta CVD)':<20}")
    print("-" * 65)
    print(f"{'Total Trades':<25} {summary1['total_trades']:<20} {summary2['total_trades']:<20}")
    print(f"{'Closed Trades':<25} {summary1['closed_trades']:<20} {summary2['closed_trades']:<20}")
    print(f"{'Open Positions':<25} {summary1['open_positions']:<20} {summary2['open_positions']:<20}")
    print(f"{'Win Rate':<25} {summary1['win_rate']:.1f}%{'':<17} {summary2['win_rate']:.1f}%")
    print(f"{'Realized P&L':<25} ₹{summary1['realized_pnl']:>14,.0f} ₹{summary2['realized_pnl']:>14,.0f}")
    print(f"{'Unrealized P&L':<25} ₹{summary1['unrealized_pnl']:>14,.0f} ₹{summary2['unrealized_pnl']:>14,.0f}")
    print(f"{'Total P&L':<25} ₹{summary1['total_pnl']:>14,.0f} ₹{summary2['total_pnl']:>14,.0f}")
    print(f"{'Return %':<25} {summary1['total_pnl_pct']:>14.2f}% {summary2['total_pnl_pct']:>14.2f}%")
    print(f"{'Avg Days Held':<25} {summary1['avg_days_held']:>14.1f} {summary2['avg_days_held']:>14.1f}")
    print(f"{'Total Charges':<25} ₹{summary1['total_charges']:>14,.0f} ₹{summary2['total_charges']:>14,.0f}")
    print(f"{'Overnight Charges':<25} ₹{summary1['total_overnight_charges']:>14,.0f} ₹{summary2['total_overnight_charges']:>14,.0f}")

    print("\n" + "-" * 65)
    if summary1['total_pnl'] > summary2['total_pnl']:
        print(f"WINNER: Sim 1 (TD Tranche) by ₹{summary1['total_pnl'] - summary2['total_pnl']:,.0f}")
    elif summary2['total_pnl'] > summary1['total_pnl']:
        print(f"WINNER: Sim 2 (Delta CVD) by ₹{summary2['total_pnl'] - summary1['total_pnl']:,.0f}")
    else:
        print("TIE: Both strategies performed equally")

    print("\n" + "=" * 70)
    print(f"Reports saved to: {backtester.output_dir}")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    main()
