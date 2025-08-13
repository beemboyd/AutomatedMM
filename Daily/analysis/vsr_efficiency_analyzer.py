#!/usr/bin/env python3
"""
VSR Efficiency Analyzer
Analyzes VSR Daily Dashboard alerts over the last 10 business days
Generates Excel reports for long and short positions
"""

import os
import sys
import json
import pandas as pd
from datetime import datetime, timedelta
import glob
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class VSREfficiencyAnalyzer:
    def __init__(self, lookback_days=10):
        """Initialize the analyzer with lookback period"""
        self.lookback_days = lookback_days
        self.base_dir = "/Users/maverick/PycharmProjects/India-TS/Daily"
        self.output_dir = os.path.join(self.base_dir, "analysis", "Efficiency")
        
        # Ensure output directory exists
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Calculate date range
        self.end_date = datetime.now()
        self.start_date = self.end_date - timedelta(days=lookback_days + 5)  # Extra days for weekends
        
        # Data storage
        self.long_alerts = {}
        self.short_alerts = {}
        
    def get_business_days(self):
        """Get list of last N business days"""
        business_days = []
        current_date = self.end_date
        
        while len(business_days) < self.lookback_days:
            # Skip weekends (Saturday=5, Sunday=6)
            if current_date.weekday() < 5:
                business_days.append(current_date)
            current_date -= timedelta(days=1)
            
        return sorted(business_days)
    
    def parse_vsr_scanner_files(self):
        """Parse VSR scanner result files to extract alerts"""
        logger.info(f"Parsing VSR scanner files from {self.start_date.strftime('%Y%m%d')} to {self.end_date.strftime('%Y%m%d')}")
        
        # Directories to search
        scanner_dirs = [
            os.path.join(self.base_dir, "scanners", "Hourly"),
            os.path.join(self.base_dir, "results"),
            os.path.join(self.base_dir, "results-h"),
            os.path.join(self.base_dir, "results-s"),
            os.path.join(self.base_dir, "results-s-h")
        ]
        
        business_days = self.get_business_days()
        
        for day in business_days:
            date_str = day.strftime("%Y%m%d")
            logger.info(f"Processing date: {date_str}")
            
            # Parse VSR scanner files
            for scanner_dir in scanner_dirs:
                if not os.path.exists(scanner_dir):
                    continue
                    
                # VSR scanner files
                vsr_pattern = os.path.join(scanner_dir, f"VSR_*{date_str}*.xlsx")
                vsr_files = glob.glob(vsr_pattern)
                
                for vsr_file in vsr_files:
                    self.parse_vsr_excel(vsr_file, day)
                    
                # Long reversal files
                long_pattern = os.path.join(scanner_dir, f"Long_Reversal_*{date_str}*.xlsx")
                long_files = glob.glob(long_pattern)
                
                for long_file in long_files:
                    self.parse_reversal_excel(long_file, day, "long")
                    
                # Short reversal files
                short_pattern = os.path.join(scanner_dir, f"Short_Reversal_*{date_str}*.xlsx")
                short_files = glob.glob(short_pattern)
                
                for short_file in short_files:
                    self.parse_reversal_excel(short_file, day, "short")
    
    def parse_vsr_excel(self, filepath, date):
        """Parse VSR Excel file to extract ticker alerts"""
        try:
            filename = os.path.basename(filepath)
            # Extract time from filename (e.g., VSR_20250808_093037.xlsx)
            time_str = filename.split('_')[-1].replace('.xlsx', '')
            if len(time_str) == 6:
                alert_time = datetime.strptime(f"{date.strftime('%Y%m%d')}_{time_str}", "%Y%m%d_%H%M%S")
            else:
                alert_time = date
            
            # Read Excel file
            df = pd.read_excel(filepath)
            
            # Check for VSR Score column
            if 'VSR Score' in df.columns or 'Score' in df.columns:
                score_col = 'VSR Score' if 'VSR Score' in df.columns else 'Score'
                
                for _, row in df.iterrows():
                    ticker = row.get('Ticker', row.get('Symbol', ''))
                    if not ticker:
                        continue
                        
                    score = row.get(score_col, 0)
                    price = row.get('Entry_Price', row.get('Current Price', row.get('Price', row.get('Close', 0))))
                    
                    # Determine if long or short based on score or other indicators
                    if score > 50:  # High VSR score indicates long opportunity
                        if ticker not in self.long_alerts:
                            self.long_alerts[ticker] = {
                                'first_alert_time': alert_time,
                                'first_price': price,
                                'alert_count': 0,
                                'all_alerts': []
                            }
                        
                        self.long_alerts[ticker]['alert_count'] += 1
                        self.long_alerts[ticker]['all_alerts'].append({
                            'time': alert_time,
                            'price': price,
                            'score': score
                        })
                    
            logger.debug(f"Parsed {filepath}: Found {len(df)} rows")
            
        except Exception as e:
            logger.warning(f"Error parsing {filepath}: {e}")
    
    def parse_reversal_excel(self, filepath, date, signal_type):
        """Parse Long/Short Reversal Excel files"""
        try:
            filename = os.path.basename(filepath)
            # Extract time from filename
            parts = filename.split('_')
            if len(parts) >= 4:
                time_str = parts[-1].replace('.xlsx', '')
                if len(time_str) == 6:
                    alert_time = datetime.strptime(f"{date.strftime('%Y%m%d')}_{time_str}", "%Y%m%d_%H%M%S")
                else:
                    alert_time = date
            else:
                alert_time = date
            
            # Read Excel file
            df = pd.read_excel(filepath)
            
            # Process based on signal type
            target_dict = self.long_alerts if signal_type == "long" else self.short_alerts
            
            for _, row in df.iterrows():
                ticker = row.get('Ticker', row.get('Symbol', ''))
                if not ticker:
                    continue
                    
                price = row.get('Entry_Price', row.get('Current Price', row.get('Price', row.get('Close', 0))))
                score = row.get('VSR Score', row.get('Score', row.get('Signal Strength', 0)))
                
                if ticker not in target_dict:
                    target_dict[ticker] = {
                        'first_alert_time': alert_time,
                        'first_price': price,
                        'alert_count': 0,
                        'all_alerts': []
                    }
                
                target_dict[ticker]['alert_count'] += 1
                target_dict[ticker]['all_alerts'].append({
                    'time': alert_time,
                    'price': price,
                    'score': score
                })
            
            logger.debug(f"Parsed {signal_type} reversal {filepath}: Found {len(df)} rows")
            
        except Exception as e:
            logger.warning(f"Error parsing {filepath}: {e}")
    
    def parse_persistence_files(self):
        """Parse VSR persistence JSON files for additional alert data"""
        persistence_files = [
            os.path.join(self.base_dir, "data", "vsr_ticker_persistence.json"),
            os.path.join(self.base_dir, "data", "vsr_ticker_persistence_hourly_long.json"),
            os.path.join(self.base_dir, "data", "short_momentum", "vsr_ticker_persistence_hourly_short.json")
        ]
        
        for persist_file in persistence_files:
            if not os.path.exists(persist_file):
                continue
                
            try:
                with open(persist_file, 'r') as f:
                    data = json.load(f)
                    
                tickers = data.get('tickers', {})
                
                for ticker, info in tickers.items():
                    first_seen = info.get('first_seen', '')
                    if first_seen:
                        try:
                            first_date = datetime.strptime(first_seen, "%Y-%m-%d")
                            
                            # Check if within our date range
                            if first_date >= self.start_date:
                                # Determine if long or short based on file path
                                if 'short' in persist_file:
                                    target_dict = self.short_alerts
                                else:
                                    target_dict = self.long_alerts
                                
                                if ticker not in target_dict:
                                    target_dict[ticker] = {
                                        'first_alert_time': first_date,
                                        'first_price': 0,  # Price not available in persistence
                                        'alert_count': info.get('days_tracked', 1),
                                        'all_alerts': [],
                                        'max_score': info.get('max_score', 0),
                                        'avg_score': info.get('avg_score', 0)
                                    }
                                else:
                                    # Update with persistence data
                                    target_dict[ticker]['max_score'] = info.get('max_score', 0)
                                    target_dict[ticker]['avg_score'] = info.get('avg_score', 0)
                                    
                        except Exception as e:
                            logger.debug(f"Error parsing date for {ticker}: {e}")
                            
            except Exception as e:
                logger.warning(f"Error parsing persistence file {persist_file}: {e}")
    
    def create_efficiency_report(self, alerts_dict, signal_type):
        """Create Excel report for efficiency analysis"""
        if not alerts_dict:
            logger.warning(f"No {signal_type} alerts found for the period")
            return None
            
        # Convert to DataFrame
        report_data = []
        
        for ticker, data in alerts_dict.items():
            row = {
                'Ticker': ticker,
                'First Alert Date': data['first_alert_time'].strftime('%Y-%m-%d') if isinstance(data['first_alert_time'], datetime) else str(data['first_alert_time']),
                'First Alert Time': data['first_alert_time'].strftime('%H:%M:%S') if isinstance(data['first_alert_time'], datetime) else '',
                'First Price': round(data['first_price'], 2) if data['first_price'] else 0,
                'Alert Count': data['alert_count'],
                'Max Score': data.get('max_score', 0),
                'Avg Score': round(data.get('avg_score', 0), 2)
            }
            
            # Add latest alert info if available
            if data['all_alerts']:
                latest = data['all_alerts'][-1]
                row['Latest Alert Time'] = latest['time'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(latest['time'], datetime) else str(latest['time'])
                row['Latest Price'] = round(latest['price'], 2) if latest['price'] else 0
                
                # Calculate price change (as decimal for percentage format)
                if data['first_price'] and latest['price']:
                    price_change = (latest['price'] - data['first_price']) / data['first_price']
                    row['Price Change %'] = round(price_change, 4)  # Keep as decimal (0.05 = 5%)
                else:
                    row['Price Change %'] = 0
            
            report_data.append(row)
        
        # Create DataFrame and sort by alert count
        df = pd.DataFrame(report_data)
        df = df.sort_values('Alert Count', ascending=False)
        
        # Generate filename with date range
        start_str = self.get_business_days()[-1].strftime('%Y%m%d')
        end_str = self.get_business_days()[0].strftime('%Y%m%d')
        filename = f"Eff_Analysis_{signal_type}_{start_str}_{end_str}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # Create Excel file with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = f"{signal_type.capitalize()} Efficiency"
        
        # Add headers
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format numbers
                if isinstance(value, (int, float)):
                    if c_idx == 4:  # First Price column
                        cell.number_format = '#,##0.00'
                    elif c_idx == 5:  # Alert Count
                        cell.number_format = '#,##0'
                    elif 'Score' in headers[c_idx-1]:
                        cell.number_format = '#,##0.00'
                    elif 'Price Change' in headers[c_idx-1]:
                        cell.number_format = '+#,##0.00%;-#,##0.00%'
                        # Color code based on positive/negative
                        if value > 0:
                            cell.font = Font(color="008000")  # Green
                        elif value < 0:
                            cell.font = Font(color="FF0000")  # Red
                
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary statistics
        ws.append([])
        ws.append(['Summary Statistics'])
        ws.append(['Total Tickers', len(df)])
        ws.append(['Total Alerts', df['Alert Count'].sum()])
        ws.append(['Avg Alerts per Ticker', round(df['Alert Count'].mean(), 2)])
        ws.append(['Most Active Ticker', df.iloc[0]['Ticker'] if len(df) > 0 else 'N/A'])
        
        # Save the workbook
        wb.save(filepath)
        logger.info(f"Created {signal_type} efficiency report: {filepath}")
        
        return filepath
    
    def run_analysis(self):
        """Main method to run the complete analysis"""
        logger.info("Starting VSR Efficiency Analysis")
        logger.info(f"Analyzing last {self.lookback_days} business days")
        
        # Parse all data sources
        self.parse_vsr_scanner_files()
        self.parse_persistence_files()
        
        # Create reports
        long_report = self.create_efficiency_report(self.long_alerts, 'long')
        short_report = self.create_efficiency_report(self.short_alerts, 'short')
        
        # Print summary
        logger.info("\n" + "="*50)
        logger.info("VSR EFFICIENCY ANALYSIS COMPLETE")
        logger.info("="*50)
        logger.info(f"Analysis Period: Last {self.lookback_days} business days")
        logger.info(f"Long Alerts Found: {len(self.long_alerts)} tickers")
        logger.info(f"Short Alerts Found: {len(self.short_alerts)} tickers")
        
        if long_report:
            logger.info(f"Long Report: {long_report}")
        if short_report:
            logger.info(f"Short Report: {short_report}")
            
        logger.info("="*50)
        
        return long_report, short_report


def main():
    """Main entry point"""
    analyzer = VSREfficiencyAnalyzer(lookback_days=10)
    long_report, short_report = analyzer.run_analysis()
    
    if long_report or short_report:
        print("\nAnalysis completed successfully!")
        print(f"Reports saved to: {analyzer.output_dir}")
    else:
        print("\nNo data found for the analysis period")


if __name__ == "__main__":
    main()