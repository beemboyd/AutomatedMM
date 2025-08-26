#!/usr/bin/env python3
"""
VSR Performance Analyzer - Weekend Analysis
Analyzes VSR alert efficiency by comparing entry prices with current market prices
"""

import os
import sys
import pandas as pd
from datetime import datetime, timedelta
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from user_context_manager import UserContextManager


def get_vsr_alerts_from_files(days_back=7):
    """Extract VSR alerts from scanner files"""
    base_dir = "/Users/maverick/PycharmProjects/India-TS/Daily/scanners/Hourly"
    alerts = {}
    
    # Calculate date range
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days_back)
    
    print(f"Analyzing VSR alerts from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
    
    # Process each VSR file
    for file in sorted(os.listdir(base_dir)):
        if not file.startswith('VSR_') or not file.endswith('.xlsx'):
            continue
            
        try:
            # Extract date from filename
            date_str = file.replace('VSR_', '').replace('.xlsx', '')
            file_date = datetime.strptime(date_str[:8], '%Y%m%d')
            
            # Check if within our date range
            if file_date < start_date or file_date > end_date:
                continue
                
            # Read the file
            filepath = os.path.join(base_dir, file)
            df = pd.read_excel(filepath)
            
            if df.empty:
                continue
                
            # Extract time from filename
            time_str = date_str[9:11] + ':' + date_str[11:13] if len(date_str) > 8 else '09:30'
            
            # Process each ticker
            for _, row in df.iterrows():
                ticker = row.get('Ticker', '')
                if not ticker:
                    continue
                    
                if ticker not in alerts:
                    alerts[ticker] = {
                        'first_seen': file_date.strftime('%Y-%m-%d'),
                        'first_time': time_str,
                        'entry_price': row.get('Entry_Price', 0),
                        'stop_loss': row.get('Stop_Loss', 0),
                        'target1': row.get('Target1', 0),
                        'target2': row.get('Target2', 0),
                        'pattern': row.get('Pattern', ''),
                        'direction': row.get('Direction', 'LONG'),
                        'vsr_score': row.get('VSR_Score', row.get('Score', 0)),
                        'probability': row.get('Probability_Score', 0),
                        'alert_count': 0,
                        'last_seen': file_date.strftime('%Y-%m-%d'),
                        'all_scores': []
                    }
                
                # Update alert count and scores
                alerts[ticker]['alert_count'] += 1
                alerts[ticker]['last_seen'] = file_date.strftime('%Y-%m-%d')
                score = row.get('VSR_Score', row.get('Score', 0))
                if score:
                    try:
                        alerts[ticker]['all_scores'].append(float(score) if isinstance(score, str) and '/' in score else score)
                    except:
                        pass
                        
        except Exception as e:
            print(f"Error processing {file}: {e}")
            continue
    
    return alerts


def get_current_prices(tickers, user='Sai'):
    """Get current prices from Zerodha"""
    prices = {}
    
    try:
        ucm = UserContextManager()
        kite, config = ucm.get_user_context(user)
        
        if not kite:
            print("Could not initialize Zerodha connection")
            return prices
            
        print(f"Fetching current prices for {len(tickers)} tickers...")
        
        for ticker in tickers:
            try:
                # Try NSE first
                instrument = f"NSE:{ticker}"
                quote = kite.quote(instrument)
                if instrument in quote:
                    prices[ticker] = quote[instrument]['last_price']
                    continue
            except:
                pass
                
            try:
                # Try BSE if NSE fails
                instrument = f"BSE:{ticker}"
                quote = kite.quote(instrument)
                if instrument in quote:
                    prices[ticker] = quote[instrument]['last_price']
            except:
                pass
                
        print(f"Successfully fetched prices for {len(prices)} tickers")
        
    except Exception as e:
        print(f"Error getting prices: {e}")
    
    return prices


def calculate_performance(alerts, current_prices):
    """Calculate performance metrics"""
    performance = []
    
    for ticker, alert_data in alerts.items():
        entry_price = alert_data['entry_price']
        current_price = current_prices.get(ticker, 0)
        
        if not entry_price or not current_price:
            continue
            
        # Calculate returns
        price_change = current_price - entry_price
        price_change_pct = (price_change / entry_price) * 100
        
        # Check targets and stop loss
        sl_hit = current_price <= alert_data['stop_loss'] if alert_data['stop_loss'] else False
        t1_hit = current_price >= alert_data['target1'] if alert_data['target1'] else False
        t2_hit = current_price >= alert_data['target2'] if alert_data['target2'] else False
        
        # Calculate efficiency
        if sl_hit:
            efficiency = 0
        elif t2_hit:
            efficiency = 100
        elif t1_hit:
            efficiency = 75
        elif price_change_pct > 0:
            efficiency = min(50 + price_change_pct * 2, 74)
        else:
            efficiency = max(0, 50 + price_change_pct)
            
        # Calculate average score
        avg_score = sum(alert_data['all_scores']) / len(alert_data['all_scores']) if alert_data['all_scores'] else alert_data['vsr_score']
        
        performance.append({
            'Ticker': ticker,
            'First Seen': alert_data['first_seen'],
            'Entry Price': entry_price,
            'Current Price': current_price,
            'Change %': price_change_pct,
            'SL': alert_data['stop_loss'],
            'T1': alert_data['target1'],
            'T2': alert_data['target2'],
            'SL Hit': 'Yes' if sl_hit else 'No',
            'T1 Hit': 'Yes' if t1_hit else 'No',
            'T2 Hit': 'Yes' if t2_hit else 'No',
            'Pattern': alert_data['pattern'],
            'Alerts': alert_data['alert_count'],
            'Avg Score': avg_score,
            'Efficiency': efficiency
        })
    
    return pd.DataFrame(performance)


def create_analysis_report(df, alerts_data):
    """Create comprehensive Excel report"""
    # Generate filename
    date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = f"/Users/maverick/PycharmProjects/India-TS/Daily/analysis/Efficiency/VSR_Performance_Analysis_{date_str}.xlsx"
    
    wb = Workbook()
    
    # Sheet 1: Performance Summary
    ws1 = wb.active
    ws1.title = "Performance Summary"
    
    # Sort by efficiency
    df_sorted = df.sort_values('Efficiency', ascending=False)
    
    # Add headers
    headers = list(df_sorted.columns)
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for r_idx, row in enumerate(df_sorted.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            
            # Format based on column
            if headers[c_idx-1] in ['Entry Price', 'Current Price', 'SL', 'T1', 'T2']:
                cell.number_format = '#,##0.00'
            elif headers[c_idx-1] == 'Change %':
                cell.number_format = '+#,##0.00%;-#,##0.00%'
                if value > 0:
                    cell.font = Font(color="008000")
                elif value < 0:
                    cell.font = Font(color="FF0000")
            elif headers[c_idx-1] == 'Efficiency':
                cell.number_format = '#,##0.0'
                if value >= 75:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value >= 50:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Sheet 2: Statistics
    ws2 = wb.create_sheet("Statistics")
    
    winners = df[df['Change %'] > 0]
    losers = df[df['Change %'] < 0]
    
    stats = [
        ['VSR PERFORMANCE STATISTICS', ''],
        ['', ''],
        ['Total Tickers', len(df)],
        ['Winners', len(winners)],
        ['Losers', len(losers)],
        ['Win Rate', f"{(len(winners)/len(df)*100):.1f}%" if len(df) > 0 else "0%"],
        ['', ''],
        ['Average Gain (Winners)', f"{winners['Change %'].mean():.2f}%" if len(winners) > 0 else "0%"],
        ['Average Loss (Losers)', f"{losers['Change %'].mean():.2f}%" if len(losers) > 0 else "0%"],
        ['', ''],
        ['Target Achievement', ''],
        ['Target 1 Hit', f"{len(df[df['T1 Hit'] == 'Yes'])} ({len(df[df['T1 Hit'] == 'Yes'])/len(df)*100:.1f}%)"],
        ['Target 2 Hit', f"{len(df[df['T2 Hit'] == 'Yes'])} ({len(df[df['T2 Hit'] == 'Yes'])/len(df)*100:.1f}%)"],
        ['Stop Loss Hit', f"{len(df[df['SL Hit'] == 'Yes'])} ({len(df[df['SL Hit'] == 'Yes'])/len(df)*100:.1f}%)"],
        ['', ''],
        ['Top 5 Winners', ''],
    ]
    
    # Add top winners
    top_winners = df.nlargest(5, 'Change %')
    for _, row in top_winners.iterrows():
        stats.append([row['Ticker'], f"+{row['Change %']:.2f}%"])
    
    stats.append(['', ''])
    stats.append(['Top 5 Losers', ''])
    
    # Add top losers
    top_losers = df.nsmallest(5, 'Change %')
    for _, row in top_losers.iterrows():
        stats.append([row['Ticker'], f"{row['Change %']:.2f}%"])
    
    # Write statistics
    for row_data in stats:
        ws2.append(row_data)
        if row_data[0] in ['VSR PERFORMANCE STATISTICS', 'Target Achievement', 'Top 5 Winners', 'Top 5 Losers']:
            ws2.cell(ws2.max_row, 1).font = Font(bold=True, size=12)
            ws2.cell(ws2.max_row, 1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Adjust column widths for all sheets
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(filepath)
    print(f"\nReport saved to: {filepath}")
    
    return filepath


def print_summary(df):
    """Print analysis summary"""
    print("\n" + "="*80)
    print("VSR WEEKEND PERFORMANCE ANALYSIS")
    print("="*80)
    
    total = len(df)
    winners = len(df[df['Change %'] > 0])
    losers = len(df[df['Change %'] < 0])
    
    print(f"Total Tickers Analyzed: {total}")
    print(f"Winners: {winners} ({winners/total*100:.1f}%)")
    print(f"Losers: {losers} ({losers/total*100:.1f}%)")
    
    if winners > 0:
        avg_gain = df[df['Change %'] > 0]['Change %'].mean()
        print(f"Average Gain: {avg_gain:.2f}%")
        
    if losers > 0:
        avg_loss = df[df['Change %'] < 0]['Change %'].mean()
        print(f"Average Loss: {avg_loss:.2f}%")
    
    print(f"\nTarget Achievement:")
    t1_hit = len(df[df['T1 Hit'] == 'Yes'])
    t2_hit = len(df[df['T2 Hit'] == 'Yes'])
    sl_hit = len(df[df['SL Hit'] == 'Yes'])
    
    print(f"  Target 1 Hit: {t1_hit} ({t1_hit/total*100:.1f}%)")
    print(f"  Target 2 Hit: {t2_hit} ({t2_hit/total*100:.1f}%)")
    print(f"  Stop Loss Hit: {sl_hit} ({sl_hit/total*100:.1f}%)")
    
    print("\nTop 5 Performers:")
    top_5 = df.nlargest(5, 'Change %')
    for _, row in top_5.iterrows():
        print(f"  {row['Ticker']}: +{row['Change %']:.2f}% (Entry: {row['Entry Price']:.2f}, Current: {row['Current Price']:.2f})")
    
    print("\nBottom 5 Performers:")
    bottom_5 = df.nsmallest(5, 'Change %')
    for _, row in bottom_5.iterrows():
        print(f"  {row['Ticker']}: {row['Change %']:.2f}% (Entry: {row['Entry Price']:.2f}, Current: {row['Current Price']:.2f})")
    
    print("="*80)


def main():
    """Main execution"""
    # Get VSR alerts from past week
    alerts = get_vsr_alerts_from_files(days_back=7)
    
    if not alerts:
        print("No VSR alerts found for the analysis period")
        return
        
    print(f"Found {len(alerts)} unique tickers with VSR alerts")
    
    # Get current prices
    current_prices = get_current_prices(list(alerts.keys()))
    
    if not current_prices:
        print("Could not fetch current prices")
        return
        
    # Calculate performance
    performance_df = calculate_performance(alerts, current_prices)
    
    if performance_df.empty:
        print("No performance data to analyze")
        return
        
    # Create report
    report_path = create_analysis_report(performance_df, alerts)
    
    # Print summary
    print_summary(performance_df)
    
    # Update Activity.md
    try:
        activity_file = "/Users/maverick/PycharmProjects/India-TS/Daily/Activity.md"
        
        winners = len(performance_df[performance_df['Change %'] > 0])
        losers = len(performance_df[performance_df['Change %'] < 0])
        win_rate = (winners/len(performance_df)*100) if len(performance_df) > 0 else 0
        
        log_entry = f"""
### {datetime.now().strftime('%Y-%m-%d %H:%M')} IST - [Claude]
**Changes:**
- Ran VSR Weekend Performance Analysis for past week
- Analyzed {len(performance_df)} tickers with VSR alerts
- Fetched current prices from Zerodha API
- Generated comprehensive performance report

**Impact:**
- Win Rate: {win_rate:.1f}% ({winners} winners, {losers} losers)
- Average Gain: {performance_df[performance_df['Change %'] > 0]['Change %'].mean():.2f}% (winners)
- Average Loss: {performance_df[performance_df['Change %'] < 0]['Change %'].mean():.2f}% (losers)
- Report: {report_path}

---
"""
        
        with open(activity_file, 'r') as f:
            content = f.read()
        
        insert_pos = content.find("## Activity Log\n")
        if insert_pos != -1:
            insert_pos = content.find("\n", insert_pos) + 1
            new_content = content[:insert_pos] + log_entry + content[insert_pos:]
            
            with open(activity_file, 'w') as f:
                f.write(new_content)
                
            print("\nActivity.md updated with analysis results")
            
    except Exception as e:
        print(f"Could not update Activity.md: {e}")


if __name__ == "__main__":
    main()