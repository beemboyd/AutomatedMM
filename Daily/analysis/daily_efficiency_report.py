#!/usr/bin/env python3
"""
Daily Efficiency Report

Analyzes Telegram alert performance by comparing:
- First alert price (from Telegram audit logs)
- Current price (from Zerodha API)

This measures how well Telegram alerts perform after being sent to users.

Data Source: Daily/data/audit_vsr.db (telegram_alerts table)
Output: Daily/analysis/Efficiency/Daily_Efficiency_Long_YYYYMMDD.xlsx
"""

import os
import sys
import sqlite3
import logging
import configparser
from datetime import datetime, timedelta
from pathlib import Path
import time

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pytz

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from kiteconnect import KiteConnect

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class DailyEfficiencyReport:
    """
    Generate efficiency report from Telegram alerts.

    Compares first alert price with current price to measure
    how well the alerts performed.
    """

    def __init__(self, lookback_days=10):
        """
        Initialize the report generator.

        Args:
            lookback_days: Number of days to look back for alerts (default: 10)
        """
        self.lookback_days = lookback_days
        self.base_dir = Path("/Users/maverick/PycharmProjects/India-TS/Daily")
        self.db_path = self.base_dir / "data" / "audit_vsr.db"
        self.output_dir = self.base_dir / "analysis" / "Efficiency"
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.ist = pytz.timezone('Asia/Kolkata')
        self.kite = None
        self.instrument_map = {}

        # Rate limiting
        self.api_calls = 0
        self.last_call_time = 0
        self.rate_limit = 3  # calls per second

        # Initialize Kite connection
        self._init_kite()

    def _init_kite(self):
        """Initialize Zerodha Kite connection"""
        try:
            config = configparser.ConfigParser()
            config.read(self.base_dir / 'config.ini')

            # Try Sai's credentials first
            if 'API_CREDENTIALS_Sai' in config:
                api_key = config['API_CREDENTIALS_Sai']['api_key']
                access_token = config['API_CREDENTIALS_Sai']['access_token']
            else:
                api_key = config['DEFAULT']['api_key']
                access_token = config['DEFAULT']['access_token']

            self.kite = KiteConnect(api_key=api_key)
            self.kite.set_access_token(access_token)

            # Load instruments
            instruments = self.kite.instruments("NSE")
            self.instrument_map = {i['tradingsymbol']: i['instrument_token'] for i in instruments}

            logger.info(f"Kite connection established, loaded {len(self.instrument_map)} instruments")

        except Exception as e:
            logger.error(f"Failed to initialize Kite: {e}")
            self.kite = None

    def _throttle(self):
        """Rate limit API calls"""
        current_time = time.time()
        time_since_last = current_time - self.last_call_time
        min_interval = 1.0 / self.rate_limit

        if time_since_last < min_interval:
            time.sleep(min_interval - time_since_last)

        self.last_call_time = time.time()
        self.api_calls += 1

    def get_telegram_alerts(self):
        """
        Get first Telegram alert per ticker in the lookback period.

        Returns:
            List of dicts with ticker, first_alert_time, first_price, score, momentum, etc.
        """
        if not self.db_path.exists():
            logger.error(f"Audit database not found: {self.db_path}")
            return []

        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row

            # Calculate date range
            end_date = datetime.now(self.ist)
            start_date = end_date - timedelta(days=self.lookback_days)
            start_str = start_date.strftime('%Y-%m-%d')

            # Query for first alert per ticker in the period
            # We get the FIRST alert (earliest timestamp) for each ticker
            query = """
                SELECT
                    ticker,
                    MIN(timestamp) as first_alert_time,
                    current_price as first_price,
                    score,
                    momentum,
                    liquidity_grade,
                    alerts_last_30_days,
                    days_tracked,
                    COUNT(*) as alert_count_period
                FROM telegram_alerts
                WHERE DATE(timestamp) >= ?
                GROUP BY ticker
                ORDER BY first_alert_time
            """

            cursor = conn.execute(query, (start_str,))
            rows = cursor.fetchall()

            alerts = []
            for row in rows:
                # Get the first alert's price by querying again
                price_query = """
                    SELECT current_price, score, momentum, liquidity_grade
                    FROM telegram_alerts
                    WHERE ticker = ? AND timestamp = ?
                """
                price_cursor = conn.execute(price_query, (row['ticker'], row['first_alert_time']))
                price_row = price_cursor.fetchone()

                # Helper function to safely convert to float
                def safe_float(val, default=0.0):
                    if val is None:
                        return default
                    if isinstance(val, bytes):
                        try:
                            return float(val.decode('utf-8'))
                        except:
                            return default
                    try:
                        return float(val)
                    except:
                        return default

                # Helper function to safely convert to string
                def safe_str(val, default=''):
                    if val is None:
                        return default
                    if isinstance(val, bytes):
                        return val.decode('utf-8')
                    return str(val)

                # Extract values with proper type handling
                first_price = safe_float(price_row['current_price'] if price_row else row['first_price'])
                score = safe_float(price_row['score'] if price_row else row['score'])
                momentum = safe_float(price_row['momentum'] if price_row else row['momentum'])
                liquidity = safe_str(price_row['liquidity_grade'] if price_row else row['liquidity_grade'])

                alerts.append({
                    'ticker': row['ticker'],
                    'first_alert_time': row['first_alert_time'],
                    'first_price': first_price,
                    'score': score,
                    'momentum': momentum,
                    'liquidity_grade': liquidity,
                    'alert_count': row['alert_count_period']
                })

            conn.close()
            logger.info(f"Found {len(alerts)} unique tickers with Telegram alerts in past {self.lookback_days} days")
            return alerts

        except Exception as e:
            logger.error(f"Error reading telegram alerts: {e}")
            return []

    def get_current_price(self, ticker):
        """
        Get current price for a ticker from Zerodha API.

        Args:
            ticker: Stock symbol

        Returns:
            Current price or None if not available
        """
        if not self.kite or ticker not in self.instrument_map:
            return None

        try:
            self._throttle()
            quote = self.kite.quote(f"NSE:{ticker}")
            ticker_data = quote.get(f"NSE:{ticker}", {})
            return ticker_data.get('last_price')
        except Exception as e:
            logger.warning(f"Error getting price for {ticker}: {e}")
            return None

    def generate_report(self):
        """
        Generate the efficiency report.

        Returns:
            Path to generated Excel file or None if failed
        """
        logger.info(f"Generating Daily Efficiency Report for past {self.lookback_days} days")

        # Get Telegram alerts
        alerts = self.get_telegram_alerts()
        if not alerts:
            logger.warning("No Telegram alerts found for the period")
            return None

        # Enrich with current prices
        report_data = []
        for i, alert in enumerate(alerts):
            ticker = alert['ticker']
            first_price = alert['first_price']

            if not first_price or first_price <= 0:
                logger.warning(f"Skipping {ticker}: no first price available")
                continue

            # Get current price
            current_price = self.get_current_price(ticker)
            if not current_price:
                logger.warning(f"Skipping {ticker}: could not get current price")
                continue

            # Calculate price change (Long position = profit from price increase)
            price_change_pct = ((current_price - first_price) / first_price) * 100

            # Parse alert time
            try:
                alert_dt = datetime.fromisoformat(alert['first_alert_time'].replace('Z', '+00:00'))
                alert_date = alert_dt.strftime('%Y-%m-%d')
                alert_time = alert_dt.strftime('%H:%M:%S')
            except:
                alert_date = alert['first_alert_time'][:10] if alert['first_alert_time'] else ''
                alert_time = alert['first_alert_time'][11:19] if alert['first_alert_time'] else ''

            report_data.append({
                'Ticker': ticker,
                'First Alert Date': alert_date,
                'First Alert Time': alert_time,
                'First Price': round(first_price, 2),
                'Current Price': round(current_price, 2),
                'Price Change %': round(price_change_pct, 2),
                'Score': alert['score'],
                'Momentum %': round(alert['momentum'], 2) if alert['momentum'] else 0,
                'Liquidity': alert['liquidity_grade'] or '',
                'Alert Count': alert['alert_count']
            })

            # Progress logging
            if (i + 1) % 10 == 0:
                logger.info(f"Processed {i + 1}/{len(alerts)} tickers")

        if not report_data:
            logger.warning("No valid data to report")
            return None

        # Create DataFrame
        df = pd.DataFrame(report_data)
        df = df.sort_values('Price Change %', ascending=False)

        # Generate filename
        today_str = datetime.now(self.ist).strftime('%Y%m%d')
        start_str = (datetime.now(self.ist) - timedelta(days=self.lookback_days)).strftime('%Y%m%d')
        filename = f"Daily_Efficiency_Long_{today_str}_{start_str}.xlsx"
        filepath = self.output_dir / filename

        # Create Excel with formatting
        self._create_excel(df, filepath)

        logger.info(f"Report generated: {filepath}")
        return filepath

    def _create_excel(self, df, filepath):
        """
        Create formatted Excel report.

        Args:
            df: DataFrame with report data
            filepath: Output file path
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Long Efficiency"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        # Add headers
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Add data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border

                # Format numbers
                header_name = headers[c_idx - 1]
                if 'Price' in header_name and isinstance(value, (int, float)):
                    cell.number_format = '₹#,##0.00'
                elif 'Change %' in header_name or 'Momentum' in header_name:
                    cell.number_format = '+#,##0.00;-#,##0.00'
                    if isinstance(value, (int, float)):
                        if value > 0:
                            cell.font = Font(color="008000", bold=True)  # Green
                        elif value < 0:
                            cell.font = Font(color="FF0000")  # Red
                elif 'Score' in header_name and isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
                    if value >= 80:
                        cell.font = Font(color="008000", bold=True)
                    elif value >= 60:
                        cell.font = Font(color="0066CC")

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add summary section
        ws.append([])
        ws.append([])

        # Summary header
        summary_row = ws.max_row + 1
        ws.cell(row=summary_row, column=1, value="SUMMARY STATISTICS").font = Font(bold=True, size=12)

        # Calculate stats
        total_tickers = len(df)
        avg_change = df['Price Change %'].mean()
        median_change = df['Price Change %'].median()
        winners = len(df[df['Price Change %'] > 0])
        losers = len(df[df['Price Change %'] < 0])
        win_rate = (winners / total_tickers * 100) if total_tickers > 0 else 0
        best_performer = df.iloc[0]['Ticker'] if len(df) > 0 else 'N/A'
        best_gain = df.iloc[0]['Price Change %'] if len(df) > 0 else 0
        worst_performer = df.iloc[-1]['Ticker'] if len(df) > 0 else 'N/A'
        worst_loss = df.iloc[-1]['Price Change %'] if len(df) > 0 else 0

        stats = [
            ['Total Tickers', total_tickers],
            ['Average Change %', f"{avg_change:.2f}%"],
            ['Median Change %', f"{median_change:.2f}%"],
            ['Winners (>0%)', winners],
            ['Losers (<0%)', losers],
            ['Win Rate', f"{win_rate:.1f}%"],
            ['Best Performer', f"{best_performer} (+{best_gain:.2f}%)"],
            ['Worst Performer', f"{worst_performer} ({worst_loss:.2f}%)"],
            ['Report Generated', datetime.now(self.ist).strftime('%Y-%m-%d %H:%M:%S IST')],
            ['Lookback Period', f"{self.lookback_days} days"],
        ]

        for stat in stats:
            ws.append(stat)

        # Save
        wb.save(filepath)
        logger.info(f"Excel file saved: {filepath}")


def main():
    """Main entry point"""
    import argparse

    parser = argparse.ArgumentParser(description='Daily Efficiency Report from Telegram Alerts')
    parser.add_argument('--days', type=int, default=10,
                       help='Number of days to look back (default: 10)')
    args = parser.parse_args()

    report = DailyEfficiencyReport(lookback_days=args.days)
    filepath = report.generate_report()

    if filepath:
        print(f"\n✅ Report generated successfully!")
        print(f"📁 File: {filepath}")
    else:
        print("\n❌ Failed to generate report")
        sys.exit(1)


if __name__ == "__main__":
    main()
