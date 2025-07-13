#!/usr/bin/env python
"""
Regime Reporter Module
=====================
Generates reports and visualizations for market regime analysis.
"""

import os
import json
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import logging
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class RegimeReporter:
    """Generate reports and visualizations for market regime analysis"""
    
    def __init__(self, base_dir: str = None):
        """Initialize reporter"""
        if base_dir is None:
            base_dir = "/Users/maverick/PycharmProjects/India-TS/Daily"
            
        self.base_dir = base_dir
        self.reports_dir = os.path.join(base_dir, "reports", "market_regime")
        os.makedirs(self.reports_dir, exist_ok=True)
        
        # Color schemes for regimes
        self.regime_colors = {
            'STRONG_BULL': '#00b359',  # Dark green
            'BULL': '#66cc99',         # Light green
            'NEUTRAL': '#ffcc00',      # Yellow
            'BEAR': '#ff6666',         # Light red
            'STRONG_BEAR': '#cc0000',  # Dark red
            'VOLATILE': '#9966ff',     # Purple
            'UNKNOWN': '#999999'       # Gray
        }
        
    def generate_daily_report(self, 
                            regime_data: Dict,
                            recommendations: Dict,
                            save_format: List[str] = ['text', 'excel']) -> List[str]:
        """
        Generate daily regime report
        
        Args:
            regime_data: Current regime and indicators
            recommendations: Regime-based recommendations
            save_format: List of formats to save ('text', 'excel', 'html')
            
        Returns:
            List of saved file paths
        """
        saved_files = []
        timestamp = datetime.now()
        
        if 'text' in save_format:
            text_file = self._generate_text_report(regime_data, recommendations, timestamp)
            saved_files.append(text_file)
            
        if 'excel' in save_format:
            excel_file = self._generate_excel_report(regime_data, recommendations, timestamp)
            saved_files.append(excel_file)
            
        if 'html' in save_format:
            html_file = self._generate_html_report(regime_data, recommendations, timestamp)
            saved_files.append(html_file)
            
        return saved_files
        
    def _generate_text_report(self, regime_data: Dict, recommendations: Dict, 
                            timestamp: datetime) -> str:
        """Generate text format report"""
        filename = os.path.join(self.reports_dir, 
                              f"regime_report_{timestamp.strftime('%Y%m%d_%H%M%S')}.txt")
        
        with open(filename, 'w') as f:
            # Header
            f.write("="*80 + "\n")
            f.write(f"MARKET REGIME ANALYSIS REPORT\n")
            f.write(f"Generated: {timestamp.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("="*80 + "\n\n")
            
            # Current Regime
            regime = recommendations.get('regime', 'UNKNOWN')
            confidence = recommendations.get('confidence', 0)
            
            f.write(f"CURRENT MARKET REGIME: {regime}\n")
            f.write(f"Confidence: {confidence:.1%}\n")
            f.write(f"Risk Level: {recommendations.get('risk_level', 'UNKNOWN')}\n\n")
            
            # Key Indicators
            f.write("KEY MARKET INDICATORS\n")
            f.write("-"*40 + "\n")
            
            indicators = regime_data.get('indicators', {})
            
            # Breadth
            breadth = indicators.get('breadth', {})
            f.write(f"Market Breadth:\n")
            f.write(f"  - Bullish %: {breadth.get('bullish_percent', 0)*100:.1f}%\n")
            f.write(f"  - A/D Ratio: {breadth.get('advance_decline_ratio', 0):.2f}\n")
            f.write(f"  - High Volume %: {breadth.get('high_volume_percent', 0)*100:.1f}%\n\n")
            
            # Momentum
            momentum = indicators.get('momentum', {})
            f.write(f"Market Momentum:\n")
            f.write(f"  - Average: {momentum.get('average_momentum', 0):.2f}%\n")
            f.write(f"  - Extreme Positive: {momentum.get('extreme_positive', 0)*100:.1f}%\n")
            f.write(f"  - Extreme Negative: {momentum.get('extreme_negative', 0)*100:.1f}%\n\n")
            
            # Volatility
            volatility = indicators.get('volatility', {})
            f.write(f"Market Volatility:\n")
            f.write(f"  - Average Range: {volatility.get('average_range', 0):.2f}%\n")
            f.write(f"  - Reversal Patterns: {volatility.get('reversal_pattern_percent', 0)*100:.1f}%\n\n")
            
            # Composite Scores
            composite = indicators.get('composite', {})
            f.write(f"Composite Indicators:\n")
            f.write(f"  - Market Strength: {composite.get('market_strength_index', 0):.1f}/100\n")
            f.write(f"  - Volatility Index: {composite.get('volatility_index', 0):.1f}/100\n")
            f.write(f"  - Market Risk Score: {composite.get('market_risk_score', 0):.1f}/100\n\n")
            
            # Position Sizing Recommendations
            f.write("POSITION SIZING RECOMMENDATIONS\n")
            f.write("-"*40 + "\n")
            
            sizing = recommendations.get('position_sizing', {})
            f.write(f"Position Size Multiplier: {sizing.get('size_multiplier', 1.0):.1f}x\n")
            f.write(f"Max Portfolio Exposure: {sizing.get('max_portfolio_exposure', 0.8)*100:.0f}%\n")
            f.write(f"Stop Loss Multiplier: {sizing.get('stop_loss_multiplier', 1.5):.1f}x\n\n")
            
            # Preferred Sectors
            sectors = recommendations.get('preferred_sectors', [])
            if sectors:
                f.write(f"Preferred Sectors: {', '.join(sectors)}\n\n")
                
            # Action Items
            f.write("ACTION ITEMS\n")
            f.write("-"*40 + "\n")
            for i, action in enumerate(recommendations.get('action_items', []), 1):
                f.write(f"{i}. {action}\n")
                
            # Alerts
            alerts = recommendations.get('alerts', [])
            if alerts:
                f.write("\n\nALERTS\n")
                f.write("-"*40 + "\n")
                for alert in alerts:
                    f.write(f"[{alert['level']}] {alert['message']}\n")
                    
            # Regime Change Signals
            regime_change = recommendations.get('regime_change_alert', {})
            if regime_change.get('regime_change_detected'):
                f.write("\n\nREGIME CHANGE DETECTED\n")
                f.write("-"*40 + "\n")
                f.write(f"Transition Type: {regime_change.get('transition_type')}\n")
                f.write(f"Confidence: {regime_change.get('confidence', 0):.1%}\n")
                f.write("\nSignals:\n")
                for signal in regime_change.get('signals', []):
                    f.write(f"  - {signal.get('description', 'Unknown signal')}\n")
                    
        logger.info(f"Text report saved to: {filename}")
        return filename
        
    def _generate_excel_report(self, regime_data: Dict, recommendations: Dict,
                             timestamp: datetime) -> str:
        """Generate Excel format report"""
        filename = os.path.join(self.reports_dir,
                              f"regime_report_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._format_summary_sheet(ws_summary, regime_data, recommendations)
        
        # Indicators Sheet
        ws_indicators = wb.create_sheet("Indicators")
        self._format_indicators_sheet(ws_indicators, regime_data.get('indicators', {}))
        
        # Sectors Sheet
        sectors_data = regime_data.get('indicators', {}).get('sectors', {})
        if sectors_data:
            ws_sectors = wb.create_sheet("Sectors")
            self._format_sectors_sheet(ws_sectors, sectors_data)
            
        # Save workbook
        wb.save(filename)
        logger.info(f"Excel report saved to: {filename}")
        return filename
        
    def _format_summary_sheet(self, ws, regime_data: Dict, recommendations: Dict):
        """Format summary sheet in Excel"""
        # Title
        ws['A1'] = "Market Regime Analysis Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Current Regime
        regime = recommendations.get('regime', 'UNKNOWN')
        confidence = recommendations.get('confidence', 0)
        
        ws['A4'] = "Current Market Regime:"
        ws['B4'] = regime
        ws['B4'].font = Font(bold=True, color=self._get_regime_color_hex(regime))
        
        ws['A5'] = "Confidence:"
        ws['B5'] = f"{confidence:.1%}"
        
        ws['A6'] = "Risk Level:"
        ws['B6'] = recommendations.get('risk_level', 'UNKNOWN')
        
        # Key Metrics
        ws['A8'] = "Key Market Metrics"
        ws['A8'].font = Font(bold=True)
        
        indicators = regime_data.get('indicators', {})
        composite = indicators.get('composite', {})
        
        ws['A9'] = "Market Strength Index:"
        ws['B9'] = f"{composite.get('market_strength_index', 0):.1f}"
        
        ws['A10'] = "Volatility Index:"
        ws['B10'] = f"{composite.get('volatility_index', 0):.1f}"
        
        ws['A11'] = "Market Risk Score:"
        ws['B11'] = f"{composite.get('market_risk_score', 0):.1f}"
        
        # Position Sizing
        ws['A13'] = "Position Sizing Recommendations"
        ws['A13'].font = Font(bold=True)
        
        sizing = recommendations.get('position_sizing', {})
        ws['A14'] = "Size Multiplier:"
        ws['B14'] = f"{sizing.get('size_multiplier', 1.0):.1f}x"
        
        ws['A15'] = "Max Exposure:"
        ws['B15'] = f"{sizing.get('max_portfolio_exposure', 0.8)*100:.0f}%"
        
        ws['A16'] = "Stop Loss Multiplier:"
        ws['B16'] = f"{sizing.get('stop_loss_multiplier', 1.5):.1f}x"
        
        # Action Items
        ws['A18'] = "Action Items"
        ws['A18'].font = Font(bold=True)
        
        row = 19
        for action in recommendations.get('action_items', []):
            ws[f'A{row}'] = f"• {action}"
            row += 1
            
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
    def _format_indicators_sheet(self, ws, indicators: Dict):
        """Format indicators sheet in Excel"""
        ws['A1'] = "Market Indicators Detail"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        # Breadth Indicators
        ws[f'A{row}'] = "Breadth Indicators"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        breadth = indicators.get('breadth', {})
        for key, value in breadth.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = f"{value:.3f}" if isinstance(value, float) else str(value)
            row += 1
            
        row += 1
        
        # Momentum Indicators
        ws[f'A{row}'] = "Momentum Indicators"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        momentum = indicators.get('momentum', {})
        for key, value in momentum.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = f"{value:.3f}" if isinstance(value, float) else str(value)
            row += 1
            
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
    def _format_sectors_sheet(self, ws, sectors_data: Dict):
        """Format sectors sheet in Excel"""
        ws['A1'] = "Sector Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Top Sectors
        if 'top_sectors' in sectors_data:
            ws['A3'] = "Top Performing Sectors"
            ws['A3'].font = Font(bold=True)
            
            row = 4
            ws[f'A{row}'] = "Sector"
            ws[f'B{row}'] = "Avg Momentum"
            ws[f'C{row}'] = "Count"
            
            row += 1
            top_sectors = sectors_data['top_sectors']
            if isinstance(top_sectors, dict) and 'mean' in top_sectors:
                for sector, momentum in top_sectors['mean'].items():
                    ws[f'A{row}'] = sector
                    ws[f'B{row}'] = f"{momentum:.2f}%"
                    ws[f'C{row}'] = top_sectors.get('count', {}).get(sector, 0)
                    row += 1
                    
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        
    def _generate_html_report(self, regime_data: Dict, recommendations: Dict,
                            timestamp: datetime) -> str:
        """Generate HTML format report"""
        filename = os.path.join(self.reports_dir,
                              f"regime_report_{timestamp.strftime('%Y%m%d_%H%M%S')}.html")
        
        regime = recommendations.get('regime', 'UNKNOWN')
        confidence = recommendations.get('confidence', 0)
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Market Regime Analysis Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1, h2, h3 {{ color: #333; }}
                .regime-badge {{
                    display: inline-block;
                    padding: 5px 15px;
                    border-radius: 5px;
                    color: white;
                    font-weight: bold;
                    background-color: {self.regime_colors.get(regime, '#999')};
                }}
                .metric-box {{
                    display: inline-block;
                    margin: 10px;
                    padding: 15px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    background-color: #f5f5f5;
                }}
                .alert {{ 
                    padding: 10px;
                    margin: 10px 0;
                    border-radius: 5px;
                }}
                .alert-high {{ background-color: #ffcccc; }}
                .alert-medium {{ background-color: #ffffcc; }}
                .alert-low {{ background-color: #ccffcc; }}
                table {{ 
                    border-collapse: collapse;
                    width: 100%;
                    margin: 20px 0;
                }}
                th, td {{
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }}
                th {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <h1>Market Regime Analysis Report</h1>
            <p>Generated: {timestamp.strftime('%Y-%m-%d %H:%M:%S')}</p>
            
            <h2>Current Market Regime</h2>
            <p><span class="regime-badge">{regime}</span></p>
            <p>Confidence: {confidence:.1%}</p>
            <p>Risk Level: {recommendations.get('risk_level', 'UNKNOWN')}</p>
        """
        
        # Add key metrics
        indicators = regime_data.get('indicators', {})
        composite = indicators.get('composite', {})
        
        html_content += """
            <h2>Key Market Metrics</h2>
            <div>
        """
        
        for metric, value in composite.items():
            if isinstance(value, (int, float)):
                html_content += f"""
                <div class="metric-box">
                    <strong>{metric.replace('_', ' ').title()}</strong><br>
                    {value:.1f}
                </div>
                """
                
        html_content += "</div>"
        
        # Add recommendations
        html_content += """
            <h2>Recommendations</h2>
            <h3>Position Sizing</h3>
            <table>
                <tr>
                    <th>Parameter</th>
                    <th>Value</th>
                </tr>
        """
        
        sizing = recommendations.get('position_sizing', {})
        html_content += f"""
                <tr>
                    <td>Size Multiplier</td>
                    <td>{sizing.get('size_multiplier', 1.0):.1f}x</td>
                </tr>
                <tr>
                    <td>Max Portfolio Exposure</td>
                    <td>{sizing.get('max_portfolio_exposure', 0.8)*100:.0f}%</td>
                </tr>
                <tr>
                    <td>Stop Loss Multiplier</td>
                    <td>{sizing.get('stop_loss_multiplier', 1.5):.1f}x</td>
                </tr>
            </table>
        """
        
        # Add action items
        html_content += "<h3>Action Items</h3><ul>"
        for action in recommendations.get('action_items', []):
            html_content += f"<li>{action}</li>"
        html_content += "</ul>"
        
        # Add alerts if any
        alerts = recommendations.get('alerts', [])
        if alerts:
            html_content += "<h2>Alerts</h2>"
            for alert in alerts:
                level_class = f"alert-{alert['level'].lower()}"
                html_content += f"""
                <div class="alert {level_class}">
                    <strong>[{alert['level']}]</strong> {alert['message']}
                </div>
                """
                
        html_content += """
        </body>
        </html>
        """
        
        with open(filename, 'w') as f:
            f.write(html_content)
            
        logger.info(f"HTML report saved to: {filename}")
        return filename
        
    def _get_regime_color_hex(self, regime: str) -> str:
        """Get color hex code for regime"""
        color = self.regime_colors.get(regime, '#999999')
        return color.replace('#', '')
        
    def create_regime_visualization(self, 
                                  historical_regimes: pd.DataFrame,
                                  save_path: str = None) -> str:
        """
        Create visualization of regime history
        
        Args:
            historical_regimes: DataFrame with regime history
            save_path: Optional path to save figure
            
        Returns:
            Path to saved figure
        """
        if save_path is None:
            save_path = os.path.join(self.reports_dir,
                                   f"regime_history_{datetime.now().strftime('%Y%m%d')}.png")
            
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 8), 
                                       gridspec_kw={'height_ratios': [2, 1]})
        
        # Regime timeline
        if not historical_regimes.empty:
            regimes = historical_regimes['regime'].unique()
            regime_map = {r: i for i, r in enumerate(regimes)}
            
            historical_regimes['regime_num'] = historical_regimes['regime'].map(regime_map)
            
            # Plot regime bars
            for i, regime in enumerate(regimes):
                regime_data = historical_regimes[historical_regimes['regime'] == regime]
                if not regime_data.empty:
                    ax1.scatter(regime_data['timestamp'], 
                               regime_data['regime_num'],
                               c=self.regime_colors.get(regime, '#999999'),
                               label=regime, s=100, alpha=0.7)
                    
            ax1.set_yticks(range(len(regimes)))
            ax1.set_yticklabels(regimes)
            ax1.set_xlabel('Date')
            ax1.set_title('Market Regime History')
            ax1.legend()
            ax1.grid(True, alpha=0.3)
            
            # Confidence timeline
            ax2.plot(historical_regimes['timestamp'], 
                    historical_regimes['confidence'],
                    color='blue', linewidth=2)
            ax2.fill_between(historical_regimes['timestamp'],
                            historical_regimes['confidence'],
                            alpha=0.3)
            ax2.set_ylabel('Confidence')
            ax2.set_xlabel('Date')
            ax2.set_title('Regime Detection Confidence')
            ax2.set_ylim(0, 1)
            ax2.grid(True, alpha=0.3)
            
        plt.tight_layout()
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        logger.info(f"Visualization saved to: {save_path}")
        return save_path
        
    def generate_summary_dashboard(self, 
                                 regime_data: Dict,
                                 recommendations: Dict,
                                 historical_data: pd.DataFrame = None) -> str:
        """Generate comprehensive dashboard with all visualizations"""
        timestamp = datetime.now()
        dashboard_path = os.path.join(self.reports_dir,
                                    f"regime_dashboard_{timestamp.strftime('%Y%m%d_%H%M%S')}.png")
        
        fig = plt.figure(figsize=(16, 10))
        
        # Create grid
        gs = fig.add_gridspec(3, 3, hspace=0.3, wspace=0.3)
        
        # Current regime display
        ax1 = fig.add_subplot(gs[0, :])
        self._plot_current_regime(ax1, recommendations)
        
        # Market indicators
        ax2 = fig.add_subplot(gs[1, 0])
        self._plot_market_strength(ax2, regime_data.get('indicators', {}))
        
        ax3 = fig.add_subplot(gs[1, 1])
        self._plot_breadth_indicators(ax3, regime_data.get('indicators', {}))
        
        ax4 = fig.add_subplot(gs[1, 2])
        self._plot_volatility_gauge(ax4, regime_data.get('indicators', {}))
        
        # Historical regime if available
        if historical_data is not None and not historical_data.empty:
            ax5 = fig.add_subplot(gs[2, :])
            self._plot_regime_history_compact(ax5, historical_data)
            
        plt.suptitle(f"Market Regime Dashboard - {timestamp.strftime('%Y-%m-%d %H:%M')}", 
                    fontsize=16)
        
        plt.savefig(dashboard_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        logger.info(f"Dashboard saved to: {dashboard_path}")
        return dashboard_path
        
    def _plot_current_regime(self, ax, recommendations: Dict):
        """Plot current regime display"""
        regime = recommendations.get('regime', 'UNKNOWN')
        confidence = recommendations.get('confidence', 0)
        
        # Clear axis
        ax.clear()
        ax.set_xlim(0, 10)
        ax.set_ylim(0, 1)
        
        # Regime box
        regime_color = self.regime_colors.get(regime, '#999999')
        rect = plt.Rectangle((1, 0.3), 8, 0.4, 
                           facecolor=regime_color, 
                           edgecolor='black',
                           linewidth=2)
        ax.add_patch(rect)
        
        # Text
        ax.text(5, 0.5, regime, 
               ha='center', va='center',
               fontsize=24, fontweight='bold',
               color='white')
        
        ax.text(5, 0.15, f'Confidence: {confidence:.1%}',
               ha='center', va='center',
               fontsize=14)
        
        ax.axis('off')
        ax.set_title('Current Market Regime', fontsize=16)
        
    def _plot_market_strength(self, ax, indicators: Dict):
        """Plot market strength gauge"""
        composite = indicators.get('composite', {})
        strength = composite.get('market_strength_index', 50)
        
        # Create gauge
        ax.clear()
        
        # Draw arc
        theta = np.linspace(0, np.pi, 100)
        r = 1
        x = r * np.cos(theta)
        y = r * np.sin(theta)
        
        # Color gradient
        colors = ['red', 'orange', 'yellow', 'lightgreen', 'green']
        for i in range(len(colors)):
            start = i * 20
            end = (i + 1) * 20
            mask = (theta >= np.pi * (100 - end) / 100) & (theta <= np.pi * (100 - start) / 100)
            ax.fill_between(x[mask], 0, y[mask], color=colors[i], alpha=0.3)
            
        # Needle
        angle = np.pi * (100 - strength) / 100
        ax.plot([0, 0.8 * np.cos(angle)], [0, 0.8 * np.sin(angle)], 
               'k-', linewidth=3)
        ax.plot(0, 0, 'ko', markersize=10)
        
        # Labels
        ax.text(0, -0.3, f'{strength:.0f}', ha='center', fontsize=16, fontweight='bold')
        
        ax.set_xlim(-1.2, 1.2)
        ax.set_ylim(-0.5, 1.2)
        ax.axis('off')
        ax.set_title('Market Strength', fontsize=12)
        
    def _plot_breadth_indicators(self, ax, indicators: Dict):
        """Plot breadth indicators"""
        breadth = indicators.get('breadth', {})
        
        labels = ['Bullish %', 'Bearish %', 'Neutral %']
        sizes = [
            breadth.get('bullish_percent', 0.33) * 100,
            breadth.get('bearish_percent', 0.33) * 100,
            100 - breadth.get('bullish_percent', 0.33) * 100 - breadth.get('bearish_percent', 0.33) * 100
        ]
        colors = ['green', 'red', 'gray']
        
        ax.clear()
        ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
        ax.set_title('Market Breadth', fontsize=12)
        
    def _plot_volatility_gauge(self, ax, indicators: Dict):
        """Plot volatility gauge"""
        volatility = indicators.get('composite', {}).get('volatility_index', 50)
        
        ax.clear()
        
        # Create horizontal bar
        ax.barh([0], [volatility], color='purple', alpha=0.7)
        ax.barh([0], [100], color='lightgray', alpha=0.3)
        
        # Add threshold lines
        ax.axvline(30, color='green', linestyle='--', alpha=0.5)
        ax.axvline(50, color='orange', linestyle='--', alpha=0.5)
        ax.axvline(70, color='red', linestyle='--', alpha=0.5)
        
        ax.set_xlim(0, 100)
        ax.set_ylim(-0.5, 0.5)
        ax.set_xticks([0, 30, 50, 70, 100])
        ax.set_xticklabels(['0', 'Low', 'Med', 'High', '100'])
        ax.set_yticks([])
        ax.set_title(f'Volatility Index: {volatility:.0f}', fontsize=12)
        
    def _plot_regime_history_compact(self, ax, historical_data: pd.DataFrame):
        """Plot compact regime history"""
        # Group by regime for cleaner display
        ax.clear()
        
        # Create regime blocks
        regimes = []
        current_regime = None
        start_idx = 0
        
        for idx, row in historical_data.iterrows():
            if row['regime'] != current_regime:
                if current_regime is not None:
                    regimes.append({
                        'regime': current_regime,
                        'start': historical_data.iloc[start_idx]['timestamp'],
                        'end': historical_data.iloc[idx-1]['timestamp']
                    })
                current_regime = row['regime']
                start_idx = idx
                
        # Add last regime
        if current_regime is not None:
            regimes.append({
                'regime': current_regime,
                'start': historical_data.iloc[start_idx]['timestamp'],
                'end': historical_data.iloc[-1]['timestamp']
            })
            
        # Plot regime blocks
        for i, regime_block in enumerate(regimes):
            color = self.regime_colors.get(regime_block['regime'], '#999999')
            ax.axvspan(regime_block['start'], regime_block['end'],
                      color=color, alpha=0.5)
                      
        ax.set_xlabel('Date')
        ax.set_title('Regime History', fontsize=12)
        ax.grid(True, alpha=0.3)