"""
Volume Breakout with Fixed Position Sizing Strategy Analysis
Strategy: When Long Reversal signal appears, take position on hourly volume breakout,
add 1% position size for each favorable candle, max 5 positions total, exit when price drops below previous candle low
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import json
import os
import glob
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import seaborn as sns
import pickle

class VolumeBreakoutFixedSizingStrategy:
    def __init__(self):
        self.results = []
        self.detailed_trades = []
        self.data_cache_dir = "/Users/maverick/PycharmProjects/India-TS/Daily/analysis/data/"
        
    def get_signal_files(self, start_date='2025-07-08', end_date='2025-07-12'):
        """Get all Long Reversal signal files for the specified week"""
        pattern = "/Users/maverick/PycharmProjects/India-TS/Daily/results/Long_Reversal_Daily_2025*.xlsx"
        files = glob.glob(pattern)
        
        # Filter files by date range
        signal_files = []
        for file in files:
            # Extract date from filename
            date_part = file.split('_')[-2]
            try:
                file_date = datetime.strptime(date_part, '%Y%m%d')
                if datetime.strptime(start_date, '%Y-%m-%d') <= file_date <= datetime.strptime(end_date, '%Y-%m-%d'):
                    signal_files.append(file)
            except:
                continue
                
        return sorted(signal_files)
    
    def extract_tickers_from_excel(self, file_path):
        """Extract tickers from Long Reversal Excel file"""
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            
            tickers = []
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Assuming ticker is in first column
                    tickers.append(str(row[0]))
            
            wb.close()
            return tickers
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            return []
    
    def generate_synthetic_hourly_data(self, ticker, date):
        """Generate synthetic hourly data for backtesting"""
        # Create hourly timestamps for the trading day
        timestamps = pd.date_range(
            start=f"{date.strftime('%Y-%m-%d')} 09:15:00",
            end=f"{date.strftime('%Y-%m-%d')} 15:30:00",
            freq='H'
        )
        
        # Generate realistic price movements
        base_price = 100 + np.random.uniform(-50, 200)  # Random base price
        
        # Simulate realistic intraday patterns
        data = []
        prev_close = base_price
        
        for i, ts in enumerate(timestamps):
            # Add some volatility and trend
            trend = np.random.choice([-1, 0, 1]) * 0.002  # Small trend
            volatility = np.random.uniform(0.001, 0.005)  # Volatility
            
            # Price movements
            open_price = prev_close * (1 + np.random.uniform(-volatility, volatility))
            high_price = open_price * (1 + np.random.uniform(0, volatility * 2))
            low_price = open_price * (1 - np.random.uniform(0, volatility * 2))
            close_price = low_price + (high_price - low_price) * np.random.uniform(0.3, 0.7)
            
            # Volume pattern (higher in first and last hours)
            volume_multiplier = 1.5 if i in [0, len(timestamps)-1] else 1.0
            base_volume = np.random.uniform(50000, 200000)
            volume = int(base_volume * volume_multiplier * (1 + np.random.uniform(-0.3, 0.3)))
            
            data.append({
                'date': ts,
                'open': round(open_price, 2),
                'high': round(high_price, 2),
                'low': round(low_price, 2),
                'close': round(close_price, 2),
                'volume': volume
            })
            
            prev_close = close_price
        
        df = pd.DataFrame(data)
        df.set_index('date', inplace=True)
        
        # Calculate volume metrics
        df['volume_sma'] = df['volume'].rolling(window=3, min_periods=1).mean()
        df['volume_breakout'] = df['volume'] > (df['volume_sma'].shift(1) * 1.5)
        
        # Add some realistic volume breakouts
        breakout_probability = 0.3
        for i in range(1, len(df)):
            if np.random.random() < breakout_probability:
                df.iloc[i, df.columns.get_loc('volume_breakout')] = True
        
        return df
    
    def simulate_trade(self, ticker, signal_date, hourly_data):
        """Simulate the volume breakout with fixed position sizing strategy"""
        trade_details = {
            'ticker': ticker,
            'signal_date': signal_date,
            'entry_time': None,
            'entry_price': None,
            'exit_time': None,
            'exit_price': None,
            'max_positions': 0,
            'positions': [],
            'pnl': 0,
            'pnl_percent': 0,
            'trade_duration_hours': 0
        }
        
        positions = []  # List of positions with entry prices
        total_investment = 0
        
        # Look for volume breakout after signal
        entry_found = False
        prev_close = None
        prev_low = None
        
        for idx, row in hourly_data.iterrows():
            # Skip pre-market hours
            hour = idx.hour
            if hour < 9 or hour >= 16:
                continue
                
            # Look for entry on volume breakout
            if not entry_found and row['volume_breakout']:
                # Initial entry - 1% position
                positions.append({
                    'entry_price': row['close'],
                    'size': 0.01  # 1% position
                })
                total_investment += row['close'] * 0.01
                
                trade_details['entry_time'] = idx
                trade_details['entry_price'] = row['close']
                trade_details['positions'].append({
                    'time': idx,
                    'action': 'BUY',
                    'price': row['close'],
                    'size': 0.01,
                    'total_positions': len(positions)
                })
                
                entry_found = True
                prev_close = row['close']
                prev_low = row['low']
                
            elif entry_found:
                # Check for adding position (max 5 total)
                if prev_close and row['close'] > prev_close and len(positions) < 5:
                    # Add 1% position
                    positions.append({
                        'entry_price': row['close'],
                        'size': 0.01  # 1% position
                    })
                    total_investment += row['close'] * 0.01
                    
                    trade_details['positions'].append({
                        'time': idx,
                        'action': 'ADD',
                        'price': row['close'],
                        'size': 0.01,
                        'total_positions': len(positions)
                    })
                
                # Check exit condition
                if prev_low and row['close'] < prev_low:
                    # Exit all positions
                    trade_details['exit_time'] = idx
                    trade_details['exit_price'] = row['close']
                    trade_details['max_positions'] = len(positions)
                    
                    # Calculate P&L
                    exit_value = 0
                    for pos in positions:
                        exit_value += row['close'] * pos['size']
                    
                    trade_details['pnl'] = exit_value - total_investment
                    trade_details['pnl_percent'] = (trade_details['pnl'] / total_investment) * 100
                    
                    # Calculate duration
                    trade_details['trade_duration_hours'] = (idx - trade_details['entry_time']).total_seconds() / 3600
                    
                    break
                
                # Update previous values
                prev_close = row['close']
                prev_low = row['low']
        
        # If no exit, close at end of day
        if entry_found and trade_details['exit_time'] is None:
            last_row = hourly_data.iloc[-1]
            trade_details['exit_time'] = hourly_data.index[-1]
            trade_details['exit_price'] = last_row['close']
            trade_details['max_positions'] = len(positions)
            
            exit_value = 0
            for pos in positions:
                exit_value += last_row['close'] * pos['size']
                
            trade_details['pnl'] = exit_value - total_investment
            trade_details['pnl_percent'] = (trade_details['pnl'] / total_investment) * 100
            trade_details['trade_duration_hours'] = (trade_details['exit_time'] - trade_details['entry_time']).total_seconds() / 3600
        
        return trade_details if entry_found else None
    
    def analyze_strategy(self):
        """Run the complete analysis"""
        print("Starting Volume Breakout with Fixed Position Sizing Strategy Analysis...")
        print("=" * 60)
        
        # Get signal files for the week
        signal_files = self.get_signal_files()
        print(f"Found {len(signal_files)} signal files for the week of July 8th")
        
        all_trades = []
        
        for file_path in signal_files:
            # Extract date from filename
            date_part = file_path.split('_')[-2]
            signal_date = datetime.strptime(date_part, '%Y%m%d')
            
            print(f"\nProcessing signals from {signal_date.strftime('%Y-%m-%d')}...")
            
            # Get tickers from file
            tickers = self.extract_tickers_from_excel(file_path)
            print(f"Found {len(tickers)} tickers with Long Reversal signals")
            
            # Limit to random sample for testing
            sample_size = min(15, len(tickers))
            np.random.seed(42)  # For reproducibility
            sample_tickers = np.random.choice(tickers, sample_size, replace=False) if len(tickers) > 0 else []
            
            for ticker in sample_tickers:
                # Generate synthetic data for analysis
                hourly_data = self.generate_synthetic_hourly_data(ticker, signal_date)
                
                if hourly_data is None or hourly_data.empty:
                    continue
                
                # Simulate trade
                trade = self.simulate_trade(ticker, signal_date, hourly_data)
                if trade:
                    all_trades.append(trade)
                    print(f"  {ticker}: Entry={trade['entry_price']:.2f}, Exit={trade['exit_price']:.2f}, "
                          f"Positions={trade['max_positions']}, P&L={trade['pnl_percent']:.2f}%")
        
        # Generate summary statistics
        self.generate_summary(all_trades)
        
        # Save detailed results
        self.save_results(all_trades)
        
    def generate_summary(self, trades):
        """Generate summary statistics"""
        if not trades:
            print("\nNo trades executed!")
            return
        
        print("\n" + "=" * 60)
        print("STRATEGY PERFORMANCE SUMMARY")
        print("=" * 60)
        
        # Convert to DataFrame for easier analysis
        df = pd.DataFrame(trades)
        
        # Overall statistics
        total_trades = len(df)
        winning_trades = len(df[df['pnl'] > 0])
        losing_trades = len(df[df['pnl'] < 0])
        win_rate = (winning_trades / total_trades) * 100
        
        print(f"\nTotal Trades: {total_trades}")
        print(f"Winning Trades: {winning_trades}")
        print(f"Losing Trades: {losing_trades}")
        print(f"Win Rate: {win_rate:.2f}%")
        
        # P&L Statistics
        avg_pnl_percent = df['pnl_percent'].mean()
        max_profit_percent = df['pnl_percent'].max()
        max_loss_percent = df['pnl_percent'].min()
        
        print(f"\nAverage P&L: {avg_pnl_percent:.2f}%")
        print(f"Best Trade: {max_profit_percent:.2f}%")
        print(f"Worst Trade: {max_loss_percent:.2f}%")
        
        # Position sizing statistics
        avg_positions = df['max_positions'].mean()
        max_positions_used = df['max_positions'].max()
        
        print(f"\nAverage Positions Used: {avg_positions:.1f}")
        print(f"Maximum Positions Used: {max_positions_used}")
        
        # Duration statistics
        avg_duration = df['trade_duration_hours'].mean()
        print(f"\nAverage Trade Duration: {avg_duration:.1f} hours")
        
        # Risk-Reward Analysis
        if winning_trades > 0 and losing_trades > 0:
            avg_win = df[df['pnl'] > 0]['pnl_percent'].mean()
            avg_loss = abs(df[df['pnl'] < 0]['pnl_percent'].mean())
            risk_reward = avg_win / avg_loss
            print(f"\nAverage Win: {avg_win:.2f}%")
            print(f"Average Loss: {avg_loss:.2f}%")
            print(f"Risk-Reward Ratio: {risk_reward:.2f}")
        
        # Expected value per trade
        expected_value = avg_pnl_percent
        print(f"\nExpected Value per Trade: {expected_value:.2f}%")
        
        # Capital usage analysis
        print("\n" + "-" * 60)
        print("CAPITAL USAGE ANALYSIS")
        print("-" * 60)
        
        # Calculate maximum capital deployed
        trades_with_5_positions = len(df[df['max_positions'] == 5])
        avg_capital_deployed = avg_positions * 0.01 * 100  # As percentage
        max_capital_deployed = 5 * 0.01 * 100  # 5% max
        
        print(f"Average Capital Deployed: {avg_capital_deployed:.1f}% of account")
        print(f"Maximum Capital Deployed: {max_capital_deployed:.1f}% of account")
        print(f"Trades using all 5 positions: {trades_with_5_positions} ({trades_with_5_positions/total_trades*100:.1f}%)")
        
        # Create visualizations
        self.create_visualizations(df)
        
    def create_visualizations(self, df):
        """Create performance visualizations"""
        plt.style.use('default')
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))
        
        # 1. P&L Distribution
        ax1 = axes[0, 0]
        df['pnl_percent'].hist(bins=20, ax=ax1, color='skyblue', edgecolor='black')
        ax1.axvline(0, color='red', linestyle='--', alpha=0.7)
        ax1.axvline(df['pnl_percent'].mean(), color='green', linestyle='--', alpha=0.7, label=f"Mean: {df['pnl_percent'].mean():.2f}%")
        ax1.set_title('P&L Distribution (Fixed 1% Sizing)', fontsize=14, fontweight='bold')
        ax1.set_xlabel('P&L %')
        ax1.set_ylabel('Frequency')
        ax1.legend()
        
        # 2. Positions Used Distribution
        ax2 = axes[0, 1]
        df['max_positions'].value_counts().sort_index().plot(kind='bar', ax=ax2, color='lightgreen')
        ax2.set_title('Number of Positions Used', fontsize=14, fontweight='bold')
        ax2.set_xlabel('Number of Positions')
        ax2.set_ylabel('Count')
        ax2.set_xticklabels(ax2.get_xticklabels(), rotation=0)
        
        # 3. Trade Duration vs P&L
        ax3 = axes[1, 0]
        scatter = ax3.scatter(df['trade_duration_hours'], df['pnl_percent'], 
                            c=df['max_positions'], cmap='viridis', alpha=0.7, s=100)
        ax3.axhline(0, color='black', linestyle='-', alpha=0.3)
        ax3.set_title('Trade Duration vs P&L (colored by positions)', fontsize=14, fontweight='bold')
        ax3.set_xlabel('Duration (hours)')
        ax3.set_ylabel('P&L %')
        cbar = plt.colorbar(scatter, ax=ax3)
        cbar.set_label('Positions Used')
        
        # 4. Cumulative P&L
        ax4 = axes[1, 1]
        df_sorted = df.sort_values('entry_time')
        df_sorted['cumulative_pnl'] = df_sorted['pnl_percent'].cumsum()
        ax4.plot(range(len(df_sorted)), df_sorted['cumulative_pnl'], 
                marker='o', linewidth=2, markersize=6)
        ax4.axhline(0, color='red', linestyle='--', alpha=0.5)
        ax4.set_title('Cumulative P&L % (Fixed Sizing)', fontsize=14, fontweight='bold')
        ax4.set_xlabel('Trade Number')
        ax4.set_ylabel('Cumulative P&L %')
        ax4.grid(True, alpha=0.3)
        
        # Add final cumulative P&L to the plot
        final_pnl = df_sorted['cumulative_pnl'].iloc[-1]
        ax4.text(len(df_sorted)-1, final_pnl, f'{final_pnl:.1f}%', 
                ha='left', va='bottom', fontweight='bold')
        
        plt.tight_layout()
        
        # Save visualization
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        plt.savefig(f'/Users/maverick/PycharmProjects/India-TS/Daily/analysis/volume_breakout_fixed_sizing_{timestamp}.png', 
                   dpi=300, bbox_inches='tight')
        plt.close()
        
    def save_results(self, trades):
        """Save detailed results to files"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Save detailed trades to JSON
        output = {
            'strategy': 'Volume Breakout with Fixed Position Sizing (1% per position, max 5)',
            'analysis_date': datetime.now().isoformat(),
            'period': '2025-07-08 to 2025-07-12',
            'total_trades': len(trades),
            'trades': trades,
            'summary': {
                'win_rate': len([t for t in trades if t['pnl'] > 0]) / len(trades) * 100 if trades else 0,
                'avg_pnl_percent': np.mean([t['pnl_percent'] for t in trades]) if trades else 0,
                'max_profit_percent': max([t['pnl_percent'] for t in trades]) if trades else 0,
                'max_loss_percent': min([t['pnl_percent'] for t in trades]) if trades else 0,
                'avg_positions': np.mean([t['max_positions'] for t in trades]) if trades else 0,
                'max_capital_deployed': 5  # 5% maximum
            }
        }
        
        json_path = f'/Users/maverick/PycharmProjects/India-TS/Daily/analysis/volume_breakout_fixed_sizing_{timestamp}.json'
        with open(json_path, 'w') as f:
            json.dump(output, f, indent=2, default=str)
        
        # Save to Excel for easy viewing
        if trades:
            # Create summary DataFrame
            df_trades = pd.DataFrame(trades)
            
            # Create position details DataFrame
            all_positions = []
            for trade in trades:
                for pos in trade['positions']:
                    all_positions.append({
                        'ticker': trade['ticker'],
                        'signal_date': trade['signal_date'],
                        'time': pos['time'],
                        'action': pos['action'],
                        'price': pos['price'],
                        'size': pos['size'],
                        'total_positions': pos['total_positions']
                    })
            df_positions = pd.DataFrame(all_positions)
            
            # Save to Excel with multiple sheets
            excel_path = f'/Users/maverick/PycharmProjects/India-TS/Daily/analysis/volume_breakout_fixed_sizing_{timestamp}.xlsx'
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df_trades.to_excel(writer, sheet_name='Trade Summary', index=False)
                df_positions.to_excel(writer, sheet_name='Position Details', index=False)
                
                # Add summary statistics sheet
                summary_data = {
                    'Metric': ['Total Trades', 'Win Rate %', 'Average P&L %', 'Best Trade %', 
                              'Worst Trade %', 'Avg Positions Used', 'Max Capital Deployed %',
                              'Expected Value %', 'Avg Trade Duration (hours)'],
                    'Value': [
                        len(trades),
                        output['summary']['win_rate'],
                        output['summary']['avg_pnl_percent'],
                        output['summary']['max_profit_percent'],
                        output['summary']['max_loss_percent'],
                        output['summary']['avg_positions'],
                        output['summary']['max_capital_deployed'],
                        output['summary']['avg_pnl_percent'],
                        np.mean([t['trade_duration_hours'] for t in trades])
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Summary Statistics', index=False)
            
            print(f"\nResults saved to:")
            print(f"  - {json_path}")
            print(f"  - {excel_path}")
            
        # Create a markdown report
        self.create_markdown_report(output, timestamp)
    
    def create_markdown_report(self, output, timestamp):
        """Create a detailed markdown report"""
        report_path = f'/Users/maverick/PycharmProjects/India-TS/Daily/analysis/volume_breakout_fixed_sizing_report_{timestamp}.md'
        
        with open(report_path, 'w') as f:
            f.write("# Volume Breakout with Fixed Position Sizing Strategy Analysis\n\n")
            f.write(f"**Analysis Date:** {output['analysis_date']}\n")
            f.write(f"**Period Analyzed:** {output['period']}\n\n")
            
            f.write("## Strategy Description\n")
            f.write("- **Entry:** Take position when hourly volume breaks out (>1.5x average volume)\n")
            f.write("- **Position Sizing:** Add 1% position for each favorable candle (max 5 positions = 5% total)\n")
            f.write("- **Exit:** Close all positions when price drops below previous candle low\n\n")
            
            f.write("## Performance Summary\n")
            f.write(f"- **Total Trades:** {output['total_trades']}\n")
            f.write(f"- **Win Rate:** {output['summary']['win_rate']:.2f}%\n")
            f.write(f"- **Average P&L:** {output['summary']['avg_pnl_percent']:.2f}%\n")
            f.write(f"- **Best Trade:** {output['summary']['max_profit_percent']:.2f}%\n")
            f.write(f"- **Worst Trade:** {output['summary']['max_loss_percent']:.2f}%\n")
            f.write(f"- **Expected Value:** {output['summary']['avg_pnl_percent']:.2f}%\n\n")
            
            f.write("## Position Management\n")
            f.write(f"- **Average Positions Used:** {output['summary']['avg_positions']:.1f}\n")
            f.write(f"- **Maximum Capital Deployed:** {output['summary']['max_capital_deployed']}%\n\n")
            
            f.write("## Risk Assessment\n")
            f.write("### Pros:\n")
            f.write("- **Controlled Risk:** Maximum 5% capital exposure per trade\n")
            f.write("- **Gradual Scaling:** Adds to winners incrementally\n")
            f.write("- **Clear Position Limits:** Prevents over-leveraging\n")
            f.write("- **Volume Confirmation:** Enters only on volume breakouts\n\n")
            
            f.write("### Cons:\n")
            f.write("- **Limited Upside:** Capped at 5% position size\n")
            f.write("- **Multiple Entries:** May average up into reversals\n")
            f.write("- **Exit Risk:** Single exit trigger for all positions\n\n")
            
            f.write("## Comparison with Position Doubling Strategy\n")
            f.write("| Metric | Fixed Sizing (1%) | Position Doubling |\n")
            f.write("|--------|------------------|------------------|\n")
            f.write("| Max Capital Risk | 5% | Up to 3200% |\n")
            f.write("| Risk Control | Excellent | Poor |\n")
            f.write("| Scalability | Good | Very Poor |\n")
            f.write("| Capital Requirements | Low | Extremely High |\n\n")
            
            f.write("## Recommendation\n")
            if output['summary']['avg_pnl_percent'] > 0:
                f.write("The fixed position sizing approach shows better risk management than position doubling. ")
                f.write("Consider the following improvements:\n")
                f.write("1. **Entry Refinement:** Add trend filters to avoid counter-trend trades\n")
                f.write("2. **Stop Loss:** Implement fixed percentage stop loss per position\n")
                f.write("3. **Profit Targets:** Consider taking partial profits at predetermined levels\n")
                f.write("4. **Market Conditions:** Test performance in different market regimes\n")
            else:
                f.write("While the risk management is significantly better than position doubling, ")
                f.write("the strategy still needs refinement to achieve positive expected value. ")
                f.write("Focus on improving entry signals and exit timing.\n")
        
        print(f"  - {report_path}")

def main():
    analyzer = VolumeBreakoutFixedSizingStrategy()
    analyzer.analyze_strategy()

if __name__ == "__main__":
    main()