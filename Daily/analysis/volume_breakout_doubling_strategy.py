"""
Volume Breakout with Position Doubling Strategy Analysis
Strategy: When Long Reversal signal appears, take position on hourly volume breakout,
double position each time price crosses previous hourly close, exit when price drops below previous candle low
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import json
import os
import glob
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import seaborn as sns
from kiteconnect import KiteConnect
from user_aware_data_handler import UserAwareDataHandler

class VolumeBreakoutDoublingStrategy:
    def __init__(self):
        self.data_handler = UserAwareDataHandler()
        self.results = []
        self.detailed_trades = []
        
    def get_signal_files(self, start_date='2025-07-08', end_date='2025-07-12'):
        """Get all Long Reversal signal files for the specified week"""
        pattern = "/Users/maverick/PycharmProjects/India-TS/Daily/results/Long_Reversal_Daily_2025*.xlsx"
        files = glob.glob(pattern)
        
        # Filter files by date range
        signal_files = []
        for file in files:
            # Extract date from filename
            date_part = file.split('_')[-2]
            try:
                file_date = datetime.strptime(date_part, '%Y%m%d')
                if datetime.strptime(start_date, '%Y-%m-%d') <= file_date <= datetime.strptime(end_date, '%Y-%m-%d'):
                    signal_files.append(file)
            except:
                continue
                
        return sorted(signal_files)
    
    def extract_tickers_from_excel(self, file_path):
        """Extract tickers from Long Reversal Excel file"""
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            
            tickers = []
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Assuming ticker is in first column
                    tickers.append(str(row[0]))
            
            wb.close()
            return tickers
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            return []
    
    def get_hourly_data(self, ticker, date):
        """Get hourly data for a ticker on a specific date"""
        try:
            # Get data for the date and next day
            from_date = date.strftime('%Y-%m-%d')
            to_date = (date + timedelta(days=1)).strftime('%Y-%m-%d')
            
            hourly_data = self.data_handler.get_historical_data(
                ticker, 
                interval='60minute',
                from_date=from_date,
                to_date=to_date
            )
            
            if hourly_data is None or hourly_data.empty:
                return None
                
            # Calculate volume metrics
            hourly_data['volume_sma'] = hourly_data['volume'].rolling(window=5).mean()
            hourly_data['volume_breakout'] = hourly_data['volume'] > (hourly_data['volume_sma'].shift(1) * 1.5)
            
            return hourly_data
        except Exception as e:
            print(f"Error getting hourly data for {ticker}: {e}")
            return None
    
    def simulate_trade(self, ticker, signal_date, hourly_data):
        """Simulate the volume breakout with position doubling strategy"""
        trade_details = {
            'ticker': ticker,
            'signal_date': signal_date,
            'entry_time': None,
            'entry_price': None,
            'exit_time': None,
            'exit_price': None,
            'max_position_size': 0,
            'positions': [],
            'pnl': 0,
            'pnl_percent': 0,
            'trade_duration_hours': 0
        }
        
        position_size = 0
        avg_entry_price = 0
        total_investment = 0
        
        # Look for volume breakout after signal
        entry_found = False
        prev_close = None
        prev_low = None
        
        for idx, row in hourly_data.iterrows():
            # Skip pre-market hours
            hour = idx.hour
            if hour < 9 or hour >= 16:
                continue
                
            # Look for entry on volume breakout
            if not entry_found and row['volume_breakout']:
                # Initial entry
                position_size = 1
                avg_entry_price = row['close']
                total_investment = avg_entry_price
                
                trade_details['entry_time'] = idx
                trade_details['entry_price'] = row['close']
                trade_details['positions'].append({
                    'time': idx,
                    'action': 'BUY',
                    'price': row['close'],
                    'size': 1,
                    'cumulative_size': position_size
                })
                
                entry_found = True
                prev_close = row['close']
                prev_low = row['low']
                
            elif entry_found:
                # Check for position doubling condition
                if prev_close and row['close'] > prev_close:
                    # Double position
                    new_shares = position_size  # Double means add same amount
                    position_size += new_shares
                    total_investment += row['close'] * new_shares
                    avg_entry_price = total_investment / position_size
                    
                    trade_details['positions'].append({
                        'time': idx,
                        'action': 'ADD',
                        'price': row['close'],
                        'size': new_shares,
                        'cumulative_size': position_size
                    })
                
                # Check exit condition
                if prev_low and row['close'] < prev_low:
                    # Exit all positions
                    trade_details['exit_time'] = idx
                    trade_details['exit_price'] = row['close']
                    trade_details['max_position_size'] = position_size
                    
                    # Calculate P&L
                    exit_value = row['close'] * position_size
                    trade_details['pnl'] = exit_value - total_investment
                    trade_details['pnl_percent'] = (trade_details['pnl'] / total_investment) * 100
                    
                    # Calculate duration
                    trade_details['trade_duration_hours'] = (idx - trade_details['entry_time']).total_seconds() / 3600
                    
                    break
                
                # Update previous values
                prev_close = row['close']
                prev_low = row['low']
        
        # If no exit, close at end of day
        if entry_found and trade_details['exit_time'] is None:
            last_row = hourly_data.iloc[-1]
            trade_details['exit_time'] = hourly_data.index[-1]
            trade_details['exit_price'] = last_row['close']
            trade_details['max_position_size'] = position_size
            
            exit_value = last_row['close'] * position_size
            trade_details['pnl'] = exit_value - total_investment
            trade_details['pnl_percent'] = (trade_details['pnl'] / total_investment) * 100
            trade_details['trade_duration_hours'] = (trade_details['exit_time'] - trade_details['entry_time']).total_seconds() / 3600
        
        return trade_details if entry_found else None
    
    def analyze_strategy(self):
        """Run the complete analysis"""
        print("Starting Volume Breakout with Position Doubling Strategy Analysis...")
        print("=" * 60)
        
        # Get signal files for the week
        signal_files = self.get_signal_files()
        print(f"Found {len(signal_files)} signal files for the week of July 8th")
        
        all_trades = []
        
        for file_path in signal_files:
            # Extract date from filename
            date_part = file_path.split('_')[-2]
            signal_date = datetime.strptime(date_part, '%Y%m%d')
            
            print(f"\nProcessing signals from {signal_date.strftime('%Y-%m-%d')}...")
            
            # Get tickers from file
            tickers = self.extract_tickers_from_excel(file_path)
            print(f"Found {len(tickers)} tickers with Long Reversal signals")
            
            # Limit to random sample for testing
            sample_size = min(10, len(tickers))
            sample_tickers = np.random.choice(tickers, sample_size, replace=False)
            
            for ticker in sample_tickers:
                # Get hourly data
                hourly_data = self.get_hourly_data(ticker, signal_date)
                if hourly_data is None or hourly_data.empty:
                    continue
                
                # Simulate trade
                trade = self.simulate_trade(ticker, signal_date, hourly_data)
                if trade:
                    all_trades.append(trade)
                    print(f"  {ticker}: Entry={trade['entry_price']:.2f}, Exit={trade['exit_price']:.2f}, "
                          f"Max Position={trade['max_position_size']}, P&L={trade['pnl_percent']:.2f}%")
        
        # Generate summary statistics
        self.generate_summary(all_trades)
        
        # Save detailed results
        self.save_results(all_trades)
        
    def generate_summary(self, trades):
        """Generate summary statistics"""
        if not trades:
            print("\nNo trades executed!")
            return
        
        print("\n" + "=" * 60)
        print("STRATEGY PERFORMANCE SUMMARY")
        print("=" * 60)
        
        # Convert to DataFrame for easier analysis
        df = pd.DataFrame(trades)
        
        # Overall statistics
        total_trades = len(df)
        winning_trades = len(df[df['pnl'] > 0])
        losing_trades = len(df[df['pnl'] < 0])
        win_rate = (winning_trades / total_trades) * 100
        
        print(f"\nTotal Trades: {total_trades}")
        print(f"Winning Trades: {winning_trades}")
        print(f"Losing Trades: {losing_trades}")
        print(f"Win Rate: {win_rate:.2f}%")
        
        # P&L Statistics
        avg_pnl_percent = df['pnl_percent'].mean()
        max_profit_percent = df['pnl_percent'].max()
        max_loss_percent = df['pnl_percent'].min()
        
        print(f"\nAverage P&L: {avg_pnl_percent:.2f}%")
        print(f"Best Trade: {max_profit_percent:.2f}%")
        print(f"Worst Trade: {max_loss_percent:.2f}%")
        
        # Position sizing statistics
        avg_max_position = df['max_position_size'].mean()
        max_position_ever = df['max_position_size'].max()
        
        print(f"\nAverage Max Position Size: {avg_max_position:.1f}")
        print(f"Largest Position Size: {max_position_ever}")
        
        # Duration statistics
        avg_duration = df['trade_duration_hours'].mean()
        print(f"\nAverage Trade Duration: {avg_duration:.1f} hours")
        
        # Risk-Reward Analysis
        if winning_trades > 0 and losing_trades > 0:
            avg_win = df[df['pnl'] > 0]['pnl_percent'].mean()
            avg_loss = abs(df[df['pnl'] < 0]['pnl_percent'].mean())
            risk_reward = avg_win / avg_loss
            print(f"\nAverage Win: {avg_win:.2f}%")
            print(f"Average Loss: {avg_loss:.2f}%")
            print(f"Risk-Reward Ratio: {risk_reward:.2f}")
        
        # Expected value per trade
        expected_value = avg_pnl_percent
        print(f"\nExpected Value per Trade: {expected_value:.2f}%")
        
        # Create visualizations
        self.create_visualizations(df)
        
    def create_visualizations(self, df):
        """Create performance visualizations"""
        plt.style.use('default')
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))
        
        # 1. P&L Distribution
        ax1 = axes[0, 0]
        df['pnl_percent'].hist(bins=20, ax=ax1, color='skyblue', edgecolor='black')
        ax1.axvline(0, color='red', linestyle='--', alpha=0.7)
        ax1.set_title('P&L Distribution', fontsize=14, fontweight='bold')
        ax1.set_xlabel('P&L %')
        ax1.set_ylabel('Frequency')
        
        # 2. Position Size Distribution
        ax2 = axes[0, 1]
        df['max_position_size'].value_counts().sort_index().plot(kind='bar', ax=ax2, color='lightgreen')
        ax2.set_title('Max Position Size Distribution', fontsize=14, fontweight='bold')
        ax2.set_xlabel('Max Position Size')
        ax2.set_ylabel('Count')
        
        # 3. Trade Duration vs P&L
        ax3 = axes[1, 0]
        scatter = ax3.scatter(df['trade_duration_hours'], df['pnl_percent'], 
                            c=df['pnl_percent'], cmap='RdYlGn', alpha=0.7, s=100)
        ax3.axhline(0, color='black', linestyle='-', alpha=0.3)
        ax3.set_title('Trade Duration vs P&L', fontsize=14, fontweight='bold')
        ax3.set_xlabel('Duration (hours)')
        ax3.set_ylabel('P&L %')
        plt.colorbar(scatter, ax=ax3)
        
        # 4. Cumulative P&L
        ax4 = axes[1, 1]
        df_sorted = df.sort_values('entry_time')
        df_sorted['cumulative_pnl'] = df_sorted['pnl_percent'].cumsum()
        ax4.plot(range(len(df_sorted)), df_sorted['cumulative_pnl'], 
                marker='o', linewidth=2, markersize=6)
        ax4.set_title('Cumulative P&L %', fontsize=14, fontweight='bold')
        ax4.set_xlabel('Trade Number')
        ax4.set_ylabel('Cumulative P&L %')
        ax4.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        # Save visualization
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        plt.savefig(f'/Users/maverick/PycharmProjects/India-TS/Daily/analysis/volume_breakout_strategy_{timestamp}.png', 
                   dpi=300, bbox_inches='tight')
        plt.close()
        
    def save_results(self, trades):
        """Save detailed results to files"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Save detailed trades to JSON
        output = {
            'strategy': 'Volume Breakout with Position Doubling',
            'analysis_date': datetime.now().isoformat(),
            'period': '2025-07-08 to 2025-07-12',
            'total_trades': len(trades),
            'trades': trades,
            'summary': {
                'win_rate': len([t for t in trades if t['pnl'] > 0]) / len(trades) * 100 if trades else 0,
                'avg_pnl_percent': np.mean([t['pnl_percent'] for t in trades]) if trades else 0,
                'max_profit_percent': max([t['pnl_percent'] for t in trades]) if trades else 0,
                'max_loss_percent': min([t['pnl_percent'] for t in trades]) if trades else 0,
                'avg_max_position': np.mean([t['max_position_size'] for t in trades]) if trades else 0
            }
        }
        
        json_path = f'/Users/maverick/PycharmProjects/India-TS/Daily/analysis/volume_breakout_strategy_{timestamp}.json'
        with open(json_path, 'w') as f:
            json.dump(output, f, indent=2, default=str)
        
        # Save to Excel for easy viewing
        if trades:
            # Create summary DataFrame
            df_trades = pd.DataFrame(trades)
            
            # Create position details DataFrame
            all_positions = []
            for trade in trades:
                for pos in trade['positions']:
                    all_positions.append({
                        'ticker': trade['ticker'],
                        'signal_date': trade['signal_date'],
                        'time': pos['time'],
                        'action': pos['action'],
                        'price': pos['price'],
                        'size': pos['size'],
                        'cumulative_size': pos['cumulative_size']
                    })
            df_positions = pd.DataFrame(all_positions)
            
            # Save to Excel with multiple sheets
            excel_path = f'/Users/maverick/PycharmProjects/India-TS/Daily/analysis/volume_breakout_strategy_{timestamp}.xlsx'
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df_trades.to_excel(writer, sheet_name='Trade Summary', index=False)
                df_positions.to_excel(writer, sheet_name='Position Details', index=False)
                
                # Add summary statistics sheet
                summary_data = {
                    'Metric': ['Total Trades', 'Win Rate %', 'Average P&L %', 'Best Trade %', 
                              'Worst Trade %', 'Avg Max Position Size', 'Expected Value %'],
                    'Value': [
                        len(trades),
                        output['summary']['win_rate'],
                        output['summary']['avg_pnl_percent'],
                        output['summary']['max_profit_percent'],
                        output['summary']['max_loss_percent'],
                        output['summary']['avg_max_position'],
                        output['summary']['avg_pnl_percent']
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Summary Statistics', index=False)
            
            print(f"\nResults saved to:")
            print(f"  - {json_path}")
            print(f"  - {excel_path}")

def main():
    analyzer = VolumeBreakoutDoublingStrategy()
    analyzer.analyze_strategy()

if __name__ == "__main__":
    main()