#!/usr/bin/env python3
"""
Action Plan Generator
=====================
Generates daily trading action plans based on Brooks Higher Probability LONG Reversal
frequency analysis. Provides tiered recommendations for the next trading day.

Default behavior: Generates 1-day analysis report
Can generate multiple reports or custom periods with parameters

Usage:
    python Action_plan.py              # Default: 1-day analysis
    python Action_plan.py --days 3     # Generate 3-day analysis
    python Action_plan.py --all        # Generate all 3 reports (1-day, 2-day, 3-day)

Author: Claude Code Assistant
Created: 2025-06-06
Modified: 2025-06-10 - Changed default to 1-day analysis, added --all option for multiple reports
"""

import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from collections import defaultdict
import logging
import configparser
from pathlib import Path

# Add parent directories to path for imports
# sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '..', 'ML'))

# Import the frequency analyzer
try:
    from ML.Frequent_ticker_performance import FrequentTickerPerformanceAnalyzer
except ImportError:
    print("Error: Could not import FrequentTickerPerformanceAnalyzer from ML module")
    sys.exit(1)

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ActionPlanGenerator:
    """Generate consolidated action plans for Brooks reversal trading"""
    
    def __init__(self, user_name="Sai", lookback_days=1):
        """Initialize the action plan generator"""
        self.user_name = user_name
        self.lookback_days = lookback_days
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.daily_dir = os.path.dirname(self.script_dir)
        self.plan_dir = os.path.join(self.daily_dir, "Plan")
        
        # Create Plan directory if it doesn't exist
        os.makedirs(self.plan_dir, exist_ok=True)
        
        # Load config for additional settings
        self.config = self._load_config()
        
        # For 1-day analysis, we need to look at yesterday's data since today's won't be available yet
        # when running in the morning
        if lookback_days == 1:
            logger.info("Note: 1-day analysis will look at yesterday's data")
        
        logger.info(f"Initialized Action Plan Generator for {lookback_days}-day analysis")
    
    def generate_all_reports(self):
        """Generate all three reports (1-day, 2-day, 3-day)"""
        logger.info("Generating comprehensive action plan reports...")
        
        results = {}
        
        # Generate 1-day report
        logger.info("\n" + "="*50)
        logger.info("Generating 1-DAY analysis...")
        logger.info("="*50)
        self.lookback_days = 1
        filename_1day = self.generate_action_plan("1day")
        results['1day'] = filename_1day
        
        # Generate 2-day report
        logger.info("\n" + "="*50)
        logger.info("Generating 2-DAY analysis...")
        logger.info("="*50)
        self.lookback_days = 2
        filename_2day = self.generate_action_plan("2day")
        results['2day'] = filename_2day
        
        # Generate 3-day report
        logger.info("\n" + "="*50)
        logger.info("Generating 3-DAY analysis...")
        logger.info("="*50)
        self.lookback_days = 3
        filename_3day = self.generate_action_plan("3day")
        results['3day'] = filename_3day
        
        logger.info("\n" + "="*50)
        logger.info("All reports generated successfully!")
        logger.info("="*50)
        
        return results
    
    def _load_config(self):
        """Load configuration from config.ini"""
        config = configparser.ConfigParser()
        config_path = os.path.join(self.daily_dir, 'config.ini')
        
        if os.path.exists(config_path):
            config.read(config_path)
        
        return config
    
    def generate_action_plan(self, period_suffix=""):
        """Generate the complete action plan - consolidated Excel only"""
        logger.info(f"Generating action plan based on {self.lookback_days}-day frequency analysis...")
        
        # Initialize the frequency analyzer
        analyzer = FrequentTickerPerformanceAnalyzer(
            user_name=self.user_name, 
            days_back=self.lookback_days
        )
        
        # Run analysis
        returns_data = analyzer.analyze_reports()
        
        if not returns_data:
            logger.error("No data available for analysis")
            return None
        
        # Categorize tickers into tiers
        tiers = self._categorize_tickers(returns_data)
        
        # Create consolidated buy list Excel file only
        txt_path, excel_path = self._save_consolidated_plan(tiers, returns_data, period_suffix)
        
        logger.info(f"Consolidated Excel plan saved")
        
        # Return the Excel path for reporting
        return excel_path
    
    def _categorize_tickers(self, returns_data):
        """Categorize tickers into investment tiers based on frequency"""
        # Adjust thresholds based on lookback period
        if self.lookback_days <= 1:
            # 1-day thresholds
            tier1_min = 3   # 3+ appearances in 1 day
            tier2_min = 2   # 2 appearances
            tier3_min = 1   # 1 appearance
            tier4_min = 0   # Not used for 1-day
            tier5_min = -1  # Not used for 1-day
        elif self.lookback_days <= 2:
            # 2-day thresholds
            tier1_min = 5   # 5+ appearances in 2 days
            tier2_min = 3   # 3-4 appearances
            tier3_min = 2   # 2 appearances
            tier4_min = 1   # 1 appearance
            tier5_min = 0   # Not used for 2-day
        else:
            # 3+ day thresholds (original)
            tier1_min = 16  # 16+ appearances
            tier2_min = 11  # 11-15 appearances
            tier3_min = 6   # 6-10 appearances
            tier4_min = 3   # 3-5 appearances
            tier5_min = 0   # <3 appearances
        
        tiers = {
            'tier_1_maximum_conviction': [],
            'tier_2_strong_buy': [],
            'tier_3_consider': [],
            'tier_4_watch': [],
            'tier_5_avoid': []
        }
        
        for ticker, data in returns_data.items():
            ticker_info = {
                'ticker': ticker,
                'appearances': data['appearances'],
                'return_pct': data['return_pct'],
                'current_price': data.get('current_price', 0),
                'first_date': data['first_appearance'].strftime('%Y-%m-%d'),
                'days_held': (datetime.now() - data['first_appearance']).days,
                'score': self._calculate_score(data)
            }
            
            if data['appearances'] >= tier1_min:
                tiers['tier_1_maximum_conviction'].append(ticker_info)
            elif data['appearances'] >= tier2_min:
                tiers['tier_2_strong_buy'].append(ticker_info)
            elif data['appearances'] >= tier3_min:
                tiers['tier_3_consider'].append(ticker_info)
            elif tier4_min > 0 and data['appearances'] >= tier4_min:
                tiers['tier_4_watch'].append(ticker_info)
            else:
                tiers['tier_5_avoid'].append(ticker_info)
        
        # Sort each tier by score
        for tier in tiers:
            tiers[tier].sort(key=lambda x: x['score'], reverse=True)
        
        return tiers
    
    def _calculate_score(self, data):
        """Calculate composite score for ranking"""
        # Weight frequency more heavily than returns for reliability
        frequency_weight = 0.7
        return_weight = 0.3
        
        # Normalize frequency (assume max 30)
        freq_score = min(data['appearances'] / 30, 1) * 100 * frequency_weight
        
        # Normalize returns (cap at 20%)
        return_score = min(max(data['return_pct'], -20), 20) / 20 * 100 * return_weight
        
        return freq_score + return_score
    
    def _create_plan_content(self, tiers, returns_data):
        """Create the formatted action plan content"""
        lines = []
        
        # Header
        lines.append("=" * 120)
        lines.append("DAILY ACTION PLAN - BROOKS HIGHER PROBABILITY LONG REVERSAL")
        lines.append("=" * 120)
        lines.append(f"\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Analysis Period: Past {self.lookback_days} days")
        lines.append(f"Total Tickers Analyzed: {len(returns_data)}")
        
        # Market Overview
        total_winners = len([d for d in returns_data.values() if d['return_pct'] > 0])
        avg_return = np.mean([d['return_pct'] for d in returns_data.values()])
        
        lines.append(f"\nMarket Performance:")
        lines.append(f"  - Overall Win Rate: {total_winners/len(returns_data)*100:.1f}%")
        lines.append(f"  - Average Return: {avg_return:.2f}%")
        
        # Executive Summary
        lines.append("\n" + "=" * 120)
        lines.append("EXECUTIVE SUMMARY")
        lines.append("=" * 120)
        
        tier_counts = {
            'Tier 1 (16+)': len(tiers['tier_1_maximum_conviction']),
            'Tier 2 (11-15)': len(tiers['tier_2_strong_buy']),
            'Tier 3 (6-10)': len(tiers['tier_3_consider']),
            'Tier 4 (3-5)': len(tiers['tier_4_watch']),
            'Tier 5 (<3)': len(tiers['tier_5_avoid'])
        }
        
        lines.append("\nTicker Distribution:")
        for tier_name, count in tier_counts.items():
            lines.append(f"  - {tier_name}: {count} tickers")
        
        actionable = tier_counts['Tier 1 (16+)'] + tier_counts['Tier 2 (11-15)'] + tier_counts['Tier 3 (6-10)']
        lines.append(f"\nTotal Actionable Tickers (6+ appearances): {actionable}")
        
        # Tier 1: Maximum Conviction
        lines.append("\n" + "=" * 120)
        lines.append("TIER 1: MAXIMUM CONVICTION (16+ appearances)")
        lines.append("Historical Win Rate: 100% | Recommended Position Size: 100%")
        lines.append("=" * 120)
        
        if tiers['tier_1_maximum_conviction']:
            lines.append(f"\n{'Rank':<5} {'Ticker':<10} {'Freq':<6} {'Return%':<10} {'Price':<12} {'Days':<6} {'Action':<25} {'Notes':<30}")
            lines.append("-" * 120)
            
            for i, ticker in enumerate(tiers['tier_1_maximum_conviction'], 1):
                action = self._get_action(ticker, 'tier_1')
                notes = self._get_notes(ticker)
                lines.append(
                    f"{i:<5} {ticker['ticker']:<10} {ticker['appearances']:<6} "
                    f"{ticker['return_pct']:<10.2f} ₹{ticker['current_price']:<11.2f} "
                    f"{ticker['days_held']:<6} {action:<25} {notes:<30}"
                )
        else:
            lines.append("\n[No tickers qualify for this tier]")
        
        # Tier 2: Strong Buy
        lines.append("\n" + "=" * 120)
        lines.append("TIER 2: STRONG BUY (11-15 appearances)")
        lines.append("Historical Win Rate: 96.7% | Recommended Position Size: 75%")
        lines.append("=" * 120)
        
        if tiers['tier_2_strong_buy']:
            lines.append(f"\n{'Rank':<5} {'Ticker':<10} {'Freq':<6} {'Return%':<10} {'Price':<12} {'Days':<6} {'Action':<25} {'Notes':<30}")
            lines.append("-" * 120)
            
            for i, ticker in enumerate(tiers['tier_2_strong_buy'], 1):
                action = self._get_action(ticker, 'tier_2')
                notes = self._get_notes(ticker)
                lines.append(
                    f"{i:<5} {ticker['ticker']:<10} {ticker['appearances']:<6} "
                    f"{ticker['return_pct']:<10.2f} ₹{ticker['current_price']:<11.2f} "
                    f"{ticker['days_held']:<6} {action:<25} {notes:<30}"
                )
        else:
            lines.append("\n[No tickers qualify for this tier]")
        
        # Tier 3: Consider (Top 15 only)
        lines.append("\n" + "=" * 120)
        lines.append("TIER 3: CONSIDER (6-10 appearances)")
        lines.append("Historical Win Rate: 80.6% | Recommended Position Size: 50%")
        lines.append("=" * 120)
        
        if tiers['tier_3_consider']:
            lines.append("\n[Showing top 15 tickers only]")
            lines.append(f"\n{'Rank':<5} {'Ticker':<10} {'Freq':<6} {'Return%':<10} {'Price':<12} {'Days':<6} {'Action':<25} {'Notes':<30}")
            lines.append("-" * 120)
            
            for i, ticker in enumerate(tiers['tier_3_consider'][:15], 1):
                action = self._get_action(ticker, 'tier_3')
                notes = self._get_notes(ticker)
                lines.append(
                    f"{i:<5} {ticker['ticker']:<10} {ticker['appearances']:<6} "
                    f"{ticker['return_pct']:<10.2f} ₹{ticker['current_price']:<11.2f} "
                    f"{ticker['days_held']:<6} {action:<25} {notes:<30}"
                )
            
            if len(tiers['tier_3_consider']) > 15:
                lines.append(f"\n[... and {len(tiers['tier_3_consider']) - 15} more tickers in this tier]")
        else:
            lines.append("\n[No tickers qualify for this tier]")
        
        # Consolidated Buy List
        lines.append("\n" + "=" * 120)
        lines.append("CONSOLIDATED BUY LIST (All Actionable Tickers)")
        lines.append("=" * 120)
        
        all_buys = []
        for ticker in tiers['tier_1_maximum_conviction']:
            ticker['tier'] = 'Tier 1'
            ticker['position_size'] = '100%'
            all_buys.append(ticker)
        
        for ticker in tiers['tier_2_strong_buy']:
            ticker['tier'] = 'Tier 2'
            ticker['position_size'] = '75%'
            all_buys.append(ticker)
        
        for ticker in tiers['tier_3_consider']:
            if ticker['return_pct'] > 3:  # Only include profitable ones
                ticker['tier'] = 'Tier 3'
                ticker['position_size'] = '50%'
                all_buys.append(ticker)
        
        # Sort by score
        all_buys.sort(key=lambda x: x['score'], reverse=True)
        
        lines.append(f"\n{'Rank':<5} {'Ticker':<10} {'Tier':<8} {'Freq':<6} {'Return%':<10} {'Price':<12} {'Position':<10}")
        lines.append("-" * 80)
        
        for i, ticker in enumerate(all_buys[:25], 1):  # Top 25
            lines.append(
                f"{i:<5} {ticker['ticker']:<10} {ticker['tier']:<8} {ticker['appearances']:<6} "
                f"{ticker['return_pct']:<10.2f} ₹{ticker['current_price']:<11.2f} {ticker['position_size']:<10}"
            )
        
        # Trading Instructions
        lines.append("\n" + "=" * 120)
        lines.append("TRADING INSTRUCTIONS")
        lines.append("=" * 120)
        
        lines.append("\n1. EXECUTION PRIORITY:")
        lines.append("   - First: All Tier 1 tickers (100% position)")
        lines.append("   - Second: Tier 2 tickers with returns > 5% (75% position)")
        lines.append("   - Third: Tier 3 tickers with returns > 3% (50% position)")
        
        lines.append("\n2. ORDER PLACEMENT:")
        lines.append("   - Time: Place orders at 9:15 AM market open")
        lines.append("   - Type: Limit orders at or slightly above current price")
        lines.append("   - Gap Rule: Skip if stock gaps up > 2% from close")
        
        lines.append("\n3. RISK MANAGEMENT:")
        lines.append("   - Stop Loss: -3% for all positions")
        lines.append("   - First Target: +5% (book 50% profit)")
        lines.append("   - Second Target: +8% (exit remaining)")
        lines.append("   - Time Stop: Exit if no movement in 3 days")
        
        lines.append("\n4. POSITION MONITORING:")
        lines.append("   - Check frequency daily - exit if drops below tier threshold")
        lines.append("   - Trail stop to breakeven after +3% move")
        lines.append("   - Scale out on weakness, scale in on strength")
        
        # Risk Disclaimer
        lines.append("\n" + "=" * 120)
        lines.append("DISCLAIMER")
        lines.append("=" * 120)
        lines.append("\nThis action plan is based on historical frequency analysis of Brooks reversal patterns.")
        lines.append("Past performance does not guarantee future results. Trade at your own risk.")
        lines.append("Always use proper position sizing and risk management.")
        
        return '\n'.join(lines)
    
    def _get_action(self, ticker_info, tier):
        """Determine specific action for a ticker"""
        if tier == 'tier_1':
            return "BUY - Full Position"
        elif tier == 'tier_2':
            if ticker_info['return_pct'] > 5:
                return "BUY - 75% Position"
            elif ticker_info['return_pct'] > 0:
                return "BUY - Half Position"
            else:
                return "WAIT - Monitor"
        elif tier == 'tier_3':
            if ticker_info['return_pct'] > 5:
                return "CONSIDER - 50% Pos"
            elif ticker_info['return_pct'] > 3:
                return "WATCH - Small Pos"
            else:
                return "AVOID"
        else:
            return "NO ACTION"
    
    def _get_notes(self, ticker_info):
        """Generate notes/warnings for a ticker"""
        notes = []
        
        if ticker_info['return_pct'] > 15:
            notes.append("Extended move")
        elif ticker_info['return_pct'] > 10:
            notes.append("Strong momentum")
        
        if ticker_info['days_held'] <= 2:
            notes.append("New signal")
        elif ticker_info['days_held'] > 7:
            notes.append("Mature signal")
        
        if ticker_info['current_price'] < 100:
            notes.append("Penny stock")
        elif ticker_info['current_price'] > 5000:
            notes.append("High priced")
        
        return ", ".join(notes) if notes else "-"
    
    def _save_html_plan(self, content, filepath):
        """Convert text content to HTML and save"""
        html_content = self._convert_to_html(content)
        with open(filepath, 'w') as f:
            f.write(html_content)
        logger.info(f"HTML action plan saved to: {filepath}")
    
    def _convert_to_html(self, text_content):
        """Convert text action plan to HTML format"""
        html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Brooks Trading Action Plan</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
            color: #333;
        }
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        h1 {
            color: #2c3e50;
            text-align: center;
            border-bottom: 3px solid #3498db;
            padding-bottom: 15px;
            margin-bottom: 30px;
        }
        h2 {
            color: #34495e;
            background-color: #ecf0f1;
            padding: 10px;
            border-left: 5px solid #3498db;
            margin-top: 30px;
        }
        .summary-box {
            background-color: #e8f4f8;
            border: 1px solid #3498db;
            border-radius: 5px;
            padding: 15px;
            margin: 20px 0;
        }
        .tier-1 {
            background-color: #d4edda;
            border-color: #28a745;
        }
        .tier-1 h2 {
            background-color: #28a745;
            color: white;
            border-left: none;
        }
        .tier-2 {
            background-color: #cce5ff;
            border-color: #007bff;
        }
        .tier-2 h2 {
            background-color: #007bff;
            color: white;
            border-left: none;
        }
        .tier-3 {
            background-color: #fff3cd;
            border-color: #ffc107;
        }
        .tier-3 h2 {
            background-color: #ffc107;
            color: #333;
            border-left: none;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 15px 0;
        }
        th {
            background-color: #34495e;
            color: white;
            padding: 12px;
            text-align: left;
            position: sticky;
            top: 0;
        }
        td {
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }
        tr:hover {
            background-color: #f5f5f5;
        }
        .instructions {
            background-color: #f8f9fa;
            border: 1px solid #dee2e6;
            border-radius: 5px;
            padding: 20px;
            margin: 20px 0;
        }
        .disclaimer {
            background-color: #f8d7da;
            border: 1px solid #f5c6cb;
            border-radius: 5px;
            padding: 15px;
            margin: 20px 0;
            color: #721c24;
        }
        .metric {
            display: inline-block;
            margin: 10px 20px 10px 0;
            font-size: 16px;
        }
        .metric-label {
            font-weight: bold;
            color: #555;
        }
        .metric-value {
            color: #2c3e50;
            font-size: 18px;
            font-weight: bold;
        }
        .positive {
            color: #28a745;
        }
        .negative {
            color: #dc3545;
        }
        pre {
            white-space: pre-wrap;
            font-family: monospace;
            background-color: #f5f5f5;
            padding: 10px;
            border-radius: 5px;
            overflow-x: auto;
        }
    </style>
</head>
<body>
    <div class="container">
"""
        
        lines = text_content.split('\n')
        current_section = None
        in_table = False
        table_buffer = []
        
        for line in lines:
            # Skip empty lines
            if not line.strip():
                if in_table and table_buffer:
                    html += self._create_html_table(table_buffer)
                    table_buffer = []
                    in_table = False
                continue
            
            # Main title
            if "DAILY ACTION PLAN" in line and "=" in lines[lines.index(line)-1] if lines.index(line) > 0 else False:
                html += "<h1>DAILY ACTION PLAN - BROOKS HIGHER PROBABILITY LONG REVERSAL</h1>\n"
            
            # Section headers
            elif line.startswith("=") and len(set(line)) == 1:
                continue
            
            # Major sections
            elif "TIER 1:" in line:
                if in_table and table_buffer:
                    html += self._create_html_table(table_buffer)
                    table_buffer = []
                    in_table = False
                html += '<div class="tier-1">\n<h2>' + line + '</h2>\n'
                current_section = 'tier-1'
            elif "TIER 2:" in line:
                if in_table and table_buffer:
                    html += self._create_html_table(table_buffer)
                    table_buffer = []
                    in_table = False
                if current_section:
                    html += '</div>\n'
                html += '<div class="tier-2">\n<h2>' + line + '</h2>\n'
                current_section = 'tier-2'
            elif "TIER 3:" in line:
                if in_table and table_buffer:
                    html += self._create_html_table(table_buffer)
                    table_buffer = []
                    in_table = False
                if current_section:
                    html += '</div>\n'
                html += '<div class="tier-3">\n<h2>' + line + '</h2>\n'
                current_section = 'tier-3'
            elif any(header in line for header in ["EXECUTIVE SUMMARY", "CONSOLIDATED BUY LIST", "TRADING INSTRUCTIONS", "DISCLAIMER"]):
                if in_table and table_buffer:
                    html += self._create_html_table(table_buffer)
                    table_buffer = []
                    in_table = False
                if current_section:
                    html += '</div>\n'
                    current_section = None
                
                if "DISCLAIMER" in line:
                    html += '<div class="disclaimer">\n<h2>' + line + '</h2>\n'
                elif "TRADING INSTRUCTIONS" in line:
                    html += '<div class="instructions">\n<h2>' + line + '</h2>\n'
                else:
                    html += '<h2>' + line + '</h2>\n'
            
            # Table headers
            elif line.startswith("Rank") and "Ticker" in line:
                in_table = True
                table_buffer = [line]
            elif line.startswith("-") and len(set(line.strip())) <= 2:
                continue
            
            # Table data
            elif in_table and line.strip() and not line.startswith("["):
                table_buffer.append(line)
            
            # Metrics
            elif "Generated:" in line or "Analysis Period:" in line or "Total Tickers Analyzed:" in line:
                parts = line.split(":")
                if len(parts) == 2:
                    html += f'<div class="metric"><span class="metric-label">{parts[0]}:</span> <span class="metric-value">{parts[1].strip()}</span></div>\n'
            elif "Win Rate:" in line or "Average Return:" in line:
                parts = line.split(":")
                if len(parts) == 2:
                    value = parts[1].strip()
                    css_class = "positive" if "%" in value and float(value.rstrip('%')) > 0 else ""
                    html += f'<div class="metric"><span class="metric-label">{parts[0].strip()}:</span> <span class="metric-value {css_class}">{value}</span></div>\n'
            
            # Lists
            elif line.strip().startswith("-"):
                html += f"<li>{line.strip()[1:].strip()}</li>\n"
            elif line.strip().startswith(tuple(str(i) + "." for i in range(1, 10))):
                html += f"<h3>{line.strip()}</h3>\n"
            
            # Regular paragraphs
            else:
                if not line.startswith(" ") or ":" not in line:
                    html += f"<p>{line.strip()}</p>\n"
                else:
                    html += f"<div style='margin-left: 20px;'>{line.strip()}</div>\n"
        
        # Close any open sections
        if in_table and table_buffer:
            html += self._create_html_table(table_buffer)
        if current_section:
            html += '</div>\n'
        
        html += """
    </div>
</body>
</html>"""
        
        return html
    
    def _create_html_table(self, table_lines):
        """Convert table lines to HTML table"""
        if not table_lines:
            return ""
        
        html = "<table>\n<thead>\n<tr>\n"
        
        # Parse header
        headers = table_lines[0].split()
        for header in headers:
            html += f"<th>{header}</th>\n"
        html += "</tr>\n</thead>\n<tbody>\n"
        
        # Parse data rows
        for line in table_lines[1:]:
            if line.strip():
                html += "<tr>\n"
                # Split by multiple spaces to handle column data
                parts = line.split()
                
                # Reconstruct columns based on expected format
                if len(parts) >= 7:
                    # Handle ticker data rows
                    rank = parts[0]
                    ticker = parts[1]
                    freq = parts[2]
                    return_pct = parts[3]
                    price = ' '.join(parts[4:6]) if len(parts) > 5 and parts[4].startswith('₹') else parts[4]
                    days = parts[5] if not parts[4].startswith('₹') else parts[6]
                    
                    # Remaining parts are action and notes
                    remaining_idx = 6 if not parts[4].startswith('₹') else 7
                    if remaining_idx < len(parts):
                        action = ' '.join(parts[remaining_idx:remaining_idx+3])
                        notes = ' '.join(parts[remaining_idx+3:]) if remaining_idx+3 < len(parts) else "-"
                    else:
                        action = "-"
                        notes = "-"
                    
                    html += f"<td>{rank}</td>\n"
                    html += f"<td><strong>{ticker}</strong></td>\n"
                    html += f"<td>{freq}</td>\n"
                    
                    # Color code returns
                    return_class = "positive" if float(return_pct) > 0 else "negative"
                    html += f'<td class="{return_class}">{return_pct}%</td>\n'
                    
                    html += f"<td>{price}</td>\n"
                    html += f"<td>{days}</td>\n"
                    
                    if len(headers) > 6:
                        html += f"<td>{action}</td>\n"
                    if len(headers) > 7:
                        html += f"<td>{notes}</td>\n"
                else:
                    # Fallback for other table formats
                    for part in parts:
                        html += f"<td>{part}</td>\n"
                
                html += "</tr>\n"
        
        html += "</tbody>\n</table>\n"
        return html
    
    def _save_plan(self, content, period_suffix=""):
        """Save the action plan to file"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = f"_{period_suffix}" if period_suffix else ""
        filename = f"Action_Plan{suffix}_{timestamp}.txt"
        filepath = os.path.join(self.plan_dir, filename)
        
        with open(filepath, 'w') as f:
            f.write(content)
        
        logger.info(f"Action plan saved to: {filepath}")
        
        # Generate and save HTML version
        html_filename = f"Action_Plan{suffix}_{timestamp}.html"
        html_filepath = os.path.join(self.plan_dir, html_filename)
        self._save_html_plan(content, html_filepath)
        
        return filepath
    
    def _save_csv_summary(self, tiers, period_suffix=""):
        """Save a CSV summary for easy import into spreadsheets"""
        all_tickers = []
        
        # Compile all tickers with tier info
        tier_map = {
            'tier_1_maximum_conviction': 'Tier 1 (16+)',
            'tier_2_strong_buy': 'Tier 2 (11-15)',
            'tier_3_consider': 'Tier 3 (6-10)',
            'tier_4_watch': 'Tier 4 (3-5)',
            'tier_5_avoid': 'Tier 5 (<3)'
        }
        
        for tier_key, tier_name in tier_map.items():
            for ticker in tiers[tier_key]:
                ticker_data = ticker.copy()
                ticker_data['tier'] = tier_name
                ticker_data['recommendation'] = self._get_action(ticker, tier_key)
                all_tickers.append(ticker_data)
        
        # Create DataFrame
        df = pd.DataFrame(all_tickers)
        
        # Reorder columns
        columns = ['ticker', 'tier', 'appearances', 'return_pct', 'current_price', 
                   'days_held', 'score', 'recommendation', 'first_date']
        df = df[columns]
        
        # Sort by tier and score
        df = df.sort_values(['tier', 'score'], ascending=[True, False])
        
        # Save CSV
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = f"_{period_suffix}" if period_suffix else ""
        csv_filename = f"Action_Plan_Summary{suffix}_{timestamp}.csv"
        csv_filepath = os.path.join(self.plan_dir, csv_filename)
        
        df.to_csv(csv_filepath, index=False)
        
        logger.info(f"CSV summary saved to: {csv_filepath}")
        
        return csv_filepath
    
    def _save_consolidated_plan(self, tiers, returns_data, period_suffix=""):
        """Save consolidated buy list to text file with frequency and recency-weighted position sizing"""
        # Collect all actionable tickers (Tier 1, 2, and profitable Tier 3)
        all_buys = []
        
        # Add Tier 1 tickers
        for ticker in tiers['tier_1_maximum_conviction']:
            ticker_copy = ticker.copy()
            ticker_copy['tier'] = 'Tier 1'
            # Get appearance dates from returns_data
            if ticker['ticker'] in returns_data:
                ticker_copy['appearance_dates'] = returns_data[ticker['ticker']].get('appearance_dates', [])
            else:
                ticker_copy['appearance_dates'] = []
            all_buys.append(ticker_copy)
        
        # Add Tier 2 tickers
        for ticker in tiers['tier_2_strong_buy']:
            ticker_copy = ticker.copy()
            ticker_copy['tier'] = 'Tier 2'
            # Get appearance dates from returns_data
            if ticker['ticker'] in returns_data:
                ticker_copy['appearance_dates'] = returns_data[ticker['ticker']].get('appearance_dates', [])
            else:
                ticker_copy['appearance_dates'] = []
            all_buys.append(ticker_copy)
        
        # Add profitable Tier 3 tickers (return > 3%)
        for ticker in tiers['tier_3_consider']:
            if ticker['return_pct'] > 3:
                ticker_copy = ticker.copy()
                ticker_copy['tier'] = 'Tier 3'
                # Get appearance dates from returns_data
                if ticker['ticker'] in returns_data:
                    ticker_copy['appearance_dates'] = returns_data[ticker['ticker']].get('appearance_dates', [])
                else:
                    ticker_copy['appearance_dates'] = []
                all_buys.append(ticker_copy)
        
        # Sort by score (highest first)
        all_buys.sort(key=lambda x: x['score'], reverse=True)
        
        # Take only top 20 tickers
        top_20_buys = all_buys[:20]
        
        # Calculate frequency-weighted and recency-weighted position sizes
        num_positions = len(top_20_buys)
        if num_positions > 0:
            # Calculate total frequency for normalization
            total_frequency = sum(ticker['appearances'] for ticker in top_20_buys)
            
            # Calculate recency scores (appearances in last 3 days get more weight)
            now = datetime.now()
            for ticker in top_20_buys:
                recency_score = 0
                recent_count = 0
                for date in ticker.get('appearance_dates', []):
                    days_ago = (now - date).days
                    if days_ago <= 3:  # Last 3 days
                        recency_score += ticker['appearances'] * 2  # Double weight for recent
                        recent_count += 2
                    elif days_ago <= 5:  # 4-5 days ago
                        recency_score += ticker['appearances'] * 1.5  # 1.5x weight
                        recent_count += 1.5
                    else:  # Older than 5 days
                        recency_score += ticker['appearances'] * 1  # Normal weight
                        recent_count += 1
                
                # Store recency score (average weighted frequency)
                if recent_count > 0:
                    ticker['recency_weighted_freq'] = recency_score / len(ticker.get('appearance_dates', [1]))
                else:
                    ticker['recency_weighted_freq'] = ticker['appearances']
            
            # Calculate total recency-weighted frequency
            total_recency_freq = sum(ticker['recency_weighted_freq'] for ticker in top_20_buys)
            
            # Calculate position sizes
            for ticker in top_20_buys:
                # Frequency-based position
                ticker['position_size'] = (ticker['appearances'] / total_frequency) * 100
                # Recency-weighted position
                ticker['recency_position_size'] = (ticker['recency_weighted_freq'] / total_recency_freq) * 100
        else:
            for ticker in top_20_buys:
                ticker['position_size'] = 0
                ticker['recency_position_size'] = 0
        
        # Create the content
        lines = []
        lines.append("CONSOLIDATED BUY LIST - TOP 20 POSITIONS (RECENCY WEIGHTED)")
        lines.append("=" * 70)
        lines.append(f"Total Tickers Selected: {num_positions} (from {len(all_buys)} qualified)")
        lines.append(f"Allocation Method: Recency-weighted (recent appearances get higher weight)")
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("=" * 70)
        lines.append("")
        
        # Header
        lines.append(f"{'Rank':<6}{'Ticker':<12}{'Price':<15}{'Position%':<10}")
        lines.append("-" * 70)
        
        # Data rows for top 20 only
        for i, ticker in enumerate(top_20_buys, 1):
            rank = str(i)
            symbol = ticker['ticker']
            price = f"₹{ticker['current_price']:.2f}"
            position = f"{ticker['recency_position_size']:.2f}%"
            
            lines.append(f"{rank:<6}{symbol:<12}{price:<15}{position:<10}")
        
        # Footer
        lines.append("-" * 70)
        lines.append(f"Total Allocation: 100.00%")
        lines.append("")
        lines.append("Notes:")
        lines.append("- Only top 20 tickers by score are included")
        lines.append("- Position allocation is weighted by recency of appearances:")
        lines.append("  • Last 3 days: 2x weight (captures immediate momentum)")
        lines.append("  • 4-5 days ago: 1.5x weight (recent but not immediate)")
        lines.append("  • Older than 5 days: 1x weight (normal weighting)")
        lines.append("- Tickers appearing frequently in recent days get higher allocation")
        lines.append("- This helps capture market pivots and emerging sector trends")
        lines.append("- Adjust position sizes based on your risk tolerance and capital")
        
        # Save the file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = f"_{period_suffix}" if period_suffix else ""
        
        # Create Excel version only
        excel_filename = f"Consolidated_Plan{suffix}_{timestamp}.xlsx"
        excel_filepath = os.path.join(self.plan_dir, excel_filename)
        
        # Prepare data for Excel
        excel_data = []
        for ticker in top_20_buys:
            excel_data.append({
                'Rank': top_20_buys.index(ticker) + 1,
                'Ticker': ticker['ticker'],
                'Price': ticker['current_price'],
                'Position%': ticker['recency_position_size']
            })
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        
        # Create Excel writer with formatting
        with pd.ExcelWriter(excel_filepath, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Consolidated_Plan', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Consolidated_Plan']
            
            # Add header formatting
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Format columns
            worksheet.column_dimensions['A'].width = 8  # Rank
            worksheet.column_dimensions['B'].width = 15  # Ticker
            worksheet.column_dimensions['C'].width = 15  # Price
            worksheet.column_dimensions['D'].width = 15  # Position%
            
            # Format price column as currency
            for row in range(2, len(df) + 2):
                worksheet[f'C{row}'].number_format = '₹#,##0.00'
                worksheet[f'D{row}'].number_format = '0.00"%"'
            
            # Add summary information
            summary_row = len(df) + 4
            worksheet[f'A{summary_row}'] = 'Summary'
            worksheet[f'A{summary_row}'].font = Font(bold=True)
            
            worksheet[f'A{summary_row + 1}'] = 'Total Positions:'
            worksheet[f'B{summary_row + 1}'] = num_positions
            
            worksheet[f'A{summary_row + 2}'] = 'Total Allocation:'
            worksheet[f'B{summary_row + 2}'] = '100.00%'
            
            worksheet[f'A{summary_row + 3}'] = 'Generated:'
            worksheet[f'B{summary_row + 3}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Add notes
            notes_row = summary_row + 5
            worksheet[f'A{notes_row}'] = 'Notes:'
            worksheet[f'A{notes_row}'].font = Font(bold=True)
            
            notes = [
                'Position allocation is weighted by recency of appearances',
                'Last 3 days: 2x weight | 4-5 days: 1.5x weight | Older: 1x weight',
                'Helps capture market pivots and emerging sector trends'
            ]
            
            for i, note in enumerate(notes):
                worksheet[f'A{notes_row + i + 1}'] = f'• {note}'
        
        # Save as latest Excel with ONLY the required columns (no summary, no notes)
        # Only update the latest file for the 1-day analysis (default)
        if not period_suffix or period_suffix == "1day":
            latest_excel_path = os.path.join(self.plan_dir, "Consolidated_Plan_Latest.xlsx")
            
            # Create a clean version for Latest file
            with pd.ExcelWriter(latest_excel_path, engine='openpyxl') as writer:
                # Write only the data - no additional information
                df.to_excel(writer, sheet_name='Orders', index=False)
            
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Orders']
                
                # Add minimal header formatting
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Format columns
                worksheet.column_dimensions['A'].width = 8  # Rank
                worksheet.column_dimensions['B'].width = 15  # Ticker
                worksheet.column_dimensions['C'].width = 15  # Price
                worksheet.column_dimensions['D'].width = 15  # Position%
                
                # Format data
                for row in range(2, len(df) + 2):
                    worksheet[f'A{row}'].alignment = Alignment(horizontal="center")  # Center align rank
                    worksheet[f'C{row}'].number_format = '₹#,##0.00'
                    worksheet[f'D{row}'].number_format = '0.00"%"'
            
            logger.info(f"Simplified Latest consolidated plan file updated (Orders only)")
        
        logger.info(f"Consolidated plan Excel saved to: {excel_filepath}")
        
        return None, excel_filepath  # No txt file anymore, only Excel

def main():
    """Main function"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate daily action plan for Brooks reversal trading')
    parser.add_argument('--user', type=str, default='Sai', help='Zerodha user name')
    parser.add_argument('--days', type=int, default=None, help='Number of days to analyze (default: 1)')
    parser.add_argument('--all', action='store_true', help='Generate all three reports (1-day, 2-day, 3-day)')
    
    args = parser.parse_args()
    
    # Check if user wants all reports
    if args.all:
        # Generate all three reports
        print("\n📊 Brooks Reversal Action Plan Generator")
        print("=" * 50)
        print("🔄 Generating comprehensive analysis (1-day, 2-day, and 3-day reports)...")
        
        generator = ActionPlanGenerator(user_name=args.user)
        results = generator.generate_all_reports()
        
        if results:
            print(f"\n✅ All consolidated Excel reports successfully generated!")
            print(f"\n📊 Excel files created:")
            if results.get('1day'):
                print(f"   1-Day: {os.path.basename(results.get('1day'))}")
            if results.get('2day'):
                print(f"   2-Day: {os.path.basename(results.get('2day'))}")
            if results.get('3day'):
                print(f"   3-Day: {os.path.basename(results.get('3day'))}")
            print(f"\n📈 Consolidated_Plan_Latest.xlsx updated (1-day version)")
            print(f"\n💡 Check the 'Plan' folder for generated Excel files")
    else:
        # Generate single report for specified days (default: 1 day)
        lookback_days = args.days if args.days is not None else 1
        print(f"\n📊 Brooks Reversal Action Plan Generator")
        print("=" * 50)
        print(f"🔍 Analyzing Brooks reversal patterns from the past {lookback_days} {'day' if lookback_days == 1 else 'days'}...")
        
        # Generate action plan
        generator = ActionPlanGenerator(user_name=args.user, lookback_days=lookback_days)
        excel_file = generator.generate_action_plan()
        
        if excel_file:
            print(f"\n✅ Consolidated Excel report successfully generated!")
            print(f"📊 Excel file: {os.path.basename(excel_file)}")
            if lookback_days == 1:  # Only 1-day updates the Latest file
                print(f"📈 Consolidated_Plan_Latest.xlsx updated")
            print(f"\n💡 Check the 'Plan' folder for the Excel file")

if __name__ == "__main__":
    main()