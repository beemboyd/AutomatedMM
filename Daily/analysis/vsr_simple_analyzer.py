#!/usr/bin/env python3
"""
VSR Simple Weekend Analyzer
Analyzes VSR signals and calculates performance without Zerodha API
Uses last known prices from scanner files as proxy for current prices
"""

import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import glob

def analyze_vsr_performance():
    """Analyze VSR performance from scanner files"""
    
    base_dir = "/Users/maverick/PycharmProjects/India-TS/Daily/scanners/Hourly"
    
    # Get date range
    end_date = datetime.now()
    start_date = end_date - timedelta(days=7)
    
    print(f"\nVSR WEEKEND EFFICIENCY ANALYSIS")
    print(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
    print("="*80)
    
    # Dictionary to store ticker data
    ticker_data = {}
    
    # Process VSR files
    vsr_files = sorted(glob.glob(os.path.join(base_dir, "VSR_*.xlsx")))
    
    for filepath in vsr_files:
        try:
            # Extract date from filename
            filename = os.path.basename(filepath)
            date_str = filename.replace('VSR_', '').replace('.xlsx', '')
            file_date = datetime.strptime(date_str[:8], '%Y%m%d')
            
            # Check if within date range
            if file_date < start_date or file_date > end_date:
                continue
                
            # Read Excel file
            df = pd.read_excel(filepath)
            
            if df.empty:
                continue
                
            # Process each ticker
            for _, row in df.iterrows():
                ticker = row.get('Ticker', '')
                if not ticker:
                    continue
                    
                if ticker not in ticker_data:
                    # First appearance - store initial data
                    ticker_data[ticker] = {
                        'first_date': file_date,
                        'first_price': row.get('Entry_Price', row.get('Close', 0)),
                        'stop_loss': row.get('Stop_Loss', 0),
                        'target1': row.get('Target1', 0),
                        'target2': row.get('Target2', 0),
                        'pattern': row.get('Pattern', ''),
                        'vsr_score': row.get('VSR_Score', row.get('Score', 0)),
                        'probability': row.get('Probability_Score', 0),
                        'alert_count': 1,
                        'last_date': file_date,
                        'last_price': row.get('Entry_Price', row.get('Close', 0)),
                        'max_score': row.get('VSR_Score', row.get('Score', 0)),
                        'direction': row.get('Direction', 'LONG')
                    }
                else:
                    # Update existing ticker data
                    ticker_data[ticker]['alert_count'] += 1
                    ticker_data[ticker]['last_date'] = file_date
                    ticker_data[ticker]['last_price'] = row.get('Entry_Price', row.get('Close', 0))
                    
                    current_score = row.get('VSR_Score', row.get('Score', 0))
                    if current_score > ticker_data[ticker]['max_score']:
                        ticker_data[ticker]['max_score'] = current_score
                        
        except Exception as e:
            print(f"Error processing {filepath}: {e}")
            continue
    
    if not ticker_data:
        print("No VSR data found for analysis period")
        return
        
    # Calculate performance metrics
    performance_data = []
    
    for ticker, data in ticker_data.items():
        if data['first_price'] and data['last_price']:
            price_change = data['last_price'] - data['first_price']
            price_change_pct = (price_change / data['first_price']) * 100
            
            # Check if targets were likely hit (based on last price)
            sl_hit = data['last_price'] <= data['stop_loss'] if data['stop_loss'] else False
            t1_hit = data['last_price'] >= data['target1'] if data['target1'] else False
            t2_hit = data['last_price'] >= data['target2'] if data['target2'] else False
            
            # Calculate efficiency score
            if sl_hit:
                efficiency = 0
            elif t2_hit:
                efficiency = 100
            elif t1_hit:
                efficiency = 75
            elif price_change_pct > 0:
                efficiency = min(50 + price_change_pct * 2, 74)
            else:
                efficiency = max(0, 50 + price_change_pct)
                
            performance_data.append({
                'Ticker': ticker,
                'First Date': data['first_date'].strftime('%Y-%m-%d'),
                'Entry Price': data['first_price'],
                'Last Price': data['last_price'],
                'Change %': price_change_pct,
                'Stop Loss': data['stop_loss'],
                'Target 1': data['target1'],
                'Target 2': data['target2'],
                'Pattern': data['pattern'],
                'Alerts': data['alert_count'],
                'Max Score': data['max_score'],
                'Efficiency': efficiency,
                'Direction': data['direction']
            })
    
    # Create DataFrame
    df = pd.DataFrame(performance_data)
    
    if df.empty:
        print("No performance data to analyze")
        return
        
    # Sort by efficiency
    df = df.sort_values('Efficiency', ascending=False)
    
    # Calculate statistics
    total = len(df)
    winners = len(df[df['Change %'] > 0])
    losers = len(df[df['Change %'] < 0])
    win_rate = (winners / total * 100) if total > 0 else 0
    
    # Print summary
    print(f"\nPERFORMANCE SUMMARY")
    print("-"*40)
    print(f"Total Tickers: {total}")
    print(f"Winners: {winners} ({win_rate:.1f}%)")
    print(f"Losers: {losers} ({100-win_rate:.1f}%)")
    
    if winners > 0:
        avg_gain = df[df['Change %'] > 0]['Change %'].mean()
        print(f"Average Gain: {avg_gain:.2f}%")
        
    if losers > 0:
        avg_loss = df[df['Change %'] < 0]['Change %'].mean()
        print(f"Average Loss: {avg_loss:.2f}%")
    
    # Top performers
    print(f"\nTOP 5 PERFORMERS")
    print("-"*40)
    top_5 = df.nlargest(5, 'Change %')
    for _, row in top_5.iterrows():
        print(f"{row['Ticker']:10} +{row['Change %']:6.2f}%  Entry: {row['Entry Price']:8.2f}  Last: {row['Last Price']:8.2f}")
    
    # Bottom performers
    print(f"\nBOTTOM 5 PERFORMERS")
    print("-"*40)
    bottom_5 = df.nsmallest(5, 'Change %')
    for _, row in bottom_5.iterrows():
        print(f"{row['Ticker']:10} {row['Change %']:7.2f}%  Entry: {row['Entry Price']:8.2f}  Last: {row['Last Price']:8.2f}")
    
    # Most active tickers
    print(f"\nMOST ACTIVE (BY ALERTS)")
    print("-"*40)
    most_active = df.nlargest(5, 'Alerts')
    for _, row in most_active.iterrows():
        print(f"{row['Ticker']:10} {row['Alerts']:3} alerts  Change: {row['Change %']:+6.2f}%  Efficiency: {row['Efficiency']:.0f}")
    
    # Pattern analysis
    print(f"\nPATTERN PERFORMANCE")
    print("-"*40)
    pattern_stats = df.groupby('Pattern').agg({
        'Ticker': 'count',
        'Change %': 'mean',
        'Efficiency': 'mean'
    }).round(2)
    pattern_stats.columns = ['Count', 'Avg Change %', 'Avg Efficiency']
    pattern_stats = pattern_stats.sort_values('Avg Efficiency', ascending=False)
    
    for pattern, row in pattern_stats.head(5).iterrows():
        if pattern:  # Skip empty patterns
            print(f"{pattern[:30]:30} Count: {int(row['Count']):3}  Avg Change: {row['Avg Change %']:+6.2f}%")
    
    # Save to Excel
    save_excel_report(df, pattern_stats)
    
    # Specific ticker analysis - IMFA
    print(f"\nSPECIFIC TICKER ANALYSIS: IMFA")
    print("-"*40)
    if 'IMFA' in df['Ticker'].values:
        imfa_data = df[df['Ticker'] == 'IMFA'].iloc[0]
        print(f"First Alert: {imfa_data['First Date']}")
        print(f"Entry Price: {imfa_data['Entry Price']:.2f}")
        print(f"Last Price: {imfa_data['Last Price']:.2f}")
        print(f"Change: {imfa_data['Change %']:.2f}%")
        print(f"Pattern: {imfa_data['Pattern']}")
        print(f"Total Alerts: {imfa_data['Alerts']}")
        print(f"Efficiency Score: {imfa_data['Efficiency']:.0f}")
    else:
        print("IMFA not found in VSR alerts for this period")
    
    print("\n" + "="*80)
    
    return df

def save_excel_report(df, pattern_stats):
    """Save analysis to Excel"""
    
    # Generate filename
    date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = f"/Users/maverick/PycharmProjects/India-TS/Daily/analysis/Efficiency/VSR_Simple_Analysis_{date_str}.xlsx"
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Performance Data
    ws1 = wb.active
    ws1.title = "Performance"
    
    # Add headers
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for r_idx, row in df.iterrows():
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx+2, column=c_idx+1, value=value)
            
            # Format numbers
            if headers[c_idx] in ['Entry Price', 'Last Price', 'Stop Loss', 'Target 1', 'Target 2']:
                cell.number_format = '#,##0.00'
            elif headers[c_idx] == 'Change %':
                cell.number_format = '+#,##0.00%;-#,##0.00%'
                if value > 0:
                    cell.font = Font(color="008000")
                elif value < 0:
                    cell.font = Font(color="FF0000")
    
    # Sheet 2: Pattern Statistics
    ws2 = wb.create_sheet("Pattern Analysis")
    
    # Add pattern stats
    ws2.append(['Pattern', 'Count', 'Avg Change %', 'Avg Efficiency'])
    for pattern, row in pattern_stats.iterrows():
        if pattern:
            ws2.append([pattern, int(row['Count']), row['Avg Change %'], row['Avg Efficiency']])
    
    # Adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save
    wb.save(filepath)
    print(f"\nReport saved to: {filepath}")

if __name__ == "__main__":
    analyze_vsr_performance()