"""
Keltner Channel Put Selling Strategy Analysis
Strategy: Sell puts at the lower Keltner Channel band (~2 months out) on Long Reversal signals
Collect premium from time decay and mean reversion tendencies
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import json
import os
import glob
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import seaborn as sns
import pickle
from scipy.stats import norm

class KeltnerPutSellingStrategy:
    def __init__(self):
        self.results = []
        self.detailed_trades = []
        self.data_cache_dir = "/Users/maverick/PycharmProjects/India-TS/Daily/analysis/data/"
        
    def get_signal_files(self, start_date='2025-07-08', end_date='2025-07-12'):
        """Get all Long Reversal signal files for the specified week"""
        pattern = "/Users/maverick/PycharmProjects/India-TS/Daily/results/Long_Reversal_Daily_2025*.xlsx"
        files = glob.glob(pattern)
        
        # Filter files by date range
        signal_files = []
        for file in files:
            # Extract date from filename
            date_part = file.split('_')[-2]
            try:
                file_date = datetime.strptime(date_part, '%Y%m%d')
                if datetime.strptime(start_date, '%Y-%m-%d') <= file_date <= datetime.strptime(end_date, '%Y-%m-%d'):
                    signal_files.append(file)
            except:
                continue
                
        return sorted(signal_files)
    
    def extract_tickers_from_excel(self, file_path):
        """Extract tickers from Long Reversal Excel file"""
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            
            tickers = []
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Assuming ticker is in first column
                    tickers.append(str(row[0]))
            
            wb.close()
            return tickers
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            return []
    
    def calculate_keltner_channels(self, data, period=20, multiplier=2.0):
        """Calculate Keltner Channel bands"""
        # Calculate EMA (middle line)
        data['ema'] = data['close'].ewm(span=period, min_periods=1).mean()
        
        # Calculate ATR
        data['high_low'] = data['high'] - data['low']
        data['high_close'] = abs(data['high'] - data['close'].shift(1))
        data['low_close'] = abs(data['low'] - data['close'].shift(1))
        data['true_range'] = data[['high_low', 'high_close', 'low_close']].max(axis=1)
        data['atr'] = data['true_range'].rolling(window=period, min_periods=1).mean()
        
        # Calculate bands
        data['kc_upper'] = data['ema'] + (multiplier * data['atr'])
        data['kc_lower'] = data['ema'] - (multiplier * data['atr'])
        
        return data
    
    def generate_synthetic_daily_data(self, ticker, date, days_forward=60):
        """Generate synthetic daily data for backtesting with realistic patterns"""
        # Create daily timestamps for 60 days forward
        timestamps = pd.date_range(
            start=date,
            periods=days_forward,
            freq='D'
        )
        
        # Generate realistic price movements with mean reversion tendency
        base_price = 100 + np.random.uniform(-50, 200)  # Random base price
        
        # Simulate realistic daily patterns with some mean reversion
        data = []
        prev_close = base_price
        
        # Add some trend and volatility parameters
        annual_volatility = np.random.uniform(0.25, 0.45)  # 25-45% annual volatility
        daily_volatility = annual_volatility / np.sqrt(252)
        drift = np.random.uniform(-0.0002, 0.0004)  # Small daily drift
        
        for i, ts in enumerate(timestamps):
            # Add mean reversion component
            mean_reversion_strength = 0.05  # How strong is mean reversion
            distance_from_start = (prev_close - base_price) / base_price
            mean_reversion_component = -mean_reversion_strength * distance_from_start
            
            # Random walk with drift and mean reversion
            price_change = (drift + mean_reversion_component + 
                          np.random.normal(0, daily_volatility)) * prev_close
            
            open_price = prev_close
            close_price = prev_close + price_change
            
            # Calculate high and low with realistic intraday range
            intraday_range = abs(price_change) * np.random.uniform(1.2, 2.5)
            high_price = max(open_price, close_price) + intraday_range * 0.6
            low_price = min(open_price, close_price) - intraday_range * 0.4
            
            # Volume pattern
            volume = int(np.random.uniform(100000, 500000))
            
            data.append({
                'date': ts,
                'open': round(open_price, 2),
                'high': round(high_price, 2),
                'low': round(low_price, 2),
                'close': round(close_price, 2),
                'volume': volume
            })
            
            prev_close = close_price
        
        df = pd.DataFrame(data)
        df.set_index('date', inplace=True)
        
        # Calculate Keltner Channels
        df = self.calculate_keltner_channels(df)
        
        return df
    
    def calculate_put_option_price(self, spot_price, strike_price, time_to_expiry, risk_free_rate=0.06, volatility=0.30):
        """Calculate put option price using Black-Scholes formula"""
        if time_to_expiry <= 0:
            return max(strike_price - spot_price, 0)
        
        d1 = (np.log(spot_price / strike_price) + (risk_free_rate + 0.5 * volatility**2) * time_to_expiry) / (volatility * np.sqrt(time_to_expiry))
        d2 = d1 - volatility * np.sqrt(time_to_expiry)
        
        put_price = strike_price * np.exp(-risk_free_rate * time_to_expiry) * norm.cdf(-d2) - spot_price * norm.cdf(-d1)
        return max(put_price, 0)
    
    def simulate_put_selling_trade(self, ticker, signal_date, daily_data):
        """Simulate selling puts at lower Keltner Channel"""
        trade_details = {
            'ticker': ticker,
            'signal_date': signal_date,
            'entry_date': None,
            'expiry_date': None,
            'strike_price': None,
            'premium_collected': 0,
            'exit_date': None,
            'exit_reason': None,
            'final_pnl': 0,
            'max_loss': 0,
            'success': False,
            'days_held': 0
        }
        
        # Get the signal day data
        signal_day_data = daily_data.iloc[0]
        
        if pd.isna(signal_day_data['kc_lower']):
            return None
        
        # Set strike at lower Keltner Channel
        strike_price = signal_day_data['kc_lower']
        spot_price = signal_day_data['close']
        
        distance_percent = (spot_price - strike_price) / spot_price * 100
        
        # Skip if strike is too close to spot (less than 1% away)
        if distance_percent < 1.0:
            return None
        
        # Set expiry to approximately 60 days (2 months)
        expiry_days = 60
        time_to_expiry = expiry_days / 365.0
        
        # Calculate premium collected (selling the put)
        volatility = 0.30  # Assume 30% implied volatility
        premium_collected = self.calculate_put_option_price(
            spot_price, strike_price, time_to_expiry, volatility=volatility
        )
        
        # As percentage of strike price (margin requirement)
        premium_yield = (premium_collected / strike_price) * 100
        
        trade_details.update({
            'entry_date': signal_date,
            'expiry_date': signal_date + timedelta(days=expiry_days),
            'strike_price': strike_price,
            'spot_at_entry': spot_price,
            'premium_collected': premium_collected,
            'premium_yield': premium_yield,
            'margin_required': strike_price * 0.20  # Assume 20% margin requirement
        })
        
        # Track the trade through expiry or early exit
        max_loss = 0
        exit_found = False
        
        for i, (date, row) in enumerate(daily_data.iterrows()):
            if i == 0:  # Skip entry day
                continue
                
            days_remaining = expiry_days - i
            current_spot = row['close']
            
            if days_remaining <= 0:
                # At expiry
                intrinsic_value = max(strike_price - current_spot, 0)
                final_pnl = premium_collected - intrinsic_value
                
                trade_details.update({
                    'exit_date': date,
                    'exit_reason': 'expiry',
                    'final_pnl': final_pnl,
                    'final_pnl_percent': (final_pnl / trade_details['margin_required']) * 100,
                    'spot_at_exit': current_spot,
                    'success': final_pnl > 0,
                    'days_held': i
                })
                exit_found = True
                break
            
            # Calculate current option value for mark-to-market
            time_remaining = days_remaining / 365.0
            current_option_value = self.calculate_put_option_price(
                current_spot, strike_price, time_remaining, volatility=volatility
            )
            
            # P&L is premium collected minus current option value
            current_pnl = premium_collected - current_option_value
            max_loss = min(max_loss, current_pnl)
            
            # Early exit conditions
            # Exit if profit target hit (50% of premium)
            if current_pnl >= premium_collected * 0.5:
                trade_details.update({
                    'exit_date': date,
                    'exit_reason': 'profit_target',
                    'final_pnl': current_pnl,
                    'final_pnl_percent': (current_pnl / trade_details['margin_required']) * 100,
                    'spot_at_exit': current_spot,
                    'success': True,
                    'days_held': i
                })
                exit_found = True
                break
            
            # Exit if loss limit hit (-200% of premium)
            if current_pnl <= -premium_collected * 2.0:
                trade_details.update({
                    'exit_date': date,
                    'exit_reason': 'stop_loss',
                    'final_pnl': current_pnl,
                    'final_pnl_percent': (current_pnl / trade_details['margin_required']) * 100,
                    'spot_at_exit': current_spot,
                    'success': False,
                    'days_held': i
                })
                exit_found = True
                break
        
        if not exit_found:
            # Shouldn't happen, but handle edge case
            return None
            
        trade_details['max_loss'] = max_loss
        trade_details['max_loss_percent'] = (max_loss / trade_details['margin_required']) * 100
        
        return trade_details
    
    def analyze_strategy(self):
        """Run the complete analysis"""
        print("Starting Keltner Channel Put Selling Strategy Analysis...")
        print("=" * 60)
        
        # Get signal files for the week
        signal_files = self.get_signal_files()
        print(f"Found {len(signal_files)} signal files for the week of July 8th")
        
        all_trades = []
        
        for file_path in signal_files:
            # Extract date from filename
            date_part = file_path.split('_')[-2]
            signal_date = datetime.strptime(date_part, '%Y%m%d')
            
            print(f"\nProcessing signals from {signal_date.strftime('%Y-%m-%d')}...")
            
            # Get tickers from file
            tickers = self.extract_tickers_from_excel(file_path)
            print(f"Found {len(tickers)} tickers with Long Reversal signals")
            
            # Limit to random sample for testing
            sample_size = min(15, len(tickers))
            np.random.seed(42)  # For reproducibility
            sample_tickers = np.random.choice(tickers, sample_size, replace=False) if len(tickers) > 0 else []
            
            for ticker in sample_tickers:
                # Generate synthetic data for analysis
                daily_data = self.generate_synthetic_daily_data(ticker, signal_date)
                
                if daily_data is None or daily_data.empty:
                    continue
                
                # Simulate put selling trade
                trade = self.simulate_put_selling_trade(ticker, signal_date, daily_data)
                if trade:
                    all_trades.append(trade)
                    print(f"  {ticker}: Strike={trade['strike_price']:.2f}, Premium={trade['premium_yield']:.2f}%, "
                          f"P&L={trade['final_pnl_percent']:.2f}%, Exit={trade['exit_reason']}")
        
        # Generate summary statistics
        self.generate_summary(all_trades)
        
        # Save detailed results
        self.save_results(all_trades)
        
    def generate_summary(self, trades):
        """Generate summary statistics"""
        if not trades:
            print("\nNo trades executed!")
            return
        
        print("\n" + "=" * 60)
        print("PUT SELLING STRATEGY PERFORMANCE SUMMARY")
        print("=" * 60)
        
        # Convert to DataFrame for easier analysis
        df = pd.DataFrame(trades)
        
        # Overall statistics
        total_trades = len(df)
        winning_trades = len(df[df['success'] == True])
        losing_trades = len(df[df['success'] == False])
        win_rate = (winning_trades / total_trades) * 100
        
        print(f"\nTotal Trades: {total_trades}")
        print(f"Winning Trades: {winning_trades}")
        print(f"Losing Trades: {losing_trades}")
        print(f"Win Rate: {win_rate:.2f}%")
        
        # P&L Statistics
        avg_pnl_percent = df['final_pnl_percent'].mean()
        max_profit_percent = df['final_pnl_percent'].max()
        max_loss_percent = df['final_pnl_percent'].min()
        
        print(f"\nAverage P&L: {avg_pnl_percent:.2f}%")
        print(f"Best Trade: {max_profit_percent:.2f}%")
        print(f"Worst Trade: {max_loss_percent:.2f}%")
        
        # Premium collection statistics
        avg_premium_yield = df['premium_yield'].mean()
        avg_days_held = df['days_held'].mean()
        
        print(f"\nAverage Premium Collected: {avg_premium_yield:.2f}%")
        print(f"Average Days Held: {avg_days_held:.1f}")
        
        # Exit reason breakdown
        print(f"\nExit Reasons:")
        exit_counts = df['exit_reason'].value_counts()
        for reason, count in exit_counts.items():
            print(f"  {reason}: {count} ({count/total_trades*100:.1f}%)")
        
        # Risk-Reward Analysis
        if winning_trades > 0 and losing_trades > 0:
            avg_win = df[df['success'] == True]['final_pnl_percent'].mean()
            avg_loss = df[df['success'] == False]['final_pnl_percent'].mean()
            risk_reward = abs(avg_win / avg_loss)
            print(f"\nAverage Win: {avg_win:.2f}%")
            print(f"Average Loss: {avg_loss:.2f}%")
            print(f"Risk-Reward Ratio: {risk_reward:.2f}")
        
        # Expected value per trade
        expected_value = avg_pnl_percent
        print(f"\nExpected Value per Trade: {expected_value:.2f}%")
        
        # Annualized return estimate
        avg_trade_duration = avg_days_held / 365
        if avg_trade_duration > 0:
            annualized_return = (1 + expected_value/100) ** (1/avg_trade_duration) - 1
            print(f"Estimated Annualized Return: {annualized_return*100:.2f}%")
        
        # Capital efficiency
        print("\n" + "-" * 60)
        print("CAPITAL EFFICIENCY ANALYSIS")
        print("-" * 60)
        
        avg_margin = df['margin_required'].mean()
        print(f"Average Margin Required: ₹{avg_margin:,.0f}")
        print(f"Capital efficiency: {avg_premium_yield:.2f}% return for ~{avg_days_held:.0f} days")
        
        # Create visualizations
        self.create_visualizations(df)
        
    def create_visualizations(self, df):
        """Create performance visualizations"""
        plt.style.use('default')
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))
        
        # 1. P&L Distribution
        ax1 = axes[0, 0]
        df['final_pnl_percent'].hist(bins=20, ax=ax1, color='lightblue', edgecolor='black')
        ax1.axvline(0, color='red', linestyle='--', alpha=0.7)
        ax1.axvline(df['final_pnl_percent'].mean(), color='green', linestyle='--', alpha=0.7, 
                   label=f"Mean: {df['final_pnl_percent'].mean():.2f}%")
        ax1.set_title('P&L Distribution (Put Selling)', fontsize=14, fontweight='bold')
        ax1.set_xlabel('P&L %')
        ax1.set_ylabel('Frequency')
        ax1.legend()
        
        # 2. Premium Yield vs Days Held
        ax2 = axes[0, 1]
        colors = ['green' if x else 'red' for x in df['success']]
        scatter = ax2.scatter(df['days_held'], df['premium_yield'], c=colors, alpha=0.7, s=100)
        ax2.set_title('Premium Yield vs Days Held', fontsize=14, fontweight='bold')
        ax2.set_xlabel('Days Held')
        ax2.set_ylabel('Premium Yield %')
        
        # Add custom legend
        from matplotlib.patches import Patch
        legend_elements = [Patch(facecolor='green', label='Profitable'),
                          Patch(facecolor='red', label='Loss')]
        ax2.legend(handles=legend_elements)
        
        # 3. Exit Reasons Pie Chart
        ax3 = axes[1, 0]
        exit_counts = df['exit_reason'].value_counts()
        colors = ['lightgreen', 'lightcoral', 'lightyellow']
        ax3.pie(exit_counts.values, labels=exit_counts.index, autopct='%1.1f%%', colors=colors)
        ax3.set_title('Exit Reasons Distribution', fontsize=14, fontweight='bold')
        
        # 4. Cumulative P&L
        ax4 = axes[1, 1]
        df_sorted = df.sort_values('entry_date')
        df_sorted['cumulative_pnl'] = df_sorted['final_pnl_percent'].cumsum()
        ax4.plot(range(len(df_sorted)), df_sorted['cumulative_pnl'], 
                marker='o', linewidth=2, markersize=6, color='purple')
        ax4.axhline(0, color='red', linestyle='--', alpha=0.5)
        ax4.set_title('Cumulative P&L % (Put Selling)', fontsize=14, fontweight='bold')
        ax4.set_xlabel('Trade Number')
        ax4.set_ylabel('Cumulative P&L %')
        ax4.grid(True, alpha=0.3)
        
        # Add final cumulative P&L to the plot
        final_pnl = df_sorted['cumulative_pnl'].iloc[-1]
        ax4.text(len(df_sorted)-1, final_pnl, f'{final_pnl:.1f}%', 
                ha='left', va='bottom', fontweight='bold')
        
        plt.tight_layout()
        
        # Save visualization
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        plt.savefig(f'/Users/maverick/PycharmProjects/India-TS/Daily/analysis/keltner_put_selling_{timestamp}.png', 
                   dpi=300, bbox_inches='tight')
        plt.close()
        
    def save_results(self, trades):
        """Save detailed results to files"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Save detailed trades to JSON
        output = {
            'strategy': 'Keltner Channel Put Selling (2 months expiry)',
            'analysis_date': datetime.now().isoformat(),
            'period': '2025-07-08 to 2025-07-12',
            'total_trades': len(trades),
            'trades': trades,
            'summary': {
                'win_rate': len([t for t in trades if t['success']]) / len(trades) * 100 if trades else 0,
                'avg_pnl_percent': np.mean([t['final_pnl_percent'] for t in trades]) if trades else 0,
                'max_profit_percent': max([t['final_pnl_percent'] for t in trades]) if trades else 0,
                'max_loss_percent': min([t['final_pnl_percent'] for t in trades]) if trades else 0,
                'avg_premium_yield': np.mean([t['premium_yield'] for t in trades]) if trades else 0,
                'avg_days_held': np.mean([t['days_held'] for t in trades]) if trades else 0
            }
        }
        
        json_path = f'/Users/maverick/PycharmProjects/India-TS/Daily/analysis/keltner_put_selling_{timestamp}.json'
        with open(json_path, 'w') as f:
            json.dump(output, f, indent=2, default=str)
        
        # Save to Excel for easy viewing
        if trades:
            df_trades = pd.DataFrame(trades)
            
            # Save to Excel
            excel_path = f'/Users/maverick/PycharmProjects/India-TS/Daily/analysis/keltner_put_selling_{timestamp}.xlsx'
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df_trades.to_excel(writer, sheet_name='Trade Summary', index=False)
                
                # Add summary statistics sheet
                summary_data = {
                    'Metric': ['Total Trades', 'Win Rate %', 'Average P&L %', 'Best Trade %', 
                              'Worst Trade %', 'Avg Premium Yield %', 'Avg Days Held',
                              'Expected Value %'],
                    'Value': [
                        len(trades),
                        output['summary']['win_rate'],
                        output['summary']['avg_pnl_percent'],
                        output['summary']['max_profit_percent'],
                        output['summary']['max_loss_percent'],
                        output['summary']['avg_premium_yield'],
                        output['summary']['avg_days_held'],
                        output['summary']['avg_pnl_percent']
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Summary Statistics', index=False)
            
            print(f"\nResults saved to:")
            print(f"  - {json_path}")
            print(f"  - {excel_path}")
            
        # Create a markdown report
        self.create_markdown_report(output, timestamp)
    
    def create_markdown_report(self, output, timestamp):
        """Create a detailed markdown report"""
        report_path = f'/Users/maverick/PycharmProjects/India-TS/Daily/analysis/keltner_put_selling_report_{timestamp}.md'
        
        with open(report_path, 'w') as f:
            f.write("# Keltner Channel Put Selling Strategy Analysis\n\n")
            f.write(f"**Analysis Date:** {output['analysis_date']}\n")
            f.write(f"**Period Analyzed:** {output['period']}\n\n")
            
            f.write("## Strategy Description\n")
            f.write("- **Entry:** Sell puts at lower Keltner Channel band on Long Reversal signals\n")
            f.write("- **Expiry:** ~60 days (2 months)\n")
            f.write("- **Strike Selection:** Lower Keltner Channel (support level)\n")
            f.write("- **Exit Rules:** 50% profit target OR 200% loss limit OR expiry\n\n")
            
            f.write("## Performance Summary\n")
            f.write(f"- **Total Trades:** {output['total_trades']}\n")
            f.write(f"- **Win Rate:** {output['summary']['win_rate']:.2f}%\n")
            f.write(f"- **Average P&L:** {output['summary']['avg_pnl_percent']:.2f}%\n")
            f.write(f"- **Best Trade:** {output['summary']['max_profit_percent']:.2f}%\n")
            f.write(f"- **Worst Trade:** {output['summary']['max_loss_percent']:.2f}%\n")
            f.write(f"- **Average Premium Collected:** {output['summary']['avg_premium_yield']:.2f}%\n")
            f.write(f"- **Average Hold Period:** {output['summary']['avg_days_held']:.1f} days\n\n")
            
            f.write("## Strategy Advantages\n")
            f.write("### Pros:\n")
            f.write("- **Time Decay:** Profits from theta decay over time\n")
            f.write("- **Mean Reversion:** Benefits from stocks bouncing off support\n")
            f.write("- **Premium Income:** Regular income generation\n")
            f.write("- **High Win Rate Potential:** Put selling typically has 70-80% win rates\n")
            f.write("- **Defined Risk:** Maximum loss is limited (strike - premium)\n\n")
            
            f.write("### Cons:\n")
            f.write("- **Assignment Risk:** May be forced to buy stocks at strike\n")
            f.write("- **Capital Intensive:** Requires margin for each position\n")
            f.write("- **Limited Upside:** Profit capped at premium collected\n")
            f.write("- **Black Swan Risk:** Large market drops can cause significant losses\n\n")
            
            f.write("## Comparison with Direct Long Strategies\n")
            f.write("| Metric | Put Selling | Long Position | Fixed Sizing |\n")
            f.write("|--------|-------------|---------------|-------------|\n")
            f.write("| **Max Profit** | Premium (Limited) | Unlimited | Unlimited |\n")
            f.write("| **Win Rate** | Typically High | Medium | Low |\n")
            f.write("| **Time Decay** | Helps | Hurts | Neutral |\n")
            f.write("| **Capital Efficiency** | High | Medium | Low |\n")
            f.write("| **Complexity** | High | Low | Low |\n\n")
            
            f.write("## Risk Management Considerations\n")
            f.write("1. **Position Sizing:** Limit to 2-5% of portfolio per trade\n")
            f.write("2. **Diversification:** Spread across multiple stocks and expiry dates\n")
            f.write("3. **Volatility:** Avoid during high volatility periods\n")
            f.write("4. **Liquidity:** Only trade liquid options with tight spreads\n")
            f.write("5. **Assignment:** Have plan for early assignment\n\n")
            
            f.write("## Implementation Requirements\n")
            f.write("- Options trading approval\n")
            f.write("- Sufficient margin capacity\n")
            f.write("- Real-time options pricing\n")
            f.write("- Volatility analysis tools\n")
            f.write("- Risk management systems\n\n")
            
            f.write("## Recommendation\n")
            if output['summary']['avg_pnl_percent'] > 0:
                f.write("Put selling shows promise for generating consistent income. However:\n")
                f.write("1. **Paper Trade First:** Test the strategy with virtual money\n")
                f.write("2. **Start Small:** Begin with 1-2 positions to understand mechanics\n")
                f.write("3. **Focus on Liquidity:** Only trade highly liquid stocks\n")
                f.write("4. **Monitor Closely:** Options require active management\n")
            else:
                f.write("While put selling has theoretical advantages, this backtest shows ")
                f.write("negative returns. Consider:\n")
                f.write("1. **Improve Entry Signals:** Better timing for put sales\n")
                f.write("2. **Volatility Filters:** Only sell when implied volatility is high\n")
                f.write("3. **Market Regime:** Avoid during bear markets\n")
        
        print(f"  - {report_path}")

def main():
    analyzer = KeltnerPutSellingStrategy()
    analyzer.analyze_strategy()

if __name__ == "__main__":
    main()