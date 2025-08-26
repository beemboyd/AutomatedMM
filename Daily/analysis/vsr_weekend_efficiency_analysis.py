#!/usr/bin/env python3
"""
VSR Weekend Efficiency Analysis
Analyzes VSR alerts from past 2 weeks and compares with current prices
Generates comprehensive weekend analysis report with efficiency metrics
"""

import os
import sys
import json
import pandas as pd
from datetime import datetime, timedelta
import glob
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from user_context_manager import UserContextManager

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class VSRWeekendEfficiencyAnalyzer:
    def __init__(self, lookback_days=14, user='Sai'):
        """Initialize the analyzer with lookback period"""
        self.lookback_days = lookback_days
        self.user = user
        self.base_dir = "/Users/maverick/PycharmProjects/India-TS/Daily"
        self.output_dir = os.path.join(self.base_dir, "analysis", "Efficiency")
        
        # Ensure output directory exists
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Calculate date range
        self.end_date = datetime.now()
        self.start_date = self.end_date - timedelta(days=lookback_days)
        
        # Data storage
        self.vsr_alerts = {}  # Store all VSR alerts
        self.ticker_performance = {}  # Store performance metrics
        
        # Initialize Zerodha connection
        self.init_zerodha()
        
    def init_zerodha(self):
        """Initialize Zerodha connection"""
        try:
            ucm = UserContextManager()
            self.kite, self.config = ucm.get_user_context(self.user)
            logger.info(f"Initialized Zerodha connection for user: {self.user}")
        except Exception as e:
            logger.error(f"Failed to initialize Zerodha: {e}")
            self.kite = None
            
    def get_current_price(self, ticker):
        """Get current price from Zerodha"""
        if not self.kite:
            return None
            
        try:
            # Try NSE first
            instrument = f"NSE:{ticker}"
            quote = self.kite.quote(instrument)
            if instrument in quote:
                return quote[instrument]['last_price']
        except:
            try:
                # Try BSE if NSE fails
                instrument = f"BSE:{ticker}"
                quote = self.kite.quote(instrument)
                if instrument in quote:
                    return quote[instrument]['last_price']
            except:
                pass
        return None
    
    def parse_vsr_scanner_files(self):
        """Parse VSR scanner result files from past 2 weeks"""
        logger.info(f"Parsing VSR scanner files from {self.start_date.strftime('%Y-%m-%d')} to {self.end_date.strftime('%Y-%m-%d')}")
        
        # Directory with VSR scanner files
        scanner_dir = os.path.join(self.base_dir, "scanners", "Hourly")
        
        if not os.path.exists(scanner_dir):
            logger.error(f"Scanner directory not found: {scanner_dir}")
            return
            
        # Get all VSR files from the date range
        for file in sorted(os.listdir(scanner_dir)):
            if not file.startswith('VSR_') or not file.endswith('.xlsx'):
                continue
                
            # Extract date from filename
            try:
                date_str = file.replace('VSR_', '').replace('.xlsx', '')
                file_date = datetime.strptime(date_str[:8], '%Y%m%d')
                
                # Check if within our date range
                if file_date < self.start_date or file_date > self.end_date:
                    continue
                    
                filepath = os.path.join(scanner_dir, file)
                self.parse_vsr_excel(filepath, date_str)
                
            except Exception as e:
                logger.debug(f"Error processing file {file}: {e}")
                continue
    
    def parse_vsr_excel(self, filepath, date_str):
        """Parse individual VSR Excel file"""
        try:
            # Read Excel file
            df = pd.read_excel(filepath)
            
            if df.empty:
                return
                
            # Extract date and time
            alert_date = datetime.strptime(date_str[:8], '%Y%m%d')
            alert_time = date_str[9:11] + ':' + date_str[11:13] if len(date_str) > 8 else '00:00'
            alert_datetime = datetime.strptime(date_str[:8] + '_' + date_str[9:15], '%Y%m%d_%H%M%S') if len(date_str) > 8 else alert_date
            
            # Process each row
            for _, row in df.iterrows():
                ticker = row.get('Ticker', '')
                if not ticker:
                    continue
                    
                # Extract relevant data
                entry_price = row.get('Entry_Price', row.get('Current Price', row.get('Close', 0)))
                stop_loss = row.get('Stop_Loss', 0)
                target1 = row.get('Target1', 0)
                target2 = row.get('Target2', 0)
                vsr_score = row.get('VSR_Score', row.get('Score', 0))
                probability = row.get('Probability_Score', 0)
                pattern = row.get('Pattern', '')
                direction = row.get('Direction', 'LONG')
                
                # Store alert data
                if ticker not in self.vsr_alerts:
                    self.vsr_alerts[ticker] = []
                    
                self.vsr_alerts[ticker].append({
                    'datetime': alert_datetime,
                    'date': alert_date.strftime('%Y-%m-%d'),
                    'time': alert_time,
                    'entry_price': entry_price,
                    'stop_loss': stop_loss,
                    'target1': target1,
                    'target2': target2,
                    'vsr_score': vsr_score,
                    'probability': probability,
                    'pattern': pattern,
                    'direction': direction
                })
                
            logger.debug(f"Parsed {filepath}: Found {len(df)} rows")
            
        except Exception as e:
            logger.warning(f"Error parsing {filepath}: {e}")
    
    def calculate_performance_metrics(self):
        """Calculate performance metrics for each ticker"""
        logger.info("Calculating performance metrics with current prices...")
        
        for ticker, alerts in self.vsr_alerts.items():
            if not alerts:
                continue
                
            # Sort alerts by datetime
            alerts.sort(key=lambda x: x['datetime'])
            
            # Get first and last alert
            first_alert = alerts[0]
            last_alert = alerts[-1]
            
            # Get current price from Zerodha
            current_price = self.get_current_price(ticker)
            
            if not current_price:
                logger.debug(f"Could not get current price for {ticker}")
                current_price = last_alert['entry_price']  # Use last entry price as fallback
            
            # Calculate metrics
            entry_price = first_alert['entry_price']
            if entry_price and entry_price > 0:
                price_change_pct = ((current_price - entry_price) / entry_price) * 100
                
                # Check if stop loss hit
                stop_loss_hit = current_price <= first_alert['stop_loss'] if first_alert['stop_loss'] else False
                
                # Check if targets achieved
                target1_achieved = current_price >= first_alert['target1'] if first_alert['target1'] else False
                target2_achieved = current_price >= first_alert['target2'] if first_alert['target2'] else False
                
                # Calculate efficiency score
                if stop_loss_hit:
                    efficiency_score = 0
                elif target2_achieved:
                    efficiency_score = 100
                elif target1_achieved:
                    efficiency_score = 75
                elif price_change_pct > 0:
                    efficiency_score = 50 + min(price_change_pct * 5, 25)  # 50-75 based on positive movement
                else:
                    efficiency_score = max(0, 50 + price_change_pct * 2)  # 0-50 based on negative movement
                
                # Store performance metrics
                self.ticker_performance[ticker] = {
                    'first_alert_date': first_alert['date'],
                    'first_alert_time': first_alert['time'],
                    'entry_price': entry_price,
                    'current_price': current_price,
                    'price_change': current_price - entry_price,
                    'price_change_pct': price_change_pct,
                    'stop_loss': first_alert['stop_loss'],
                    'target1': first_alert['target1'],
                    'target2': first_alert['target2'],
                    'stop_loss_hit': stop_loss_hit,
                    'target1_achieved': target1_achieved,
                    'target2_achieved': target2_achieved,
                    'alert_count': len(alerts),
                    'max_vsr_score': max(a['vsr_score'] for a in alerts),
                    'avg_vsr_score': sum(a['vsr_score'] for a in alerts) / len(alerts),
                    'max_probability': max(a['probability'] for a in alerts),
                    'direction': first_alert['direction'],
                    'pattern': first_alert['pattern'],
                    'efficiency_score': efficiency_score,
                    'days_since_first_alert': (datetime.now() - first_alert['datetime']).days
                }
            else:
                logger.debug(f"Invalid entry price for {ticker}")
    
    def create_comprehensive_report(self):
        """Create comprehensive Excel report with multiple sheets"""
        if not self.ticker_performance:
            logger.warning("No performance data to report")
            return None
            
        # Generate filename
        date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"VSR_Weekend_Efficiency_Analysis_{date_str}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Performance Summary
        ws1 = wb.active
        ws1.title = "Performance Summary"
        self.create_performance_sheet(ws1)
        
        # Sheet 2: Winners (positive returns)
        ws2 = wb.create_sheet("Winners")
        self.create_winners_sheet(ws2)
        
        # Sheet 3: Losers (negative returns)
        ws3 = wb.create_sheet("Losers")
        self.create_losers_sheet(ws3)
        
        # Sheet 4: Statistics
        ws4 = wb.create_sheet("Statistics")
        self.create_statistics_sheet(ws4)
        
        # Sheet 5: Alert Details
        ws5 = wb.create_sheet("Alert Details")
        self.create_alert_details_sheet(ws5)
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"Created comprehensive report: {filepath}")
        
        return filepath
    
    def create_performance_sheet(self, ws):
        """Create performance summary sheet"""
        # Prepare data
        data = []
        for ticker, perf in self.ticker_performance.items():
            data.append({
                'Ticker': ticker,
                'First Alert': perf['first_alert_date'],
                'Entry Price': perf['entry_price'],
                'Current Price': perf['current_price'],
                'Change %': perf['price_change_pct'],
                'Stop Loss': perf['stop_loss'],
                'Target 1': perf['target1'],
                'Target 2': perf['target2'],
                'SL Hit': 'Yes' if perf['stop_loss_hit'] else 'No',
                'T1 Hit': 'Yes' if perf['target1_achieved'] else 'No',
                'T2 Hit': 'Yes' if perf['target2_achieved'] else 'No',
                'Alerts': perf['alert_count'],
                'Max VSR': perf['max_vsr_score'],
                'Efficiency': perf['efficiency_score'],
                'Days': perf['days_since_first_alert']
            })
        
        # Create DataFrame and sort by efficiency
        df = pd.DataFrame(data)
        df = df.sort_values('Efficiency', ascending=False)
        
        # Add headers with formatting
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data with formatting
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format numbers
                if c_idx in [3, 4, 6, 7, 8]:  # Price columns
                    cell.number_format = '#,##0.00'
                elif c_idx == 5:  # Change %
                    cell.number_format = '+#,##0.00%;-#,##0.00%'
                    value_float = float(value) if value else 0
                    if value_float > 0:
                        cell.font = Font(color="008000")  # Green
                    elif value_float < 0:
                        cell.font = Font(color="FF0000")  # Red
                elif c_idx == 14:  # Efficiency
                    cell.number_format = '#,##0.0'
                    # Color code efficiency
                    if value >= 75:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value >= 50:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_winners_sheet(self, ws):
        """Create winners sheet with positive returns"""
        winners = {k: v for k, v in self.ticker_performance.items() if v['price_change_pct'] > 0}
        
        if not winners:
            ws.append(["No winners found"])
            return
            
        # Sort by return percentage
        winners = dict(sorted(winners.items(), key=lambda x: x[1]['price_change_pct'], reverse=True))
        
        # Create table
        headers = ['Rank', 'Ticker', 'Entry', 'Current', 'Gain %', 'T1 Hit', 'T2 Hit', 'Pattern', 'Days']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for rank, (ticker, perf) in enumerate(winners.items(), 1):
            ws.append([
                rank,
                ticker,
                perf['entry_price'],
                perf['current_price'],
                perf['price_change_pct'],
                'Yes' if perf['target1_achieved'] else 'No',
                'Yes' if perf['target2_achieved'] else 'No',
                perf['pattern'],
                perf['days_since_first_alert']
            ])
            
            # Format the row
            row = ws.max_row
            ws.cell(row, 3).number_format = '#,##0.00'
            ws.cell(row, 4).number_format = '#,##0.00'
            ws.cell(row, 5).number_format = '+#,##0.00%'
            ws.cell(row, 5).font = Font(color="008000")
    
    def create_losers_sheet(self, ws):
        """Create losers sheet with negative returns"""
        losers = {k: v for k, v in self.ticker_performance.items() if v['price_change_pct'] < 0}
        
        if not losers:
            ws.append(["No losers found"])
            return
            
        # Sort by return percentage (most negative first)
        losers = dict(sorted(losers.items(), key=lambda x: x[1]['price_change_pct']))
        
        # Create table
        headers = ['Rank', 'Ticker', 'Entry', 'Current', 'Loss %', 'SL Hit', 'Pattern', 'Days']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for rank, (ticker, perf) in enumerate(losers.items(), 1):
            ws.append([
                rank,
                ticker,
                perf['entry_price'],
                perf['current_price'],
                perf['price_change_pct'],
                'Yes' if perf['stop_loss_hit'] else 'No',
                perf['pattern'],
                perf['days_since_first_alert']
            ])
            
            # Format the row
            row = ws.max_row
            ws.cell(row, 3).number_format = '#,##0.00'
            ws.cell(row, 4).number_format = '#,##0.00'
            ws.cell(row, 5).number_format = '-#,##0.00%'
            ws.cell(row, 5).font = Font(color="FF0000")
    
    def create_statistics_sheet(self, ws):
        """Create statistics sheet"""
        # Calculate statistics
        total_tickers = len(self.ticker_performance)
        winners = [p for p in self.ticker_performance.values() if p['price_change_pct'] > 0]
        losers = [p for p in self.ticker_performance.values() if p['price_change_pct'] < 0]
        
        win_rate = (len(winners) / total_tickers * 100) if total_tickers > 0 else 0
        
        avg_gain = sum(w['price_change_pct'] for w in winners) / len(winners) if winners else 0
        avg_loss = sum(l['price_change_pct'] for l in losers) / len(losers) if losers else 0
        
        target1_hit_count = sum(1 for p in self.ticker_performance.values() if p['target1_achieved'])
        target2_hit_count = sum(1 for p in self.ticker_performance.values() if p['target2_achieved'])
        sl_hit_count = sum(1 for p in self.ticker_performance.values() if p['stop_loss_hit'])
        
        # Create statistics table
        stats = [
            ['OVERALL STATISTICS', ''],
            ['', ''],
            ['Analysis Period', f"{self.start_date.strftime('%Y-%m-%d')} to {self.end_date.strftime('%Y-%m-%d')}"],
            ['Total Tickers Analyzed', total_tickers],
            ['Total Alerts', sum(p['alert_count'] for p in self.ticker_performance.values())],
            ['', ''],
            ['PERFORMANCE METRICS', ''],
            ['Winners', len(winners)],
            ['Losers', len(losers)],
            ['Win Rate', f"{win_rate:.2f}%"],
            ['Average Gain', f"{avg_gain:.2f}%"],
            ['Average Loss', f"{avg_loss:.2f}%"],
            ['', ''],
            ['TARGET ACHIEVEMENT', ''],
            ['Target 1 Hit', f"{target1_hit_count} ({target1_hit_count/total_tickers*100:.1f}%)"],
            ['Target 2 Hit', f"{target2_hit_count} ({target2_hit_count/total_tickers*100:.1f}%)"],
            ['Stop Loss Hit', f"{sl_hit_count} ({sl_hit_count/total_tickers*100:.1f}%)"],
            ['', ''],
            ['EFFICIENCY DISTRIBUTION', ''],
            ['High Efficiency (75-100)', sum(1 for p in self.ticker_performance.values() if p['efficiency_score'] >= 75)],
            ['Medium Efficiency (50-75)', sum(1 for p in self.ticker_performance.values() if 50 <= p['efficiency_score'] < 75)],
            ['Low Efficiency (0-50)', sum(1 for p in self.ticker_performance.values() if p['efficiency_score'] < 50)],
            ['', ''],
            ['TOP PATTERNS', ''],
        ]
        
        # Add pattern statistics
        pattern_counts = {}
        for perf in self.ticker_performance.values():
            pattern = perf['pattern']
            if pattern:
                if pattern not in pattern_counts:
                    pattern_counts[pattern] = {'count': 0, 'wins': 0}
                pattern_counts[pattern]['count'] += 1
                if perf['price_change_pct'] > 0:
                    pattern_counts[pattern]['wins'] += 1
        
        # Sort patterns by count
        sorted_patterns = sorted(pattern_counts.items(), key=lambda x: x[1]['count'], reverse=True)[:5]
        for pattern, data in sorted_patterns:
            win_rate = (data['wins'] / data['count'] * 100) if data['count'] > 0 else 0
            stats.append([pattern, f"{data['count']} signals, {win_rate:.1f}% win rate"])
        
        # Write to sheet
        for row_data in stats:
            ws.append(row_data)
            row = ws.max_row
            if row_data[0] in ['OVERALL STATISTICS', 'PERFORMANCE METRICS', 'TARGET ACHIEVEMENT', 'EFFICIENCY DISTRIBUTION', 'TOP PATTERNS']:
                ws.cell(row, 1).font = Font(bold=True, size=12)
                ws.cell(row, 1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
    
    def create_alert_details_sheet(self, ws):
        """Create detailed alert history sheet"""
        # Headers
        headers = ['Ticker', 'Date', 'Time', 'Entry Price', 'VSR Score', 'Probability', 'Pattern', 'Direction']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add all alerts
        row_num = 2
        for ticker, alerts in self.vsr_alerts.items():
            for alert in alerts:
                ws.append([
                    ticker,
                    alert['date'],
                    alert['time'],
                    alert['entry_price'],
                    alert['vsr_score'],
                    alert['probability'],
                    alert['pattern'],
                    alert['direction']
                ])
                
                # Format the row
                ws.cell(row_num, 4).number_format = '#,##0.00'
                ws.cell(row_num, 5).number_format = '#,##0.0'
                ws.cell(row_num, 6).number_format = '#,##0.00'
                
                row_num += 1
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def print_summary(self):
        """Print summary to console"""
        print("\n" + "="*80)
        print("VSR WEEKEND EFFICIENCY ANALYSIS")
        print("="*80)
        print(f"Analysis Period: {self.start_date.strftime('%Y-%m-%d')} to {self.end_date.strftime('%Y-%m-%d')}")
        print(f"Total Tickers Analyzed: {len(self.ticker_performance)}")
        print(f"Total Alerts Processed: {sum(len(alerts) for alerts in self.vsr_alerts.values())}")
        
        if self.ticker_performance:
            winners = [p for p in self.ticker_performance.values() if p['price_change_pct'] > 0]
            losers = [p for p in self.ticker_performance.values() if p['price_change_pct'] < 0]
            
            print(f"\nPerformance Summary:")
            print(f"  Winners: {len(winners)} ({len(winners)/len(self.ticker_performance)*100:.1f}%)")
            print(f"  Losers: {len(losers)} ({len(losers)/len(self.ticker_performance)*100:.1f}%)")
            
            if winners:
                avg_gain = sum(w['price_change_pct'] for w in winners) / len(winners)
                print(f"  Average Gain: {avg_gain:.2f}%")
                
                # Top 5 winners
                top_winners = sorted(self.ticker_performance.items(), 
                                   key=lambda x: x[1]['price_change_pct'], 
                                   reverse=True)[:5]
                print(f"\n  Top 5 Winners:")
                for ticker, perf in top_winners:
                    print(f"    {ticker}: +{perf['price_change_pct']:.2f}% (Entry: {perf['entry_price']:.2f}, Current: {perf['current_price']:.2f})")
            
            if losers:
                avg_loss = sum(l['price_change_pct'] for l in losers) / len(losers)
                print(f"  Average Loss: {avg_loss:.2f}%")
                
                # Top 5 losers
                top_losers = sorted(self.ticker_performance.items(), 
                                  key=lambda x: x[1]['price_change_pct'])[:5]
                print(f"\n  Top 5 Losers:")
                for ticker, perf in top_losers:
                    print(f"    {ticker}: {perf['price_change_pct']:.2f}% (Entry: {perf['entry_price']:.2f}, Current: {perf['current_price']:.2f})")
            
            # Target achievement
            target1_hit = sum(1 for p in self.ticker_performance.values() if p['target1_achieved'])
            target2_hit = sum(1 for p in self.ticker_performance.values() if p['target2_achieved'])
            sl_hit = sum(1 for p in self.ticker_performance.values() if p['stop_loss_hit'])
            
            print(f"\nTarget Achievement:")
            print(f"  Target 1 Hit: {target1_hit} ({target1_hit/len(self.ticker_performance)*100:.1f}%)")
            print(f"  Target 2 Hit: {target2_hit} ({target2_hit/len(self.ticker_performance)*100:.1f}%)")
            print(f"  Stop Loss Hit: {sl_hit} ({sl_hit/len(self.ticker_performance)*100:.1f}%)")
        
        print("="*80)
    
    def run_analysis(self):
        """Main method to run the complete analysis"""
        logger.info("Starting VSR Weekend Efficiency Analysis")
        
        # Parse VSR scanner files
        self.parse_vsr_scanner_files()
        
        if not self.vsr_alerts:
            logger.warning("No VSR alerts found for the analysis period")
            return None
            
        logger.info(f"Found {len(self.vsr_alerts)} unique tickers with alerts")
        
        # Calculate performance metrics
        self.calculate_performance_metrics()
        
        # Create comprehensive report
        report_path = self.create_comprehensive_report()
        
        # Print summary
        self.print_summary()
        
        # Update Activity.md
        self.update_activity_log(report_path)
        
        return report_path
    
    def update_activity_log(self, report_path):
        """Update Activity.md with analysis results"""
        try:
            activity_file = "/Users/maverick/PycharmProjects/India-TS/Daily/Activity.md"
            
            # Prepare summary statistics
            total_tickers = len(self.ticker_performance)
            winners = [p for p in self.ticker_performance.values() if p['price_change_pct'] > 0]
            losers = [p for p in self.ticker_performance.values() if p['price_change_pct'] < 0]
            win_rate = (len(winners) / total_tickers * 100) if total_tickers > 0 else 0
            
            # Create log entry
            log_entry = f"""
### {datetime.now().strftime('%Y-%m-%d %H:%M')} IST - [Claude]
**Changes:**
- Ran VSR Weekend Efficiency Analysis for past 2 weeks
- Analyzed {total_tickers} unique tickers from VSR scanner alerts
- Compared entry prices with current Zerodha prices
- Generated comprehensive efficiency report with 5 analysis sheets

**Impact:**
- Performance Summary: {len(winners)} winners ({win_rate:.1f}% win rate), {len(losers)} losers
- Average gain: {(sum(w['price_change_pct'] for w in winners) / len(winners)):.2f}% (for winners)
- Average loss: {(sum(l['price_change_pct'] for l in losers) / len(losers)):.2f}% (for losers)
- Target 1 achievement: {sum(1 for p in self.ticker_performance.values() if p['target1_achieved'])} tickers
- Target 2 achievement: {sum(1 for p in self.ticker_performance.values() if p['target2_achieved'])} tickers
- Stop loss hit: {sum(1 for p in self.ticker_performance.values() if p['stop_loss_hit'])} tickers
- Report saved: {report_path}

**Key Findings:**
- Most efficient patterns identified and documented
- Tickers with highest efficiency scores tracked
- Complete alert history maintained for audit trail

---
"""
            
            # Read existing content
            with open(activity_file, 'r') as f:
                content = f.read()
            
            # Find the position to insert (after the header section)
            insert_pos = content.find("## Activity Log\n")
            if insert_pos != -1:
                insert_pos = content.find("\n", insert_pos) + 1
                new_content = content[:insert_pos] + log_entry + content[insert_pos:]
                
                # Write back
                with open(activity_file, 'w') as f:
                    f.write(new_content)
                    
                logger.info("Updated Activity.md with analysis results")
                
        except Exception as e:
            logger.warning(f"Could not update Activity.md: {e}")


def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='VSR Weekend Efficiency Analysis')
    parser.add_argument('--days', type=int, default=14, help='Number of days to analyze (default: 14)')
    parser.add_argument('--user', type=str, default='Sai', help='User for Zerodha connection (default: Sai)')
    
    args = parser.parse_args()
    
    analyzer = VSRWeekendEfficiencyAnalyzer(lookback_days=args.days, user=args.user)
    report_path = analyzer.run_analysis()
    
    if report_path:
        print(f"\nAnalysis completed successfully!")
        print(f"Report saved to: {report_path}")
    else:
        print("\nNo data found for the analysis period")


if __name__ == "__main__":
    main()