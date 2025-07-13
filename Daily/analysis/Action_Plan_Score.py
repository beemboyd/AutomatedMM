#!/usr/bin/env python3
"""
Consolidated Score Analyzer
==========================
Analyzes the past 4 StrategyB report sheets from Daily/results folder
and creates a consolidated score report.

Output: Consolidated_Score_{date}_{time}.xlsx
Columns: Rank, Ticker, Score, SL, TP

Usage:
    python consolidated_score_analyzer.py
    python consolidated_score_analyzer.py --sheets 6  # Analyze past 6 sheets instead of 4
"""

import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime
import logging
from pathlib import Path
import glob
from collections import defaultdict

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ConsolidatedScoreAnalyzer:
    """Analyze past N sheets and create consolidated score report"""
    
    def __init__(self, num_sheets=4):
        """Initialize the analyzer"""
        self.num_sheets = num_sheets
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.daily_dir = os.path.dirname(self.script_dir)
        self.results_dir = os.path.join(self.daily_dir, "results")
        self.output_dir = os.path.join(self.daily_dir, "Plan")
        
        # Create output directory if it doesn't exist
        os.makedirs(self.output_dir, exist_ok=True)
        
        logger.info(f"Initialized Consolidated Score Analyzer for {num_sheets} sheets")
    
    def get_latest_reports(self):
        """Get the latest N StrategyB report files"""
        # Find all StrategyB report files
        pattern = os.path.join(self.results_dir, "StrategyB_Report_*.xlsx")
        files = glob.glob(pattern)
        
        if not files:
            logger.error("No StrategyB report files found")
            return []
        
        # Sort files by modification time (most recent first)
        files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        
        # Take only the requested number of files
        latest_files = files[:self.num_sheets]
        
        logger.info(f"Found {len(files)} total files, using latest {len(latest_files)}")
        for file in latest_files:
            logger.info(f"  - {os.path.basename(file)}")
        
        return latest_files
    
    def parse_score(self, score_str):
        """Parse score from format like '7/7' to numeric value"""
        if isinstance(score_str, str) and '/' in score_str:
            numerator = int(score_str.split('/')[0])
            return numerator
        else:
            try:
                return float(score_str)
            except:
                return 0
    
    def load_and_consolidate_data(self, files):
        """Load data from files and consolidate"""
        all_data = []
        
        for file in files:
            try:
                df = pd.read_excel(file)
                
                # Extract date from filename for tracking
                filename = os.path.basename(file)
                date_str = filename.split('_')[2]
                time_str = filename.split('_')[3].replace('.xlsx', '')
                
                # Add source info
                df['source_file'] = filename
                df['scan_date'] = date_str
                df['scan_time'] = time_str
                
                all_data.append(df)
                logger.info(f"Loaded {len(df)} entries from {filename}")
                
            except Exception as e:
                logger.error(f"Error loading {file}: {e}")
        
        if not all_data:
            logger.error("No data loaded from files")
            return pd.DataFrame()
        
        # Combine all data
        combined_df = pd.concat(all_data, ignore_index=True)
        logger.info(f"Total entries loaded: {len(combined_df)}")
        
        return combined_df
    
    def calculate_consolidated_scores(self, df):
        """Calculate consolidated scores for each ticker"""
        ticker_stats = {}
        
        # Group by ticker
        for ticker in df['Ticker'].unique():
            ticker_df = df[df['Ticker'] == ticker]
            
            # Parse scores
            scores = [self.parse_score(score) for score in ticker_df['Score']]
            
            # Calculate total score (sum of all appearances)
            total_score = sum(scores)
            
            # Get most recent entry for SL and TP
            most_recent = ticker_df.iloc[-1]  # Last entry (most recent based on how we loaded)
            
            # Extract stop loss and target prices
            stop_loss = most_recent.get('Stop_Loss', 0)
            target1 = most_recent.get('Target1', 0)
            target2 = most_recent.get('Target2', 0)
            
            # Use the higher target as TP
            target_price = max(target1, target2) if target1 and target2 else (target1 or target2 or 0)
            
            ticker_stats[ticker] = {
                'Ticker': ticker,
                'Score': total_score,
                'SL': stop_loss,
                'TP': target_price,
                'Frequency': len(ticker_df),
                'Avg_Score': np.mean(scores) if scores else 0,
                'Max_Score': max(scores) if scores else 0,
                'Last_Entry_Price': most_recent.get('Entry_Price', 0),
                'Pattern': most_recent.get('Pattern', ''),
                'Direction': most_recent.get('Direction', '')
            }
        
        # Convert to DataFrame
        result_df = pd.DataFrame.from_dict(ticker_stats, orient='index')
        
        # Sort by Score (descending)
        result_df = result_df.sort_values('Score', ascending=False)
        
        # Add Rank
        result_df['Rank'] = range(1, len(result_df) + 1)
        
        # Reorder columns
        column_order = ['Rank', 'Ticker', 'Score', 'SL', 'TP']
        extra_columns = ['Frequency', 'Avg_Score', 'Max_Score', 'Last_Entry_Price', 'Pattern', 'Direction']
        
        # Include all columns in the final output
        all_columns = column_order + extra_columns
        result_df = result_df[all_columns]
        
        return result_df
    
    def save_report(self, df):
        """Save the consolidated report to Excel"""
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Consolidated_Score_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # Create Excel writer with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Consolidated_Scores', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Consolidated_Scores']
            
            # Add header formatting
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Format headers
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Format columns
            worksheet.column_dimensions['A'].width = 8   # Rank
            worksheet.column_dimensions['B'].width = 15  # Ticker
            worksheet.column_dimensions['C'].width = 10  # Score
            worksheet.column_dimensions['D'].width = 12  # SL
            worksheet.column_dimensions['E'].width = 12  # TP
            worksheet.column_dimensions['F'].width = 12  # Frequency
            worksheet.column_dimensions['G'].width = 12  # Avg_Score
            worksheet.column_dimensions['H'].width = 12  # Max_Score
            worksheet.column_dimensions['I'].width = 15  # Last_Entry_Price
            worksheet.column_dimensions['J'].width = 30  # Pattern
            worksheet.column_dimensions['K'].width = 12  # Direction
            
            # Format data rows
            for row in range(2, len(df) + 2):
                # Center align rank and score
                worksheet[f'A{row}'].alignment = Alignment(horizontal="center")
                worksheet[f'C{row}'].alignment = Alignment(horizontal="center")
                
                # Format prices
                worksheet[f'D{row}'].number_format = '#,##0.00'  # SL
                worksheet[f'E{row}'].number_format = '#,##0.00'  # TP
                worksheet[f'I{row}'].number_format = '#,##0.00'  # Last Entry Price
                
                # Format averages
                worksheet[f'G{row}'].number_format = '0.00'  # Avg Score
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for col in ['A', 'B', 'C', 'D', 'E']:
                    worksheet[f'{col}{row}'].border = thin_border
            
            # Add summary section
            summary_row = len(df) + 4
            worksheet[f'A{summary_row}'] = 'Summary'
            worksheet[f'A{summary_row}'].font = Font(bold=True, size=12)
            
            worksheet[f'A{summary_row + 1}'] = 'Total Tickers:'
            worksheet[f'B{summary_row + 1}'] = len(df)
            
            worksheet[f'A{summary_row + 2}'] = 'Analysis Period:'
            worksheet[f'B{summary_row + 2}'] = f'Last {self.num_sheets} reports'
            
            worksheet[f'A{summary_row + 3}'] = 'Generated:'
            worksheet[f'B{summary_row + 3}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Add color coding for top performers
            # Green for top 10
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            for row in range(2, min(12, len(df) + 2)):  # Top 10
                for col in ['A', 'B', 'C', 'D', 'E']:
                    worksheet[f'{col}{row}'].fill = green_fill
            
            # Yellow for 11-20
            if len(df) > 10:
                yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                for row in range(12, min(22, len(df) + 2)):  # 11-20
                    for col in ['A', 'B', 'C', 'D', 'E']:
                        worksheet[f'{col}{row}'].fill = yellow_fill
        
        logger.info(f"Report saved to: {filepath}")
        
        # Also save as latest version for easy access
        latest_filepath = os.path.join(self.output_dir, "Consolidated_Score_Latest.xlsx")
        df.to_excel(latest_filepath, index=False)
        logger.info(f"Latest version saved to: {latest_filepath}")
        
        return filepath
    
    def generate_summary_stats(self, df):
        """Generate summary statistics"""
        logger.info("\n" + "="*60)
        logger.info("SUMMARY STATISTICS")
        logger.info("="*60)
        
        logger.info(f"Total unique tickers: {len(df)}")
        logger.info(f"Average score: {df['Score'].mean():.2f}")
        logger.info(f"Highest score: {df['Score'].max()} ({df.iloc[0]['Ticker']})")
        
        # Score distribution
        logger.info("\nScore Distribution:")
        score_bins = [0, 10, 20, 30, 40, 50, 100]
        score_labels = ['0-10', '11-20', '21-30', '31-40', '41-50', '50+']
        df['Score_Bin'] = pd.cut(df['Score'], bins=score_bins, labels=score_labels)
        
        for bin_label in score_labels:
            count = len(df[df['Score_Bin'] == bin_label])
            if count > 0:
                logger.info(f"  {bin_label}: {count} tickers")
        
        # Top 10 by score
        logger.info("\nTop 10 Tickers by Score:")
        logger.info(f"{'Rank':<6} {'Ticker':<12} {'Score':<8} {'Freq':<8} {'Avg':<8}")
        logger.info("-" * 50)
        
        for _, row in df.head(10).iterrows():
            logger.info(
                f"{row['Rank']:<6} {row['Ticker']:<12} {row['Score']:<8} "
                f"{row['Frequency']:<8} {row['Avg_Score']:<8.1f}"
            )
    
    def run(self):
        """Run the complete analysis"""
        logger.info("Starting consolidated score analysis...")
        
        # Get latest report files
        files = self.get_latest_reports()
        if not files:
            logger.error("No files to analyze")
            return None
        
        # Load and consolidate data
        combined_df = self.load_and_consolidate_data(files)
        if combined_df.empty:
            logger.error("No data to analyze")
            return None
        
        # Calculate consolidated scores
        result_df = self.calculate_consolidated_scores(combined_df)
        
        # Generate summary statistics
        self.generate_summary_stats(result_df)
        
        # Save report
        output_file = self.save_report(result_df)
        
        logger.info("\n" + "="*60)
        logger.info("Analysis complete!")
        logger.info(f"Output file: {output_file}")
        
        return output_file

def main():
    """Main function"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate consolidated score report from recent scans')
    parser.add_argument('--sheets', type=int, default=8, help='Number of past sheets to analyze (default: 4)')
    
    args = parser.parse_args()
    
    print(f"\n📊 Consolidated Score Analyzer")
    print("=" * 50)
    print(f"🔍 Analyzing past {args.sheets} StrategyB reports...")
    print(f"📂 Reading from: Daily/results/")
    print(f"💾 Saving to: Daily/Plan/")
    
    # Run analysis
    analyzer = ConsolidatedScoreAnalyzer(num_sheets=args.sheets)
    output_file = analyzer.run()
    
    if output_file:
        print(f"\n✅ Analysis successfully completed!")
        print(f"📊 Output files created:")
        print(f"   - {os.path.basename(output_file)}")
        print(f"   - Consolidated_Score_Latest.xlsx (for easy access)")
        print(f"\n📁 Location: Daily/Plan/")
    else:
        print(f"\n❌ Analysis failed. Check the logs for details.")

if __name__ == "__main__":
    main()